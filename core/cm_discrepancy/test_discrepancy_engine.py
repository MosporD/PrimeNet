"""Engine tests for the CM discrepancy audit (no CM API access required)."""

from __future__ import annotations

import os
import tempfile
import unittest

from openpyxl import load_workbook

from core.cm_discrepancy import store
from core.cm_discrepancy.audit import _audit_one_mo
from core.cm_discrepancy.common_settings import audit_mo_records, common_value, distribution_text
from core.cm_discrepancy.excel_export import export_run_workbook
from core.cm_discrepancy.records import record_key, rows_to_records, sheet_to_records


def _records(rows: dict[str, dict]) -> dict[str, dict]:
    return rows


class CommonSettingsTests(unittest.TestCase):
    def test_common_value_mode_with_tiebreak(self):
        from collections import Counter

        self.assertEqual(common_value(Counter({'a': 3, 'b': 1})), 'a')
        # Tie -> deterministic (higher value text wins the max() tie-break)
        self.assertEqual(common_value(Counter({'a': 2, 'b': 2})), 'b')
        self.assertEqual(common_value(Counter()), '')

    def test_distribution_text(self):
        from collections import Counter

        text = distribution_text(Counter({'20': 3, '40': 1}))
        self.assertEqual(text, '20: 3, 40: 1')

    def test_audit_mo_records_flags_minority_values(self):
        records = _records({
            'DN=MRBTS-1/LNCEL-1': {'DN': 'MRBTS-1/LNCEL-1', 'qRxLevMin': '-128', 'tac': '100'},
            'DN=MRBTS-2/LNCEL-1': {'DN': 'MRBTS-2/LNCEL-1', 'qRxLevMin': '-128', 'tac': '100'},
            'DN=MRBTS-3/LNCEL-1': {'DN': 'MRBTS-3/LNCEL-1', 'qRxLevMin': '-124', 'tac': '100'},
        })
        result = audit_mo_records(records)
        master = {row['parameter']: row for row in result['master']}
        # Identity column is excluded from the audit
        self.assertNotIn('DN', master)
        self.assertEqual(master['qRxLevMin']['common_setting'], '-128')
        self.assertEqual(master['qRxLevMin']['mismatch_count'], 1)
        self.assertEqual(master['qRxLevMin']['unique_count'], 2)
        self.assertEqual(master['tac']['mismatch_count'], 0)
        # Summary only lists mismatched parameters
        self.assertEqual([row['parameter'] for row in result['summary']], ['qRxLevMin'])
        self.assertEqual(list(result['mismatched_objects']), ['DN=MRBTS-3/LNCEL-1'])
        deviation = result['mismatched_objects']['DN=MRBTS-3/LNCEL-1'][0]
        self.assertEqual(deviation, {'parameter': 'qRxLevMin', 'value': '-124', 'common': '-128'})

    def test_empty_values_ignored_by_default(self):
        records = _records({
            'k1': {'p': '1'},
            'k2': {'p': ''},
            'k3': {'p': '1'},
        })
        result = audit_mo_records(records)
        master = {row['parameter']: row for row in result['master']}
        self.assertEqual(master['p']['total_samples'], 2)
        self.assertEqual(master['p']['mismatch_count'], 0)
        self.assertEqual(result['mismatched_objects'], {})


class RecordHelpersTests(unittest.TestCase):
    def test_record_key_priority(self):
        self.assertEqual(record_key({'DN': 'MRBTS-1', 'x': '2'}, fallback='f'), 'DN=MRBTS-1')
        self.assertEqual(record_key({'x': ''}, fallback='row-9'), 'row-9')

    def test_sheet_and_rows_to_records(self):
        sheet = {'headers': ['DN', 'p1'], 'rows': [['MRBTS-1/LNCEL-1', '5']]}
        records = sheet_to_records(sheet)
        self.assertIn('DN=MRBTS-1/LNCEL-1', records)
        rows = [{'NE': 'eNB1', 'Local Cell ID': '2', 'p': '7'}]
        records = rows_to_records(rows, ignore_columns={'NE'})
        self.assertIn('Local Cell ID=2', records)
        self.assertNotIn('NE', records['Local Cell ID=2'])


class StoreAndFlagsTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.db_path = os.path.join(self.tmp.name, 'cm_discrepancy.db')
        self.conn = store.connect(self.db_path)

    def tearDown(self):
        self.conn.close()
        self.tmp.cleanup()

    def _run_mo(self, run_id, prev_run_id, records):
        ne_by_key = {key: key.split('/', 1)[0].replace('DN=', '') for key in records}
        return _audit_one_mo(
            self.conn,
            run_id=run_id,
            prev_run_id=prev_run_id,
            mo='NOKLTE:LNCEL',
            records=records,
            ne_by_key=ne_by_key,
            include_empty=False,
            detected_date='2026-07-08',
        )

    def test_added_removed_against_previous_run(self):
        day1 = {
            'DN=MRBTS-1/LNCEL-1': {'DN': 'MRBTS-1/LNCEL-1', 'p': '1'},
            'DN=MRBTS-2/LNCEL-1': {'DN': 'MRBTS-2/LNCEL-1', 'p': '1'},
        }
        run1 = store.create_run(self.conn, vendor='nokia', run_date='2026-07-07')
        stats1 = self._run_mo(run1, None, day1)
        store.finish_run(self.conn, run1, status='success', stats=stats1)
        # First run: no baseline -> no added/removed
        self.assertEqual((stats1['added'], stats1['removed']), (0, 0))

        day2 = {
            'DN=MRBTS-2/LNCEL-1': {'DN': 'MRBTS-2/LNCEL-1', 'p': '1'},
            'DN=MRBTS-3/LNCEL-1': {'DN': 'MRBTS-3/LNCEL-1', 'p': '2'},
        }
        run2 = store.create_run(self.conn, vendor='nokia', run_date='2026-07-08')
        prev = store.previous_successful_run(self.conn, vendor='nokia', before_run_id=run2)
        self.assertEqual(int(prev['id']), run1)
        stats2 = self._run_mo(run2, int(prev['id']), day2)
        store.finish_run(self.conn, run2, status='success', stats=stats2)

        self.assertEqual(stats2['added'], 1)
        self.assertEqual(stats2['removed'], 1)
        detail = store.get_detail(self.conn, run2)
        flags = {(item['object_key'], item['flag']) for item in detail['items']}
        self.assertIn(('DN=MRBTS-3/LNCEL-1', 'added'), flags)
        self.assertIn(('DN=MRBTS-1/LNCEL-1', 'removed'), flags)
        # p=2 is the minority (p=1 wins the mode across day2? both appear once ->
        # tie-break picks '2', so MRBTS-2 becomes the mismatch)
        mismatched = [item for item in detail['items'] if item['flag'] == 'mismatched']
        self.assertEqual(len(mismatched), 1)

    def test_supersede_rerun_same_day(self):
        run1 = store.create_run(self.conn, vendor='huawei', run_date='2026-07-08')
        self._run_mo(run1, None, {'k': {'DN': 'x', 'p': '1'}})
        store.finish_run(self.conn, run1, status='success', stats={})
        store.supersede_runs(self.conn, vendor='huawei', run_date='2026-07-08')
        self.assertIsNone(store.find_run(self.conn, vendor='huawei', run_date='2026-07-08'))
        self.assertEqual(store.get_master(self.conn, run1), [])

    def test_excel_export_sheets(self):
        records = {
            'DN=MRBTS-1/LNCEL-1': {'DN': 'MRBTS-1/LNCEL-1', 'p': '1'},
            'DN=MRBTS-2/LNCEL-1': {'DN': 'MRBTS-2/LNCEL-1', 'p': '2'},
            'DN=MRBTS-3/LNCEL-1': {'DN': 'MRBTS-3/LNCEL-1', 'p': '2'},
        }
        run_id = store.create_run(self.conn, vendor='nokia', run_date='2026-07-08')
        stats = self._run_mo(run_id, None, records)
        store.append_trend(
            self.conn, vendor='nokia', run_date='2026-07-08',
            run_id=run_id, total=stats['mismatches'],
        )
        store.finish_run(self.conn, run_id, status='success', stats=stats)
        out = os.path.join(self.tmp.name, 'Nokia_disc_08_07_2026.xlsx')
        path = export_run_workbook(self.conn, run_id, output_path=out)
        wb = load_workbook(path)
        self.assertEqual(
            wb.sheetnames[:3], ['Summary', 'Master Sheet', 'Accumulated Data']
        )
        self.assertIn('LNCEL', wb.sheetnames)
        summary_rows = list(wb['Summary'].iter_rows(values_only=True))
        self.assertEqual(summary_rows[0], ('MO', 'Parameter', 'No. of Mismatches'))
        self.assertEqual(summary_rows[1], ('NOKLTE:LNCEL', 'p', 1))
        master_rows = list(wb['Master Sheet'].iter_rows(values_only=True))
        self.assertEqual(
            master_rows[0],
            ('MO', 'Parameter', 'Distribution', 'Common Settings', 'Unique count'),
        )
        self.assertEqual(master_rows[1][3], '2')
        trend_rows = list(wb['Accumulated Data'].iter_rows(values_only=True))
        self.assertEqual(trend_rows[1], ('2026-07-08', 1))
        detail_rows = list(wb['LNCEL'].iter_rows(values_only=True))
        self.assertEqual(detail_rows[0][-2], 'Flag')
        self.assertEqual(detail_rows[1][-2], 'mismatched')


if __name__ == '__main__':
    unittest.main()
