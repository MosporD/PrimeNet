"""Excel workbook export mirroring the legacy RAN Multi Vendor CM Auditor files.

Produces ``{Vendor}_disc_DD_MM_YYYY.xlsx`` under
``uploads/cm_discrepancy/DD_MM_YYYY/`` with:

- ``Summary``          — MO, Parameter, No. of Mismatches
- ``Master Sheet``     — MO, Parameter, Distribution, Common Settings, Unique count
- ``Accumulated Data`` — Date, Total Mismatches (daily trend)
- one sheet per MO     — object identity + parameter values + ``Flag`` + ``Date``
"""

from __future__ import annotations

import os
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.cm_discrepancy import store

_HEADER_FILL = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
_FLAG_FILLS = {
    'mismatched': PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid'),
    'added': PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid'),
    'removed': PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid'),
}
_INVALID_SHEET_CHARS = re.compile(r'[\[\]:*?/\\]')
_DETAIL_PAGE_SIZE = 1000
_MAX_DETAIL_ROWS_PER_SHEET = 50000


def _sheet_title(name: str, used: set[str]) -> str:
    title = _INVALID_SHEET_CHARS.sub('_', str(name))[:31] or 'Sheet'
    base = title
    counter = 2
    while title.lower() in used:
        suffix = f'~{counter}'
        title = f'{base[:31 - len(suffix)]}{suffix}'
        counter += 1
    used.add(title.lower())
    return title


def _write_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical='center')
    ws.freeze_panes = 'A2'
    for idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(idx)].width = max(14, min(40, len(str(header)) + 4))


def export_run_workbook(conn, run_id: int, *, output_path: str = '') -> str:
    """Write the discrepancy workbook for a stored run; returns the file path."""
    run = store.get_run(conn, run_id)
    if not run:
        raise ValueError(f'Unknown discrepancy run id: {run_id}')
    vendor = str(run['vendor'])
    run_date = str(run['run_date'])

    if not output_path:
        from core.cm_discrepancy.audit import workbook_path

        output_path = workbook_path(vendor, run_date)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook(write_only=False)
    wb.remove(wb.active)
    used_titles: set[str] = set()

    # Summary
    ws = wb.create_sheet(_sheet_title('Summary', used_titles))
    _write_header(ws, ['MO', 'Parameter', 'No. of Mismatches'])
    for row in store.get_summary(conn, run_id):
        ws.append([row['mo'], row['parameter'], int(row['mismatch_count'])])

    # Master Sheet
    ws = wb.create_sheet(_sheet_title('Master Sheet', used_titles))
    _write_header(ws, ['MO', 'Parameter', 'Distribution', 'Common Settings', 'Unique count'])
    for row in store.get_master(conn, run_id):
        ws.append([
            row['mo'], row['parameter'], row['distribution'],
            row['common_setting'], int(row['unique_count']),
        ])

    # Accumulated Data (daily trend for this vendor)
    ws = wb.create_sheet(_sheet_title('Accumulated Data', used_titles))
    _write_header(ws, ['Date', 'Total Mismatches'])
    for row in store.get_trend(conn, vendor=vendor, limit=365):
        ws.append([row['run_date'], int(row['total_mismatches'])])

    # Per-MO detail sheets (flagged objects only, matching legacy discrepancy files)
    for mo_info in store.list_detail_mos(conn, run_id):
        mo = str(mo_info['mo'])
        headers: list[str] = []
        rows_buffer: list[dict] = []
        page = 1
        while True:
            chunk = store.get_detail(
                conn, run_id, mo=mo, page=page, page_size=_DETAIL_PAGE_SIZE
            )
            items = chunk['items']
            for item in items:
                for column in item['payload'].keys():
                    if column not in headers:
                        headers.append(column)
            rows_buffer.extend(items)
            if page >= chunk['pages'] or len(rows_buffer) >= _MAX_DETAIL_ROWS_PER_SHEET:
                break
            page += 1

        ws = wb.create_sheet(_sheet_title(mo.split(':', 1)[-1], used_titles))
        _write_header(ws, ['Object', 'NE'] + headers + ['Mismatched Parameters', 'Flag', 'Date'])
        flag_col = len(headers) + 4
        for item in rows_buffer[:_MAX_DETAIL_ROWS_PER_SHEET]:
            payload = item['payload']
            mismatch_text = '; '.join(
                f"{m.get('parameter')}={m.get('value')} (common {m.get('common')})"
                for m in (item['mismatches'] or [])
            )
            ws.append(
                [item['object_key'], item['ne_name']]
                + [str(payload.get(col, '')) for col in headers]
                + [mismatch_text, item['flag'], item['detected_date']]
            )
            fill = _FLAG_FILLS.get(str(item['flag']))
            if fill:
                ws.cell(row=ws.max_row, column=flag_col).fill = fill

    if not wb.sheetnames:
        wb.create_sheet('Summary')
    wb.save(output_path)
    return output_path
