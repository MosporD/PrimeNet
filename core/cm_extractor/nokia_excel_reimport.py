"""Nokia CM Excel round-trip preview and execution helpers."""

from __future__ import annotations

import json
import os
import posixpath
import shutil
import uuid
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

from core.cm_extractor.config import build_nokia_operations_client, nokia_export_ssh_settings
from core.cm_extractor.nokia_operations_client import NokiaOperationsError

try:
    import paramiko
except ImportError:  # pragma: no cover - deployment image includes paramiko
    paramiko = None


IDENTITY_COLUMNS = {'DN', 'moId', 'distName'}
INTERNAL_SHEETS = {'INDEX', 'Empty', 'Skipped_NEs'}
CONFIRMATION_PHRASE = 'APPLY NOKIA EXCEL CHANGES'
_DEFAULT_REIMPORT_REMOTE_DIR = '/d/oss/global/var/racops/import'
_PROJECT_ROOT = Path(__file__).resolve().parents[2]


@dataclass(frozen=True)
class ParsedSheet:
    name: str
    headers: list[str]
    hierarchy_count: int
    rows: dict[str, dict[str, Any]]
    row_numbers: dict[str, int]


def _cell_text(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'true' if value else 'false'
    return str(value).strip()


def _safe_username(username: str) -> str:
    out = ''.join(ch if ch.isalnum() or ch in '._-' else '_' for ch in (username or 'unknown').strip())
    return out[:64] or 'unknown'


def _sheet_class(sheet_name: str) -> str:
    token = (sheet_name or '').strip()
    if '_' in token:
        return token.split('_', 1)[-1]
    return token


def _header_row_and_hierarchy(ws) -> tuple[int, int]:
    first = _cell_text(ws.cell(row=1, column=1).value)
    if first.startswith('HIERARCHY_COLS:'):
        try:
            return 2, max(0, int(first.split(':', 1)[1]))
        except ValueError:
            return 2, 0
    return 1, 0


def _row_identity(row: dict[str, Any], hierarchy_headers: list[str]) -> str:
    for col in ('DN', 'moId', 'distName'):
        val = _cell_text(row.get(col))
        if val:
            return val
    parts = ['PLMN-PLMN']
    for col in hierarchy_headers:
        val = _cell_text(row.get(col))
        if val:
            parts.append(f'{col}-{val}')
    return '/'.join(parts) if len(parts) > 1 else ''


def parse_nokia_workbook(path: str | os.PathLike[str]) -> dict[str, ParsedSheet]:
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    parsed: dict[str, ParsedSheet] = {}
    try:
        for ws in wb.worksheets:
            if ws.title in INTERNAL_SHEETS:
                continue
            header_row, hierarchy_count = _header_row_and_hierarchy(ws)
            headers = [
                _cell_text(ws.cell(row=header_row, column=idx).value)
                for idx in range(1, ws.max_column + 1)
            ]
            while headers and not headers[-1]:
                headers.pop()
            if not headers:
                continue
            hierarchy_headers = headers[:hierarchy_count]
            rows: dict[str, dict[str, Any]] = {}
            row_numbers: dict[str, int] = {}
            for row_idx in range(header_row + 1, ws.max_row + 1):
                row = {
                    header: _cell_text(ws.cell(row=row_idx, column=col_idx).value)
                    for col_idx, header in enumerate(headers, start=1)
                    if header
                }
                if not any(row.values()):
                    continue
                identity = _row_identity(row, hierarchy_headers)
                if not identity:
                    continue
                rows[identity] = row
                row_numbers[identity] = row_idx
            parsed[ws.title] = ParsedSheet(
                name=ws.title,
                headers=headers,
                hierarchy_count=hierarchy_count,
                rows=rows,
                row_numbers=row_numbers,
            )
    finally:
        wb.close()
    return parsed


def _is_readonly_column(sheet: ParsedSheet, column: str) -> bool:
    if column in IDENTITY_COLUMNS:
        return True
    try:
        return sheet.headers.index(column) < sheet.hierarchy_count
    except ValueError:
        return True


def compare_nokia_workbooks(
    baseline_path: str | os.PathLike[str],
    edited_path: str | os.PathLike[str],
    *,
    allow_blank: bool = False,
    max_changes: int = 100,
) -> dict[str, Any]:
    baseline = parse_nokia_workbook(baseline_path)
    edited = parse_nokia_workbook(edited_path)
    changes: list[dict[str, Any]] = []
    blocked: list[dict[str, Any]] = []
    warnings: list[str] = []

    for sheet_name, edited_sheet in edited.items():
        base_sheet = baseline.get(sheet_name)
        if not base_sheet:
            warnings.append(f'Sheet {sheet_name} is new and will be ignored.')
            continue
        base_headers = set(base_sheet.headers)
        for identity, edited_row in edited_sheet.rows.items():
            base_row = base_sheet.rows.get(identity)
            if not base_row:
                blocked.append({
                    'sheet': sheet_name,
                    'row': edited_sheet.row_numbers.get(identity),
                    'target': identity,
                    'reason': 'New rows are not supported for Nokia reimport.',
                })
                continue
            for column, new_value in edited_row.items():
                if column not in base_headers:
                    blocked.append({
                        'sheet': sheet_name,
                        'row': edited_sheet.row_numbers.get(identity),
                        'target': identity,
                        'parameter': column,
                        'reason': 'Unknown/new columns are not supported.',
                    })
                    continue
                old_value = _cell_text(base_row.get(column))
                new_value = _cell_text(new_value)
                if old_value == new_value:
                    continue
                if _is_readonly_column(base_sheet, column):
                    blocked.append({
                        'sheet': sheet_name,
                        'row': edited_sheet.row_numbers.get(identity),
                        'target': identity,
                        'parameter': column,
                        'old_value': old_value,
                        'new_value': new_value,
                        'reason': 'Identity and hierarchy columns cannot be changed.',
                    })
                    continue
                if column.startswith('Item-') or old_value == 'List' or new_value == 'List':
                    blocked.append({
                        'sheet': sheet_name,
                        'row': edited_sheet.row_numbers.get(identity),
                        'target': identity,
                        'parameter': column,
                        'old_value': old_value,
                        'new_value': new_value,
                        'reason': 'List parameters are not supported in the first Nokia reimport pass.',
                    })
                    continue
                if not new_value and old_value and not allow_blank:
                    blocked.append({
                        'sheet': sheet_name,
                        'row': edited_sheet.row_numbers.get(identity),
                        'target': identity,
                        'parameter': column,
                        'old_value': old_value,
                        'new_value': new_value,
                        'reason': 'Blanking a value requires allow_blank=true.',
                    })
                    continue
                changes.append({
                    'sheet': sheet_name,
                    'mo_class': _sheet_class(sheet_name),
                    'row': edited_sheet.row_numbers.get(identity),
                    'target': identity,
                    'parameter': column,
                    'old_value': old_value,
                    'new_value': new_value,
                })

    for sheet_name, base_sheet in baseline.items():
        edited_sheet = edited.get(sheet_name)
        if not edited_sheet:
            warnings.append(f'Sheet {sheet_name} was removed and will be ignored.')
            continue
        removed = set(base_sheet.rows) - set(edited_sheet.rows)
        if removed:
            warnings.append(f'{len(removed)} row(s) removed from {sheet_name}; row deletion is ignored.')

    over_limit = len(changes) > max_changes
    if over_limit:
        blocked.append({
            'reason': f'{len(changes)} changes detected; max allowed per execution is {max_changes}.',
        })

    return {
        'change_count': len(changes),
        'blocked_count': len(blocked),
        'changes': changes,
        'blocked': blocked,
        'warnings': warnings,
        'executable': bool(changes) and not blocked and not over_limit,
    }


def preview_root(username: str, token: str | None = None) -> Path:
    root = _PROJECT_ROOT / 'uploads' / 'cm_extractor' / 'reimport' / _safe_username(username)
    if token:
        root = root / token
    return root


def reimport_remote_dir() -> str:
    """OMC SFTP folder for CM Operations actualImport uploads (not the export dir)."""
    return (
        os.environ.get('NOKIA_CM_REIMPORT_REMOTE_DIR', '').strip()
        or _DEFAULT_REIMPORT_REMOTE_DIR
    )


def create_preview(
    *,
    username: str,
    baseline_path: str,
    edited_path: str,
    edited_filename: str,
    allow_blank: bool = False,
    max_changes: int = 100,
) -> dict[str, Any]:
    token = uuid.uuid4().hex
    dest = preview_root(username, token)
    dest.mkdir(parents=True, exist_ok=True)
    baseline_copy = dest / 'baseline.xlsx'
    edited_copy = dest / 'edited.xlsx'
    shutil.copy2(baseline_path, baseline_copy)
    shutil.copy2(edited_path, edited_copy)
    diff = compare_nokia_workbooks(
        baseline_copy,
        edited_copy,
        allow_blank=allow_blank,
        max_changes=max_changes,
    )
    payload = {
        'token': token,
        'username': username,
        'edited_filename': edited_filename,
        'created_at': datetime.now().isoformat(timespec='seconds'),
        'allow_blank': bool(allow_blank),
        'max_changes': int(max_changes),
        'baseline_path': str(baseline_copy),
        'edited_path': str(edited_copy),
        **diff,
    }
    (dest / 'preview.json').write_text(json.dumps(payload, indent=2), encoding='utf-8')
    return payload


def load_preview(username: str, token: str) -> dict[str, Any]:
    path = preview_root(username, token) / 'preview.json'
    if not path.is_file():
        raise FileNotFoundError('Preview token not found or expired.')
    return json.loads(path.read_text(encoding='utf-8'))


def write_changes_xml(preview: dict[str, Any]) -> str:
    token = preview['token']
    out_path = preview_root(preview['username'], token) / 'changes.xml'
    root = ET.Element('raml', {'version': '2.0', 'xmlns': 'raml21.xsd'})
    cm_data = ET.SubElement(root, 'cmData', {
        'type': 'plan',
        'name': f'PrimeNet_Reimport_{token[:8]}',
        'version': os.environ.get('NOKIA_CM_REIMPORT_RAML_VERSION', 'xL21A_2012_003'),
    })
    header = ET.SubElement(cm_data, 'header')
    ET.SubElement(header, 'log', {
        'dateTime': datetime.now().strftime('%Y-%m-%dT%H:%M:%S'),
        'action': 'created',
        'appInfo': 'PrimeNet Nokia Excel Reimport',
    })
    grouped: dict[tuple[str, str], dict[str, str]] = {}
    for change in preview.get('changes') or []:
        key = (str(change['mo_class']), str(change['target']))
        grouped.setdefault(key, {})[str(change['parameter'])] = str(change['new_value'])
    for (mo_class, target), params in sorted(grouped.items()):
        mo_elem = ET.SubElement(cm_data, 'managedObject', {
            'class': mo_class,
            'distName': target,
            'operation': os.environ.get('NOKIA_CM_REIMPORT_MO_OPERATION', 'update'),
        })
        for name, value in sorted(params.items()):
            ET.SubElement(mo_elem, 'p', {'name': name}).text = value
    tree = ET.ElementTree(root)
    ET.indent(tree, space='  ')
    tree.write(out_path, encoding='utf-8', xml_declaration=True)
    return str(out_path)


def _upload_to_omc(local_path: str, remote_name: str) -> str:
    if paramiko is None:
        raise RuntimeError('paramiko is required for Nokia CM reimport SFTP upload.')
    cfg = nokia_export_ssh_settings()
    if not cfg.get('configured'):
        raise RuntimeError('Nokia CM SFTP is not configured. Set NOKIA_CM_SSH_* or NOKIA_PM_* in .env.')
    remote_dir = reimport_remote_dir()
    remote_path = posixpath.join(remote_dir, remote_name)
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    try:
        ssh.connect(
            hostname=cfg['host'],
            port=int(cfg.get('port') or 22),
            username=cfg['username'],
            password=cfg['password'],
            timeout=int(cfg.get('timeout') or 60),
        )
        sftp = ssh.open_sftp()
        try:
            try:
                sftp.chdir(remote_dir)
            except OSError as exc:
                try:
                    sftp.mkdir(remote_dir)
                    sftp.chdir(remote_dir)
                except OSError as mkdir_exc:
                    raise RuntimeError(
                        f'Cannot access Nokia CM import folder {remote_dir!r} on '
                        f'{cfg["host"]} as {cfg["username"]!r}. Set '
                        f'NOKIA_CM_REIMPORT_REMOTE_DIR in .env to a writable import path. '
                        f'({mkdir_exc or exc})'
                    ) from mkdir_exc
            try:
                sftp.put(local_path, remote_path)
            except (OSError, PermissionError) as exc:
                raise RuntimeError(
                    f'SFTP upload denied for {remote_path!r} on {cfg["host"]} '
                    f'as {cfg["username"]!r}. The CM import folder must be writable by '
                    f'the SFTP user (not the CM REST export folder). '
                    f'Configure NOKIA_CM_REIMPORT_REMOTE_DIR if needed. ({exc})'
                ) from exc
            return remote_path
        finally:
            sftp.close()
    finally:
        ssh.close()


def execute_preview(username: str, token: str, *, wait: bool = False) -> dict[str, Any]:
    preview = load_preview(username, token)
    if not preview.get('executable'):
        raise ValueError('Preview has blocked items or no executable changes.')
    xml_path = write_changes_xml(preview)
    remote_name = f'primenet_reimport_{token[:12]}.xml'
    remote_path = _upload_to_omc(xml_path, remote_name)
    operation_name = os.environ.get('NOKIA_CM_REIMPORT_OPERATION_NAME', 'Import_Export')
    attributes = {
        'importExportOperation': os.environ.get('NOKIA_CM_REIMPORT_OPERATION_MODE', 'actualImport'),
        'fileFormat': os.environ.get('NOKIA_CM_REIMPORT_FILE_FORMAT', 'RAML2'),
        'fileName': remote_name,
        'inputFile': remote_path,
        'DN': os.environ.get('NOKIA_CM_REIMPORT_DN', 'PLMN-PLMN'),
        'useQualifiedClassAbbreviation': os.environ.get('NOKIA_CM_REIMPORT_QUALIFIED_CLASS', 'true'),
    }
    client = build_nokia_operations_client()
    operation_id = client.start_operation(
        operation_name,
        operation_alias=f'PrimeNet Excel reimport {token[:8]}',
        attributes={k: str(v) for k, v in attributes.items() if str(v).strip()},
    )
    result: dict[str, Any] = {
        'operation_id': operation_id,
        'operation_name': operation_name,
        'remote_path': remote_path,
        'xml_path': xml_path,
        'change_count': preview.get('change_count', 0),
    }
    if wait:
        status, feedbacks = client.wait_for_operation(operation_id, timeout_sec=900)
        result['status'] = status
        result['feedbacks'] = feedbacks[:50]
    return result
