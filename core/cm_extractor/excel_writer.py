"""Write CM extraction results to Excel workbooks."""

from __future__ import annotations

import re
from collections import defaultdict
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from ncm_core import (
    managed_objects_to_ncm_sheet,
    merge_ncm_sheet_parts,
    query_table_to_ncm_sheet,
)


def _mo_class_from_nokia_object(mo: dict[str, Any]) -> str:
    mo_class = mo.get('moClass') or {}
    class_id = mo_class.get('id') or 'Unknown'
    if ':' in class_id:
        return class_id.split(':', 1)[1]
    return class_id


def managed_objects_to_sheet(managed_objects: list[dict[str, Any]]) -> dict[str, Any]:
    """Convert getManagedObjects payload to NCM-layout {headers, rows, hierarchy_col_count}."""
    return managed_objects_to_ncm_sheet(managed_objects)


def merge_sheet_parts(parts: list[dict[str, Any]]) -> dict[str, Any]:
    """Merge sheet parts using NCM hierarchy discovery and column ordering."""
    return merge_ncm_sheet_parts(parts)


def query_rows_to_ncm_sheet(headers: list[str], rows: list[list[Any]]) -> dict[str, Any]:
    """Convert CM query rows to NCM-layout sheet data."""
    return query_table_to_ncm_sheet(headers, rows)


_INVALID_SHEET_TITLE_CHARS = re.compile(r'[\:\\/?*\[\]]')


def sanitize_sheet_title(name: str) -> str:
    """Excel sheet titles cannot contain : \\ / ? * [ ]."""
    cleaned = _INVALID_SHEET_TITLE_CHARS.sub('_', (name or '').strip())
    return (cleaned or 'Sheet')[:31]


def allocate_sheet_title(mo_class_id: str, used: set[str]) -> str:
    """
    Return a unique Excel-safe tab name for one MO class selection.

    Prefer the short abbreviation (LNCEL); add adaptation prefix or a numeric
    suffix when the same abbreviation appears more than once.
    """
    raw = (mo_class_id or '').strip()
    if ':' in raw:
        adaptation, abbreviation = raw.split(':', 1)
    else:
        adaptation, abbreviation = '', raw

    candidates = [
        sanitize_sheet_title(abbreviation),
        sanitize_sheet_title(f'{adaptation}_{abbreviation}'),
        sanitize_sheet_title(raw.replace(':', '_')),
    ]
    seen_bases: set[str] = set()
    for base in candidates:
        if not base or base in seen_bases:
            continue
        seen_bases.add(base)
        if base not in used:
            used.add(base)
            return base

    stem = sanitize_sheet_title(f'{adaptation}_{abbreviation}')[:25] or 'Sheet'
    suffix = 2
    while True:
        title = sanitize_sheet_title(f'{stem}_{suffix}')
        if title not in used:
            used.add(title)
            return title
        suffix += 1


def write_nokia_multi_sheet_excel(path: str, sheets: dict[str, dict[str, Any]]) -> None:
    """Write multiple sheets: {sheet_name: {headers, rows, hierarchy_col_count?}}."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    used_titles: set[str] = set()
    for sheet_name, data in sheets.items():
        base = sanitize_sheet_title(sheet_name)
        title = base
        suffix = 2
        while title in used_titles:
            tail = f'_{suffix}'
            title = f'{base[:31 - len(tail)]}{tail}'
            suffix += 1
        used_titles.add(title)
        ws = wb.create_sheet(title=title)
        _write_ncm_sheet(
            ws,
            data.get('headers') or [],
            data.get('rows') or [],
            hierarchy_col_count=int(data.get('hierarchy_col_count') or 0),
        )

    if not wb.sheetnames:
        ws = wb.create_sheet('Empty')
        ws.cell(row=1, column=1, value='No data')

    wb.save(path)


def write_nokia_query_excel(path: str, expressions: list[str], rows: list[list[Any]]) -> None:
    sheet = query_table_to_ncm_sheet(
        list(expressions) if expressions else [],
        rows,
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'QueryResult'
    _write_ncm_sheet(
        ws,
        sheet.get('headers') or [],
        sheet.get('rows') or [],
        hierarchy_col_count=int(sheet.get('hierarchy_col_count') or 0),
    )
    wb.save(path)


def write_nokia_managed_objects_excel(path: str, managed_objects: list[dict[str, Any]]) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    by_class: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for mo in managed_objects:
        by_class[_mo_class_from_nokia_object(mo)].append(mo)

    for mo_class, objects in sorted(by_class.items()):
        sheet_name = mo_class[:31] or 'MO'
        ws = wb.create_sheet(title=sheet_name)
        sheet = managed_objects_to_ncm_sheet(objects)
        _write_ncm_sheet(
            ws,
            sheet.get('headers') or [],
            sheet.get('rows') or [],
            hierarchy_col_count=int(sheet.get('hierarchy_col_count') or 0),
        )

    if not wb.sheetnames:
        ws = wb.create_sheet('Empty')
        ws.cell(row=1, column=1, value='No data')

    wb.save(path)


def write_huawei_mml_excel(path: str, rows: list[dict[str, Any]], sheet_name: str = 'MML_Result') -> None:
    write_huawei_sheets_excel(path, {sheet_name: rows})


def write_huawei_sheets_excel(path: str, sheets: dict[str, list[dict[str, Any]]]) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _INTERNAL_COLS = {'_mml_warnings', 'NE'}
    _GARBAGE_COL_RE = re.compile(r'^[\d=]+$')
    _MASHED_COL_RE = re.compile(r'^\d+(?:\s+\d+)+$')

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        columns: list[str] = []
        for row in rows:
            for key in row:
                key_str = str(key)
                if key_str in _INTERNAL_COLS:
                    continue
                if _GARBAGE_COL_RE.match(key_str):
                    continue
                if _MASHED_COL_RE.match(key_str.strip()):
                    continue
                if key_str not in columns:
                    columns.append(key_str)
        for col in ('NE',):
            if any(col in row for row in rows) and col not in columns:
                columns.append(col)
        table_rows = [[row.get(c, '') for c in columns] for row in rows]
        _write_simple_sheet(ws, columns, table_rows)

    if not wb.sheetnames:
        ws = wb.create_sheet('Empty')
        ws.cell(row=1, column=1, value='No data')

    wb.save(path)


def _write_ncm_sheet(
    ws,
    headers: list[str],
    rows: list[list[Any]],
    *,
    hierarchy_col_count: int = 0,
) -> None:
    """Write NCM-format sheet: metadata row, headers, data (matches bulk RAML export)."""
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    if headers:
        metadata_cell = ws.cell(row=1, column=1, value=f'HIERARCHY_COLS:{hierarchy_col_count}')
        metadata_cell.font = Font(bold=True, italic=True, color='FF0000')
        header_row = 2
        data_start = 3
    else:
        header_row = 1
        data_start = 2

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, row in enumerate(rows, start=data_start):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18


def _write_simple_sheet(ws, headers: list[str], rows: list[list[Any]]) -> None:
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18
