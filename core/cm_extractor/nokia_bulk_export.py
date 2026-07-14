"""
Bulk RNC/BSC export via CM Operations Import_Export (actualExport) + RAML/XML → Excel.
"""

from __future__ import annotations

import os
import re
import shutil
import socket
import tempfile
import time
import uuid
from typing import Any

import paramiko
from paramiko.ssh_exception import SSHException

from core.cm_extractor.config import nokia_export_ssh_settings, nokia_bulk_export_settings
from core.cm_extractor.nokia_client import NokiaCmClient
from core.cm_extractor.nokia_operations_client import NokiaOperationsClient, NokiaOperationsError
from core.cm_extractor.site_catalog import normalize_scope_level, resolve_bulk_export_dns
from ncm_core import FilterConfig, XMLToExcelConverter

_EXPORT_OBJECT_RE = re.compile(r'exported objects:\s*(\d+)', re.I)
_EXPORT_OK_RE = re.compile(r'exported\s+(\d+)\s+managed objects', re.I)
_TRANSIENT_WINERRORS = frozenset({10053, 10054, 10060})
_SFTP_CHUNK_BYTES = 1024 * 1024
_FALLBACK_EXPORT_DIRS = (
    '/var/opt/oss/global/racops/export',
)


class NokiaBulkExportError(Exception):
    pass


def mo_class_id_to_class_filter_include(mo_class_id: str) -> str:
    """
    Convert UI MO class id to NetAct Import_Export ``classFilterInclude`` token.

    NetAct expects ``*:ABBREV`` (e.g. ``*:FMCS``), not ``NOKRNC:FMCS``.
    """
    token = (mo_class_id or '').strip()
    if not token:
        return ''
    if token.startswith('*:'):
        return token
    if ':' in token:
        _qualifier, abbr = token.split(':', 1)
        return f'*:{abbr.strip()}'
    return f'*:{token}'


def normalize_class_filter_include(class_filter: str) -> str:
    """Normalize comma-separated MO class filter for Import_Export."""
    parts: list[str] = []
    seen: set[str] = set()
    for raw in (class_filter or '').split(','):
        token = mo_class_id_to_class_filter_include(raw.strip())
        if token and token not in seen:
            seen.add(token)
            parts.append(token)
    return ','.join(parts)


def selections_to_class_filter_include(selections: list[dict[str, Any]]) -> str:
    """Build NetAct Import_Export ``classFilterInclude`` from UI MO picks."""
    parts: list[str] = []
    seen: set[str] = set()
    for sel in selections or []:
        mo_class_id = (sel.get('mo_class_id') or sel.get('id') or '').strip()
        token = mo_class_id_to_class_filter_include(mo_class_id)
        if not token or token in seen:
            continue
        seen.add(token)
        parts.append(token)
    return ','.join(parts)


def build_param_filter_from_selections(
    selections: list[dict[str, Any]],
) -> FilterConfig | None:
    """
    Excel column filter keyed by RAML MO abbreviation (e.g. ``WCEL``).

    Empty parameter list for a class means full MO export (all parameters).
    """
    filter_dict: dict[str, list[str]] = {}
    for sel in selections or []:
        mo_class_id = (sel.get('mo_class_id') or sel.get('id') or '').strip()
        if ':' not in mo_class_id:
            continue
        abbr = mo_class_id.split(':', 1)[1]
        export_mode = (sel.get('export_mode') or '').strip().lower()
        params = [
            str(param).lstrip('@')
            for param in (sel.get('parameters') or [])
            if str(param).strip()
        ]
        if export_mode == 'full':
            filter_dict[abbr] = []
        elif params:
            filter_dict[abbr] = params
    if not filter_dict:
        return None
    return FilterConfig(filter_dict=filter_dict, all_mos=set(filter_dict.keys()))


def resolve_bulk_export_filters(
    selections: list[dict[str, Any]] | None,
    class_filter_include: str = '',
) -> tuple[str, FilterConfig | None]:
    """Derive NetAct class filter and Excel param filter from UI selections."""
    if selections:
        class_filter = selections_to_class_filter_include(selections)
        param_filter = build_param_filter_from_selections(selections)
        if class_filter:
            return class_filter, param_filter
    manual = normalize_class_filter_include(class_filter_include)
    return manual, None


def _stringify_attributes(attributes: dict[str, Any]) -> dict[str, str]:
    result: dict[str, str] = {}
    for key, value in attributes.items():
        if value is None:
            continue
        if isinstance(value, bool):
            result[key] = 'true' if value else 'false'
        else:
            result[key] = str(value)
    return result


def _parse_export_object_count(feedbacks: list[dict[str, Any]]) -> int:
    total = 0
    for item in feedbacks:
        text = ' '.join(
            part for part in (
                str(item.get('title') or ''),
                str(item.get('details') or ''),
            ) if part
        )
        for pattern in (_EXPORT_OK_RE, _EXPORT_OBJECT_RE):
            match = pattern.search(text)
            if match:
                total = max(total, int(match.group(1)))
    return total


def _collect_export_errors(feedbacks: list[dict[str, Any]]) -> list[str]:
    errors: list[str] = []
    for item in feedbacks:
        if (item.get('type') or '').upper() != 'ERROR':
            continue
        title = str(item.get('title') or 'Export error').strip()
        details = str(item.get('details') or '').strip()
        errors.append(f'{title}: {details}' if details else title)
    return errors


def _operation_output_paths(attributes: list[dict[str, Any]]) -> tuple[str, str]:
    """Return ``(basename, full_remote_path)`` from operation attributes."""
    for item in attributes:
        attrs = item.get('operationAttributes') or {}
        output_file = str(attrs.get('outputFile') or '').strip()
        if output_file:
            return os.path.basename(output_file), output_file
        file_name = str(attrs.get('fileName') or '').strip()
        if file_name:
            return os.path.basename(file_name), ''
    return '', ''


def _operation_output_file(attributes: list[dict[str, Any]]) -> str:
    basename, _full = _operation_output_paths(attributes)
    return basename


def _export_remote_dirs(cfg: dict[str, Any]) -> list[str]:
    dirs: list[str] = []
    primary = str(cfg.get('remote_dir') or '').strip().rstrip('/')
    if primary:
        dirs.append(primary)
    extra = str(cfg.get('remote_dir_extra') or '').strip()
    for part in extra.split(','):
        part = part.strip().rstrip('/')
        if part and part not in dirs:
            dirs.append(part)
    for fallback in _FALLBACK_EXPORT_DIRS:
        if fallback not in dirs:
            dirs.append(fallback)
    return dirs


def _is_transient_sftp_error(exc: BaseException) -> bool:
    if isinstance(exc, (EOFError, ConnectionResetError, BrokenPipeError, TimeoutError, SSHException)):
        return True
    if isinstance(exc, (OSError, socket.error)):
        if getattr(exc, 'winerror', None) in _TRANSIENT_WINERRORS:
            return True
        if getattr(exc, 'errno', None) in (104, 110, 111, 113):
            return True
    msg = str(exc).lower()
    return any(
        token in msg
        for token in ('forcibly closed', 'connection reset', 'broken pipe', 'connection lost')
    )


def _open_sftp_session(cfg: dict[str, Any]) -> tuple[paramiko.SSHClient, paramiko.SFTPClient]:
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(
        cfg['host'],
        port=int(cfg.get('port') or 22),
        username=cfg['username'],
        password=cfg['password'],
        timeout=int(cfg.get('timeout') or 60),
        banner_timeout=int(cfg.get('timeout') or 60),
        auth_timeout=int(cfg.get('timeout') or 60),
        allow_agent=False,
        look_for_keys=False,
    )
    transport = ssh.get_transport()
    if transport:
        transport.set_keepalive(30)
    return ssh, ssh.open_sftp()


def _close_sftp_session(
    ssh: paramiko.SSHClient | None,
    sftp: paramiko.SFTPClient | None,
) -> None:
    if sftp is not None:
        try:
            sftp.close()
        except Exception:
            pass
    if ssh is not None:
        try:
            ssh.close()
        except Exception:
            pass


def _remote_file_stat(sftp: paramiko.SFTPClient, remote_path: str):
    try:
        return sftp.stat(remote_path)
    except FileNotFoundError:
        return None
    except OSError as exc:
        if getattr(exc, 'errno', None) == 2:
            return None
        raise


def _find_remote_export_file(
    sftp: paramiko.SFTPClient,
    remote_dirs: list[str],
    remote_filename: str,
) -> tuple[str, Any] | tuple[None, None]:
    target = remote_filename.strip()
    if not target:
        return None, None
    for remote_dir in remote_dirs:
        remote_path = f'{remote_dir.rstrip("/")}/{target}'
        stat = _remote_file_stat(sftp, remote_path)
        if stat is not None:
            return remote_path, stat
        try:
            for entry in sftp.listdir_attr(remote_dir):
                if entry.filename == target:
                    return f'{remote_dir.rstrip("/")}/{entry.filename}', entry
        except OSError:
            continue
    return None, None


def _resolve_remote_export_path(
    sftp: paramiko.SFTPClient,
    remote_dirs: list[str],
    remote_filename: str,
    *,
    remote_path_hint: str = '',
) -> tuple[str, Any] | tuple[None, None]:
    hint = (remote_path_hint or '').strip()
    if hint.startswith('/'):
        stat = _remote_file_stat(sftp, hint)
        if stat is not None:
            return hint, stat
    return _find_remote_export_file(sftp, remote_dirs, remote_filename)


def _download_sftp_file_resumable(
    sftp: paramiko.SFTPClient,
    remote_path: str,
    local_path: str,
) -> None:
    remote_size = int(_remote_file_stat(sftp, remote_path).st_size or 0)
    offset = 0
    if os.path.exists(local_path):
        offset = os.path.getsize(local_path)
        if remote_size and offset >= remote_size:
            return
        if offset and remote_size and offset > remote_size:
            os.remove(local_path)
            offset = 0

    with sftp.open(remote_path, 'rb') as remote_f:
        if offset:
            remote_f.seek(offset)
        mode = 'ab' if offset else 'wb'
        with open(local_path, mode) as local_f:
            while True:
                chunk = remote_f.read(_SFTP_CHUNK_BYTES)
                if not chunk:
                    break
                local_f.write(chunk)
                offset += len(chunk)

    if remote_size and os.path.getsize(local_path) < remote_size:
        raise OSError(
            f'Incomplete SFTP download ({os.path.getsize(local_path)} of {remote_size} bytes).',
        )


def fetch_export_file_via_sftp(
    remote_filename: str,
    local_path: str,
    *,
    remote_path_hint: str = '',
) -> None:
    cfg = nokia_export_ssh_settings()
    bulk_cfg = nokia_bulk_export_settings()
    remote_dirs = _export_remote_dirs(cfg)
    primary_dir = remote_dirs[0] if remote_dirs else '/d/oss/global/var/racops/export'
    if not cfg.get('configured'):
        raise NokiaBulkExportError(
            'SFTP is not configured to retrieve the RAML/XML export from the NetAct OMC. '
            'Set NOKIA_CM_SSH_HOST / NOKIA_CM_SSH_USER / NOKIA_CM_SSH_PASSWORD, or reuse '
            'NOKIA_PM_HOST / NOKIA_PM_USER / NOKIA_PM_PASSWORD in .env. '
            f'Expected file: {primary_dir}/{remote_filename}'
        )

    wait_sec = int(bulk_cfg.get('file_wait_sec') or cfg.get('file_wait_sec') or 600)
    poll_interval = 3
    deadline = time.time() + wait_sec
    last_error = 'Unknown SFTP error'
    credential_hint = (
        'Use OMC SFTP credentials (NOKIA_CM_SSH_* or NOKIA_PM_* / ftpuser on '
        f'{cfg["host"]}), not the CM REST API login.'
    )

    os.makedirs(os.path.dirname(local_path) or '.', exist_ok=True)

    while time.time() < deadline:
        ssh: paramiko.SSHClient | None = None
        sftp: paramiko.SFTPClient | None = None
        try:
            ssh, sftp = _open_sftp_session(cfg)
            remote_path, stat = _resolve_remote_export_path(
                sftp,
                remote_dirs,
                remote_filename,
                remote_path_hint=remote_path_hint,
            )
            if remote_path is None or stat is None:
                last_error = (
                    f'Export file not found yet: {primary_dir}/{remote_filename}'
                )
            else:
                _download_sftp_file_resumable(sftp, remote_path, local_path)
                if os.path.getsize(local_path) > 0:
                    return
                last_error = 'Downloaded export file is empty.'
        except FileNotFoundError:
            last_error = f'Export file not found yet: {primary_dir}/{remote_filename}'
        except Exception as exc:
            if _is_transient_sftp_error(exc):
                last_error = str(exc)
            else:
                raise NokiaBulkExportError(
                    f'Cannot retrieve export file from NetAct via SFTP '
                    f'({cfg["host"]}:{primary_dir}/{remote_filename}): {exc}. '
                    f'{credential_hint}',
                ) from exc
        finally:
            _close_sftp_session(ssh, sftp)

        time.sleep(poll_interval)

    raise NokiaBulkExportError(
        f'Timed out after {wait_sec}s waiting for export file on NetAct SFTP '
        f'({cfg["host"]}:{primary_dir}/{remote_filename}). '
        f'Last error: {last_error}. {credential_hint} '
        'Transient connection drops are retried automatically; if this persists, verify '
        'network access to port 22 on the OMC and that the export completed on NetAct.',
    )


def raml_xml_to_excel(
    xml_path: str,
    excel_path: str,
    *,
    param_filter: FilterConfig | None = None,
) -> tuple[int, list[str]]:
    converter = XMLToExcelConverter(xml_path, excel_path)
    if param_filter and param_filter.filter_dict:
        converter.filter_config = param_filter
    ok, message = converter.convert()
    if not ok:
        raise NokiaBulkExportError(message or 'Failed to convert RAML/XML export to Excel')

    import openpyxl

    wb = openpyxl.load_workbook(excel_path, read_only=True)
    sheet_names = list(wb.sheetnames)
    row_count = 0
    for name in sheet_names:
        ws = wb[name]
        row_count += max(0, ws.max_row - 2)
    wb.close()
    return row_count, sheet_names


def run_controller_bulk_export(
    cm_client: NokiaCmClient,
    ops_client: NokiaOperationsClient,
    *,
    scope_level: str,
    site_ids: list[str],
    selections: list[dict[str, Any]] | None = None,
    class_filter_include: str = '',
    file_format: str = 'RAML2',
    operation_timeout_sec: int | None = None,
) -> dict[str, Any]:
    """
    Trigger Import_Export actualExport for RNC/BSC scope and return local Excel path metadata.

    Returns dict with keys: xml_path, excel_path, operation_id, export_dn, remote_file,
    object_count, sheet_names, row_count, warnings.
    """
    level = normalize_scope_level(scope_level)
    if level not in ('MRBTS', 'RNC', 'BSC'):
        raise NokiaBulkExportError('Bulk CM Operations export supports MRBTS, RNC, and BSC scope.')

    bulk_cfg = nokia_bulk_export_settings()
    op_timeout = operation_timeout_sec or bulk_cfg['operation_timeout_sec']

    class_filter, param_filter = resolve_bulk_export_filters(
        selections,
        class_filter_include,
    )
    if not class_filter:
        raise NokiaBulkExportError(
            'Select at least one managed object class (and parameters) in the MO picker '
            'before running bulk export.'
        )

    dns = resolve_bulk_export_dns(cm_client, site_ids, scope_level=level)
    if not dns:
        raise NokiaBulkExportError('Could not resolve any controller DNs for the selected site id(s).')

    export_file = f'primenet_{level.lower()}_{uuid.uuid4().hex[:10]}.xml'
    attributes = _stringify_attributes({
        'importExportOperation': 'actualExport',
        'fileFormat': file_format or 'RAML2',
        'DN': ','.join(dns),
        'fileName': export_file,
        'useQualifiedClassAbbreviation': True,
        'includeSiteInfo': False,
        'includeMaintenanceregionInfo': False,
        'btsm': False,
    })
    if class_filter.strip():
        attributes['classFilterInclude'] = class_filter.strip()

    mo_summary = class_filter.strip()
    operation_id = ops_client.start_operation(
        'Import_Export',
        operation_alias=f'PrimeNet {level} bulk export ({mo_summary[:80]})',
        attributes=attributes,
    )
    status, feedbacks = ops_client.wait_for_operation(
        operation_id,
        timeout_sec=op_timeout,
    )
    if status == 'FAILED':
        errors = _collect_export_errors(feedbacks)
        raise NokiaBulkExportError(
            errors[0] if errors else f'CM Operations export failed (operation {operation_id}).',
        )
    if status == 'INTERRUPTED':
        raise NokiaBulkExportError(f'CM Operations export was interrupted (operation {operation_id}).')

    object_count = _parse_export_object_count(feedbacks)
    if object_count == 0:
        errors = _collect_export_errors(feedbacks)
        hint = (
            f'NetAct exported 0 MO instances for DN scope {", ".join(dns)} '
            f'with MO filter {class_filter.strip()!r}. '
            'Use the PrimeNet site id (e.g. 2012 → PLMN-PLMN/RNC-2012), not the short '
            'CM instance id (RNC-12). If the DN is correct, verify the selected MO '
            'classes exist under that controller.'
        )
        if errors:
            hint = f'{errors[0]} {hint}'
        raise NokiaBulkExportError(hint)

    attr_rows = ops_client.get_attributes([operation_id])
    remote_file, remote_path_hint = _operation_output_paths(attr_rows)
    if not remote_file:
        remote_file = export_file

    tmp_dir = tempfile.mkdtemp(prefix='cm_bulk_export_')
    xml_path = os.path.join(tmp_dir, remote_file)
    excel_path = os.path.join(tmp_dir, f'{level.lower()}_bulk_export.xlsx')

    fetch_export_file_via_sftp(
        remote_file,
        xml_path,
        remote_path_hint=remote_path_hint,
    )
    row_count, sheet_names = raml_xml_to_excel(xml_path, excel_path, param_filter=param_filter)

    warnings: list[str] = []
    if len(dns) > 1:
        scope_label = 'site(s)' if level == 'MRBTS' else 'controller(s)'
        warnings.append(f'Combined export for {len(dns)} {scope_label} in one RAML file.')
    warnings.append(
        f'CM Operations export: {object_count} MO instance(s) from NetAct Import_Export → '
        f'{len(sheet_names)} Excel sheet(s). MO filter: {mo_summary}.'
    )
    if param_filter and param_filter.filter_dict:
        param_classes = [
            abbr for abbr, params in param_filter.filter_dict.items() if params
        ]
        if param_classes:
            warnings.append(
                'Excel includes selected parameters only for: '
                + ', '.join(param_classes)
                + '. Other picked classes were exported with all parameters.'
            )

    return {
        'xml_path': xml_path,
        'excel_path': excel_path,
        'tmpdir': tmp_dir,
        'operation_id': operation_id,
        'export_dns': dns,
        'remote_file': remote_file,
        'object_count': object_count,
        'row_count': row_count,
        'sheet_names': sheet_names,
        'class_filter_include': mo_summary,
        'warnings': warnings,
    }


def export_controller_selection_to_excel(
    cm_client: NokiaCmClient,
    ops_client: NokiaOperationsClient,
    output_path: str,
    *,
    scope_level: str,
    site_ids: list[str],
    selections: list[dict[str, Any]],
) -> tuple[int, list[str], str]:
    """
    Extract via CM Operations Import_Export (full MO tree), not persistency API.

    Supported scopes: MRBTS (heavy LTE MOs), RNC, and BSC.
    """
    result = run_controller_bulk_export(
        cm_client,
        ops_client,
        scope_level=scope_level,
        site_ids=site_ids,
        selections=selections,
    )
    try:
        shutil.copy2(result['excel_path'], output_path)
    finally:
        tmpdir = result.get('tmpdir')
        if tmpdir:
            shutil.rmtree(tmpdir, ignore_errors=True)

    param_summary = []
    for sel in selections:
        abbr = (sel.get('mo_class_id') or sel.get('id') or '').split(':')[-1]
        params = sel.get('parameters') or []
        if abbr and params:
            param_summary.append(f'{abbr} ({len(params)} params)')

    sheet_names = result.get('sheet_names') or []
    summary = (
        f'{result["row_count"]} row(s) across {len(sheet_names)} sheet(s): '
        f'{", ".join(sheet_names[:8])}'
        f'{"…" if len(sheet_names) > 8 else ""}. '
    )
    if param_summary:
        summary += f'MO classes: {"; ".join(param_summary)}. '
    summary += (
        f'NetAct exported {result["object_count"]} MO instance(s) '
        f'via CM Operations Import_Export (operation {result["operation_id"]}).'
    )
    if result.get('warnings'):
        summary += ' ' + ' '.join(result['warnings'])
    return int(result['row_count'] or 0), sheet_names, summary
