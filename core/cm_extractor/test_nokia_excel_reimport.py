from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import openpyxl

from core.cm_extractor.nokia_excel_reimport import compare_nokia_workbooks


def _write_workbook(path: Path, *, pci: str = '100', mrbts: str = '1234', extra_header: str = '') -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'LNCEL'
    ws.cell(row=1, column=1, value='HIERARCHY_COLS:3')
    headers = ['MRBTS', 'LNBTS', 'LNCEL', 'name', 'phyCellId']
    if extra_header:
        headers.append(extra_header)
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=2, column=idx, value=header)
    values = [mrbts, '10', '20', 'Cell_A', pci]
    if extra_header:
        values.append('x')
    for idx, value in enumerate(values, start=1):
        ws.cell(row=3, column=idx, value=value)
    wb.save(path)
    wb.close()


class NokiaExcelReimportTests(unittest.TestCase):
    def test_detects_parameter_change(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp) / 'base.xlsx'
            edited = Path(tmp) / 'edited.xlsx'
            _write_workbook(base, pci='100')
            _write_workbook(edited, pci='101')

            result = compare_nokia_workbooks(base, edited)

            self.assertTrue(result['executable'])
            self.assertEqual(result['change_count'], 1)
            self.assertEqual(result['changes'][0]['parameter'], 'phyCellId')
            self.assertEqual(result['changes'][0]['old_value'], '100')
            self.assertEqual(result['changes'][0]['new_value'], '101')

    def test_blocks_hierarchy_change_as_new_row(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp) / 'base.xlsx'
            edited = Path(tmp) / 'edited.xlsx'
            _write_workbook(base, mrbts='1234')
            _write_workbook(edited, mrbts='9999')

            result = compare_nokia_workbooks(base, edited)

            self.assertFalse(result['executable'])
            self.assertGreaterEqual(result['blocked_count'], 1)

    def test_blocks_blank_without_flag(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp) / 'base.xlsx'
            edited = Path(tmp) / 'edited.xlsx'
            _write_workbook(base, pci='100')
            _write_workbook(edited, pci='')

            result = compare_nokia_workbooks(base, edited)

            self.assertFalse(result['executable'])
            self.assertEqual(result['changes'], [])
            self.assertIn('Blanking', result['blocked'][0]['reason'])

    def test_blocks_unknown_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp) / 'base.xlsx'
            edited = Path(tmp) / 'edited.xlsx'
            _write_workbook(base)
            _write_workbook(edited, extra_header='newParam')

            result = compare_nokia_workbooks(base, edited)

            self.assertFalse(result['executable'])
            self.assertIn('Unknown/new columns', result['blocked'][0]['reason'])


if __name__ == '__main__':
    unittest.main()
