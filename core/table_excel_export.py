"""Generic Excel workbook builder for PrimeNet table exports."""

from __future__ import annotations

import io
import re
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _safe_filename_part(value: str, *, fallback: str = 'export') -> str:
    cleaned = re.sub(r'[^\w\-]+', '_', (value or '').strip())
    return cleaned.strip('_')[:48] or fallback


def build_table_workbook(
    *,
    filename_stem: str,
    report_title: str,
    sheet_title: str,
    columns: list[str],
    rows: list[dict[str, Any]],
    meta: dict[str, Any] | list[tuple[str, Any]] | None = None,
    column_labels: dict[str, str] | None = None,
) -> tuple[io.BytesIO, str]:
    if not columns:
        raise ValueError('No columns to export')
    if not rows:
        raise ValueError('No rows to export')

    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'{_safe_filename_part(filename_stem)}_{stamp}.xlsx'

    wb = Workbook()
    hdr_fill = PatternFill(start_color='1F6FEB', end_color='1F6FEB', fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True)

    ws_meta = wb.active
    ws_meta.title = 'Report Info'
    ws_meta.column_dimensions['A'].width = 24
    ws_meta.column_dimensions['B'].width = 52
    generated = datetime.now(timezone.utc).replace(microsecond=0).isoformat()

    meta_rows: list[tuple[str, Any]] = [('Report', report_title), ('Generated (UTC)', generated)]
    if isinstance(meta, dict):
        meta_rows.extend((str(k), v) for k, v in meta.items())
    elif isinstance(meta, list):
        meta_rows.extend(meta)

    ws_meta['A1'] = 'Field'
    ws_meta['B1'] = 'Value'
    for col, label in enumerate(('Field', 'Value'), 1):
        cell = ws_meta.cell(row=1, column=col, value=label)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')
    for row_idx, (field, value) in enumerate(meta_rows, start=2):
        ws_meta.cell(row=row_idx, column=1, value=field)
        ws_meta.cell(row=row_idx, column=2, value=value)

    ws_data = wb.create_sheet((sheet_title or 'Data')[:31])
    labels = column_labels or {}
    headers = [str(labels.get(col) or col) for col in columns]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_data.cell(row=1, column=col_idx, value=header)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for row_idx, item in enumerate(rows, start=2):
        if not isinstance(item, dict):
            continue
        for col_idx, col in enumerate(columns, start=1):
            value = item.get(col)
            if value is None:
                value = ''
            ws_data.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, col in enumerate(columns, start=1):
        max_len = len(str(labels.get(col) or col))
        for row_idx in range(2, min(len(rows) + 2, 102)):
            cell_val = ws_data.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(cell_val or '')))
        ws_data.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 48)

    ws_data.freeze_panes = 'A2'
    if rows:
        ws_data.auto_filter.ref = f'A1:{get_column_letter(len(columns))}{len(rows) + 1}'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, filename
