"""
Nokia Configuration Manager - Core Processing Logic
Extracted from NCM_V3.py without GUI components
"""

import re
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import datetime
from typing import Any

_INVALID_SHEET_TITLE_CHARS = re.compile(r'[\:\\/?*\[\]]')


def _sanitize_sheet_title(name: str) -> str:
    """Excel sheet titles cannot contain : \\ / ? * [ ]."""
    cleaned = _INVALID_SHEET_TITLE_CHARS.sub('_', (name or '').strip())
    return (cleaned or 'Sheet')[:31]


def _unique_sheet_title(mo_class: str, used: set[str]) -> str:
    base = _sanitize_sheet_title(mo_class)
    title = base
    suffix = 2
    while title in used:
        tail = f'_{suffix}'
        title = f'{base[:31 - len(tail)]}{tail}'
        suffix += 1
    used.add(title)
    return title

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    raise ImportError("openpyxl not installed. Install with: pip install openpyxl")

try:
    import pandas as pd
except ImportError:
    raise ImportError("pandas not installed. Install with: pip install pandas")

def import_parameters_from_excel(excel_file):
    """
    Import parameter selections from Excel file.
    Supports both .xlsx and .xls formats.
    Excel must have two columns: 'MO' and 'Parameter'
    
    Returns: dict {mo_class: set(parameters)}
    """
    try:
        df = pd.read_excel(excel_file)
        
        if 'MO' not in df.columns or 'Parameter' not in df.columns:
            raise ValueError("Excel file must have 'MO' and 'Parameter' columns")
        
        param_dict = defaultdict(set)
        for _, row in df.iterrows():
            mo = str(row['MO']).strip()
            param = str(row['Parameter']).strip()
            
            if mo and param and mo != 'nan' and param != 'nan':
                param_dict[mo].add(param)
        
        return dict(param_dict)
        
    except Exception as e:
        raise Exception(f"Error importing parameters from Excel: {str(e)}")


class FilterConfig:
    """Stores filter configuration with both filtering dict and all MOs list"""
    def __init__(self, filter_dict=None, all_mos=None):
        self.filter_dict = filter_dict or {}
        self.all_mos = all_mos or set()


def discover_hierarchy_elements(hierarchy_elements: list[str], distname: str) -> None:
    """Record MO class tokens from a distName path in first-seen depth order."""
    for part in (distname or '').split('/'):
        if '-' in part:
            mo_class = part.split('-', 1)[0]
            if mo_class != 'PLMN' and mo_class not in hierarchy_elements:
                hierarchy_elements.append(mo_class)


def parse_distname_hierarchy(distname: str) -> dict[str, str]:
    """Parse distName into {MO_CLASS: instance_id} (excludes PLMN)."""
    hierarchy: dict[str, str] = {}
    for part in (distname or '').split('/'):
        if '-' in part:
            mo_class, mo_id = part.split('-', 1)
            if mo_class != 'PLMN':
                hierarchy[mo_class] = mo_id
    return hierarchy


def sort_hierarchy_columns(hierarchy_elements: list[str], hierarchy_cols: list[str]) -> list[str]:
    """Sort hierarchy columns by depth order discovered from distNames."""
    if not hierarchy_cols:
        return []
    sorted_cols: list[str] = []
    for col in hierarchy_elements:
        if col in hierarchy_cols:
            sorted_cols.append(col)
    for col in hierarchy_cols:
        if col not in sorted_cols:
            sorted_cols.append(col)
    return sorted_cols


def sort_parameter_columns(param_cols: list[str]) -> list[str]:
    """Sort parameter columns with list markers followed by Item-* children."""
    list_markers: set[str] = set()
    list_details: dict[str, list[str]] = {}
    regular_params: list[str] = []

    for col in param_cols:
        if col.startswith('Item-'):
            parts = col.split('-', 2)
            if len(parts) >= 2:
                list_name = parts[1]
                list_details.setdefault(list_name, []).append(col)
        else:
            has_children = any(
                detail_col.startswith(f'Item-{col}-') for detail_col in param_cols
            )
            if has_children:
                list_markers.add(col)
            else:
                regular_params.append(col)

    result: list[str] = []
    result.extend(sorted(regular_params))
    for list_name in sorted(list_markers):
        result.append(list_name)
        result.extend(sorted(list_details.get(list_name, [])))
    return result


class NokiaMoSheetBuilder:
    """
    Build Excel sheet rows using NCM core hierarchy discovery and list integration.

    Matches ``XMLToExcelConverter`` column layout: hierarchy columns first (depth
    order), then scalar parameters, then list markers with ``Item-{list}-{param}``
    columns. Semicolon-joins repeated list-item values.
    """

    _SKIP_PARAM_NAMES = frozenset({'PLMN'})
    _IDENTITY_COLS = frozenset({'moId', 'DN'})

    def __init__(self) -> None:
        self.hierarchy_elements: list[str] = []
        self.row_dicts: list[dict[str, Any]] = []
        self.all_columns: set[str] = set()

    def discover_hierarchy(self, distname: str) -> None:
        discover_hierarchy_elements(self.hierarchy_elements, distname)

    def parse_distname(self, distname: str) -> dict[str, str]:
        return parse_distname_hierarchy(distname)

    def flatten_api_parameters(self, parameters: dict[str, Any] | None) -> dict[str, str]:
        result: dict[str, str] = {}
        for name, value in (parameters or {}).items():
            if name in self._SKIP_PARAM_NAMES:
                continue
            result.update(self._flatten_value(name, value))
        return result

    def _flatten_value(self, name: str, value: Any) -> dict[str, str]:
        if value is None:
            return {name: ''}
        if isinstance(value, bool):
            return {name: 'true' if value else 'false'}
        if isinstance(value, (int, float)):
            return {name: str(value)}
        if isinstance(value, str):
            return {name: value}
        if isinstance(value, list):
            return self._flatten_list(name, value)
        if isinstance(value, dict):
            return self._flatten_structured_value(name, value)
        return {name: str(value)}

    def _flatten_list(self, list_name: str, items: list[Any]) -> dict[str, str]:
        if not items:
            return {list_name: ''}
        if all(isinstance(item, dict) for item in items):
            result: dict[str, str] = {list_name: 'List'}
            list_params: dict[str, list[str]] = defaultdict(list)
            for item in items:
                for key, val in item.items():
                    if key in self._SKIP_PARAM_NAMES:
                        continue
                    for flat_key, flat_val in self._flatten_value(str(key), val).items():
                        list_params[flat_key].append(flat_val)
            for param_name, values in list_params.items():
                full_key = f'Item-{list_name}-{param_name}'
                if len(values) > 1:
                    result[full_key] = ';'.join(values)
                elif len(values) == 1:
                    result[full_key] = values[0]
            return result
        scalar_values = [str(item) for item in items]
        if len(scalar_values) > 1:
            return {list_name: ';'.join(scalar_values)}
        return {list_name: scalar_values[0]}

    def _flatten_structured_value(self, name: str, value: dict[str, Any]) -> dict[str, str]:
        items = value.get('items')
        if isinstance(items, list):
            return self._flatten_list(name, items)

        nested_scalars: dict[str, str] = {}
        nested_lists: dict[str, list[Any]] = {}
        for key, val in value.items():
            if key in self._SKIP_PARAM_NAMES:
                continue
            if isinstance(val, list):
                nested_lists[str(key)] = val
            elif isinstance(val, dict):
                nested_scalars.update(self._flatten_structured_value(str(key), val))
            elif val is not None:
                nested_scalars[str(key)] = str(val)

        if nested_lists:
            result = {name: 'List'}
            list_params: dict[str, list[str]] = defaultdict(list)
            for list_key, list_val in nested_lists.items():
                flat = self._flatten_list(list_key, list_val)
                for flat_key, flat_val in flat.items():
                    if flat_key == list_key:
                        continue
                    if flat_key.startswith('Item-'):
                        list_params[flat_key.split('-', 2)[-1]].append(flat_val)
                    else:
                        list_params[flat_key].append(flat_val)
            for param_name, values in list_params.items():
                full_key = f'Item-{name}-{param_name}'
                if len(values) > 1:
                    result[full_key] = ';'.join(values)
                elif len(values) == 1:
                    result[full_key] = values[0]
            result.update(nested_scalars)
            return result

        if nested_scalars:
            return nested_scalars
        return {name: str(value)}

    def add_managed_object(
        self,
        distname: str,
        parameters: dict[str, Any] | None = None,
        *,
        extra: dict[str, Any] | None = None,
    ) -> None:
        self.discover_hierarchy(distname)
        row: dict[str, Any] = {}
        row.update(self.parse_distname(distname))
        if extra:
            for key, val in extra.items():
                if key in self._IDENTITY_COLS:
                    continue
                row[key] = val
        row.update(self.flatten_api_parameters(parameters))
        self.row_dicts.append(row)
        self.all_columns.update(row.keys())

    def add_flat_row(self, row_data: dict[str, Any], *, distname: str = '') -> None:
        """Add a row that may already contain hierarchy columns and/or a DN/moId."""
        dn = str(
            distname
            or row_data.get('DN')
            or row_data.get('moId')
            or ''
        ).strip()
        if dn:
            self.discover_hierarchy(dn)
        row: dict[str, Any] = {}
        if dn:
            row.update(self.parse_distname(dn))
        for key, val in row_data.items():
            if key in self._IDENTITY_COLS:
                continue
            if key in row and row[key] == val:
                continue
            if isinstance(val, (dict, list)):
                row.update(self.flatten_api_parameters({str(key): val}))
            else:
                row[key] = '' if val is None else val
        self.row_dicts.append(row)
        self.all_columns.update(row.keys())

    def seed_hierarchy_columns(self, columns: list[str]) -> None:
        for col in columns:
            if col and col not in self.hierarchy_elements and col not in self._IDENTITY_COLS:
                self.hierarchy_elements.append(col)

    def ordered_columns(self) -> tuple[list[str], int]:
        hierarchy_cols = sort_hierarchy_columns(
            self.hierarchy_elements,
            [col for col in self.all_columns if col in self.hierarchy_elements],
        )
        param_cols = [
            col for col in self.all_columns
            if col not in self.hierarchy_elements and col not in self._IDENTITY_COLS
        ]
        param_cols_sorted = sort_parameter_columns(param_cols)
        ordered = hierarchy_cols + param_cols_sorted
        return ordered, len(hierarchy_cols)

    def to_sheet(self) -> dict[str, Any]:
        if not self.row_dicts:
            return {
                'headers': ['moId'],
                'rows': [],
                'hierarchy_col_count': 0,
                'mo_count': 0,
            }
        ordered, hierarchy_col_count = self.ordered_columns()
        rows = [[row.get(col, '') for col in ordered] for row in self.row_dicts]
        return {
            'headers': ordered,
            'rows': rows,
            'hierarchy_col_count': hierarchy_col_count,
            'mo_count': len(self.row_dicts),
        }


def managed_objects_to_ncm_sheet(managed_objects: list[dict[str, Any]]) -> dict[str, Any]:
    """Convert getManagedObjects payload to NCM-layout {headers, rows, hierarchy_col_count}."""
    builder = NokiaMoSheetBuilder()
    for mo in managed_objects or []:
        distname = str(mo.get('moId') or mo.get('distName') or '').strip()
        if not distname:
            continue
        builder.add_managed_object(distname, mo.get('parameters') or {})
    return builder.to_sheet()


def query_table_to_ncm_sheet(headers: list[str], rows: list[list[Any]]) -> dict[str, Any]:
    """Convert CM query result table (DN + parameters) to NCM-layout sheet data."""
    builder = NokiaMoSheetBuilder()
    if not headers:
        return builder.to_sheet()

    for row in rows or []:
        row_map = {
            headers[idx]: row[idx] if idx < len(row) else ''
            for idx in range(len(headers))
        }
        builder.add_flat_row(row_map)
    return builder.to_sheet()


def merge_ncm_sheet_parts(parts: list[dict[str, Any]]) -> dict[str, Any]:
    """Merge per-site/per-chunk sheet parts preserving NCM hierarchy + list column order."""
    builder = NokiaMoSheetBuilder()
    for part in parts or []:
        headers = list(part.get('headers') or [])
        hierarchy_col_count = int(part.get('hierarchy_col_count') or 0)
        if hierarchy_col_count and headers:
            builder.seed_hierarchy_columns(headers[:hierarchy_col_count])
        elif headers and headers[0] == 'DN':
            for row in part.get('rows') or []:
                if row:
                    builder.discover_hierarchy(str(row[0]))
        for row in part.get('rows') or []:
            row_map = {
                headers[idx]: row[idx] if idx < len(row) else ''
                for idx in range(len(headers))
            }
            builder.add_flat_row(row_map)
    return builder.to_sheet()


def _mo_class_abbreviation(mo_class: str) -> str:
    """RAML ``class`` may be ``FMCS`` or ``com.nokia.asrnc:FMCS`` — return ``FMCS``."""
    token = (mo_class or '').strip()
    if ':' in token:
        return token.rsplit(':', 1)[-1]
    return token


class XMLToExcelConverter:
    """Convert Nokia XML configuration to Excel format"""
    
    def __init__(self, xml_file, output_file):
        self.xml_file = xml_file
        self.output_file = output_file
        self.data_by_mo = defaultdict(list)
        self.all_columns_by_mo = defaultdict(set)
        self.hierarchy_elements = []
        self.filter_config = FilterConfig()
    
    def set_parameters(self, param_dict):
        """Set parameter filtering"""
        self.filter_config.filter_dict = param_dict
        self.filter_config.all_mos = set(param_dict.keys())
    
    def discover_hierarchy(self, distname):
        """Extract hierarchy elements from distName path"""
        discover_hierarchy_elements(self.hierarchy_elements, distname)
    
    def parse_distname(self, distname):
        """Parse distName into hierarchy dictionary"""
        return parse_distname_hierarchy(distname)
    
    def sort_hierarchy_columns(self, hierarchy_cols):
        """Sort hierarchy columns by depth order"""
        return sort_hierarchy_columns(self.hierarchy_elements, hierarchy_cols)
    
    def sort_parameter_columns(self, param_cols):
        """Sort parameter columns with list markers followed by their children"""
        return sort_parameter_columns(param_cols)
    
    def extract_parameters(self, mo_elem):
        """Extract all parameters from a managedObject element"""
        result = {}
        
        for child in mo_elem:
            tag = child.tag
            
            if tag.endswith('p'):
                name = child.attrib.get("name", "")
                if name == "PLMN":
                    continue
                value = child.text.strip() if child.text else ""
                result[name] = value
            
            elif tag.endswith('list'):
                list_name = child.attrib.get("name", "list")
                items = [item for item in child if item.tag.endswith('item')]
                
                result[list_name] = "List"
                
                list_params = {}
                for item in items:
                    item_params = {}
                    self._extract_list_item_params(item, item_params)
                    
                    for param_name, param_value in item_params.items():
                        if param_name not in list_params:
                            list_params[param_name] = []
                        list_params[param_name].append(param_value)
                
                for param_name, values in list_params.items():
                    full_key = f"Item-{list_name}-{param_name}"
                    if len(values) > 1:
                        result[full_key] = ';'.join(values)
                    elif len(values) == 1:
                        result[full_key] = values[0]
            
            else:
                self._extract_nested_parameters(child, result)
        
        return result
    
    def _extract_list_item_params(self, item, result):
        """Extract parameters from a single list item"""
        for child in item:
            tag = child.tag
            
            if tag.endswith('p'):
                name = child.attrib.get("name", "")
                value = child.text.strip() if child.text else ""
                result[name] = value
            elif tag.endswith('list'):
                list_name = child.attrib.get("name", "list")
                nested_items = [i for i in child if i.tag.endswith('item')]
                
                nested_values = []
                for nested_item in nested_items:
                    nested_params = {}
                    self._extract_list_item_params(nested_item, nested_params)
                    nested_values.extend(nested_params.values())
                
                if nested_values:
                    result[list_name] = ';'.join(nested_values)
            else:
                self._extract_list_item_params(child, result)
    
    def _extract_nested_parameters(self, element, result):
        """Recursively extract parameters from nested elements"""
        for child in element:
            tag = child.tag
            
            if tag.endswith('p'):
                name = child.attrib.get("name", "")
                if name == "PLMN":
                    continue
                value = child.text.strip() if child.text else ""
                result[name] = value
            elif tag.endswith('list'):
                list_name = child.attrib.get("name", "list")
                items = [item for item in child if item.tag.endswith('item')]
                
                result[list_name] = "List"
                
                list_params = {}
                for item in items:
                    item_params = {}
                    self._extract_list_item_params(item, item_params)
                    
                    for param_name, param_value in item_params.items():
                        if param_name not in list_params:
                            list_params[param_name] = []
                        list_params[param_name].append(param_value)
                
                for param_name, values in list_params.items():
                    full_key = f"Item-{list_name}-{param_name}"
                    if len(values) > 1:
                        result[full_key] = ';'.join(values)
                    elif len(values) == 1:
                        result[full_key] = values[0]
            else:
                self._extract_nested_parameters(child, result)
    
    def _filter_key_for_mo_class(self, mo_class):
        """Map RAML class attribute to a filter_dict key (UI uses abbreviations like FMCS)."""
        if not self.filter_config.filter_dict:
            return None
        if mo_class in self.filter_config.filter_dict:
            return mo_class
        abbr = _mo_class_abbreviation(mo_class)
        if abbr in self.filter_config.filter_dict:
            return abbr
        return None

    def should_include_mo_param(self, mo_class, param_name):
        """Check if MO/param combo should be included."""
        if not self.filter_config.filter_dict:
            return True

        filter_key = self._filter_key_for_mo_class(mo_class)
        if filter_key is None:
            return False

        allowed = self.filter_config.filter_dict[filter_key]
        if not allowed:
            return True
        return param_name in allowed
    
    def should_process_mo(self, mo_class):
        """Check if this MO class matches any filter"""
        if not self.filter_config.filter_dict:
            return True
        
        return self._filter_key_for_mo_class(mo_class) is not None
    
    def convert(self, progress_callback=None):
        """Main conversion method - returns (success, message)"""
        try:
            print(f"\nConverting: {self.xml_file}")
            print(f"Output: {self.output_file}")
            
            # PASS 1: Discover hierarchy
            if progress_callback:
                progress_callback(10, "Discovering hierarchy structure...")
            
            processed_count = 0
            skipped_count = 0
            
            for event, elem in ET.iterparse(self.xml_file, events=('end',)):
                if elem.tag.endswith('managedObject'):
                    mo_class = elem.get('class', '')
                    
                    if not self.should_process_mo(mo_class):
                        skipped_count += 1
                        elem.clear()
                        continue
                    
                    distname = elem.get('distName', '')
                    self.discover_hierarchy(distname)
                    processed_count += 1
                    elem.clear()
            
            if progress_callback:
                progress_callback(40, "Extracting data...")
            
            # PASS 2: Extract data
            for event, elem in ET.iterparse(self.xml_file, events=('end',)):
                if elem.tag.endswith('managedObject'):
                    mo_class = elem.get('class', '')
                    
                    if not self.should_process_mo(mo_class):
                        elem.clear()
                        continue
                    
                    distname = elem.get('distName', '')
                    hierarchy = self.parse_distname(distname)
                    params_dict = self.extract_parameters(elem)
                    
                    # Apply parameter filtering
                    if self.filter_config.filter_dict:
                        filtered_params = {k: v for k, v in params_dict.items() 
                                          if self.should_include_mo_param(mo_class, k)}
                        params_dict = filtered_params
                    
                    row_data = {}
                    row_data.update(hierarchy)
                    row_data.update(params_dict)
                    
                    mo_class_key = _mo_class_abbreviation(mo_class)
                    self.data_by_mo[mo_class_key].append(row_data)
                    self.all_columns_by_mo[mo_class_key].update(row_data.keys())
                    
                    elem.clear()
            
            if progress_callback:
                progress_callback(70, "Writing Excel file...")
            
            # Write Excel
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            used_sheet_titles: set[str] = set()

            for mo_class in sorted(self.data_by_mo.keys()):
                mo_data = self.data_by_mo[mo_class]
                
                if not mo_data:
                    continue
                
                sheet_name = _unique_sheet_title(mo_class, used_sheet_titles)
                ws = wb.create_sheet(title=sheet_name)
                
                all_cols = self.all_columns_by_mo[mo_class]
                hierarchy_cols = self.sort_hierarchy_columns([c for c in all_cols if c in self.hierarchy_elements])
                param_cols = [c for c in all_cols if c not in self.hierarchy_elements]
                param_cols_sorted = self.sort_parameter_columns(param_cols)
                ordered_cols = hierarchy_cols + param_cols_sorted
                
                # Write metadata
                metadata_cell = ws.cell(row=1, column=1)
                metadata_cell.value = f"HIERARCHY_COLS:{len(hierarchy_cols)}"
                metadata_cell.font = Font(bold=True, italic=True, color="FF0000")
                
                # Write headers
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for col_idx, col_name in enumerate(ordered_cols, start=1):
                    cell = ws.cell(row=2, column=col_idx)
                    cell.value = col_name
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Write data
                for row_idx, row_data in enumerate(mo_data, start=3):
                    for col_idx, col_name in enumerate(ordered_cols, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = row_data.get(col_name, '')
            
            if progress_callback:
                progress_callback(90, "Saving file...")

            if not wb.sheetnames:
                raml_classes = sorted(self.data_by_mo.keys()) or sorted(
                    cls for cls in (self.all_columns_by_mo or {}) if cls
                )
                filter_keys = sorted(self.filter_config.filter_dict.keys())
                raise ValueError(
                    'No MO data matched the Excel filter after RAML conversion. '
                    f'RAML classes: {raml_classes or "(none parsed)"}; '
                    f'filter keys: {filter_keys}. '
                    'Qualified RAML class names (e.g. com.nokia.asrnc:FMCS) must map to '
                    'the MO abbreviation (FMCS) from your selection.'
                )
            
            wb.save(self.output_file)
            
            if progress_callback:
                progress_callback(100, "Complete!")
            
            return True, "Conversion successful"
            
        except Exception as e:
            error_msg = f"Error during conversion: {str(e)}"
            print(error_msg)
            import traceback
            traceback.print_exc()
            return False, error_msg


class ExcelToXMLConverter:
    """Convert Excel back to Nokia XML format"""
    
    def __init__(self, excel_file, output_xml):
        self.excel_file = excel_file
        self.output_xml = output_xml
        self.operations = {}
        self.version = "xL21A_2012_003"
        self.mo_order = None
    
    def discover_sheets(self):
        """Discover all MO sheets in Excel file"""
        try:
            wb = openpyxl.load_workbook(self.excel_file, read_only=True)
            mo_classes = []
            
            for sheet_name in wb.sheetnames:
                if sheet_name == 'INDEX':
                    continue
                
                if '_' in sheet_name:
                    mo_class = sheet_name.split('_', 1)[1]
                else:
                    mo_class = sheet_name
                
                if mo_class not in mo_classes:
                    mo_classes.append(mo_class)
            
            wb.close()
            return sorted(mo_classes)
        except Exception as e:
            raise Exception(f"Error reading Excel file: {str(e)}")
    
    def set_operations(self, operations):
        """Set operations for each MO class"""
        self.operations = operations
    
    def set_mo_order(self, mo_order):
        """Set processing order for MO classes"""
        self.mo_order = mo_order
    
    def build_distname(self, row, hierarchy_cols):
        """Build distName from hierarchy columns"""
        parts = ["PLMN-PLMN"]
        
        for col in hierarchy_cols:
            value = str(row.get(col, '')).strip()
            if value:
                parts.append(f"{col}-{value}")
        
        return '/'.join(parts)
    
    def parse_list_column(self, col_name):
        """Parse Item-listname-param into components"""
        if not col_name.startswith('Item-'):
            return None, None
        
        parts = col_name.split('-', 2)
        if len(parts) >= 3:
            return parts[1], parts[2]
        return None, None
    
    def build_parameters(self, row, param_cols):
        """Build parameter structure from row data"""
        params = {}
        lists = {}
        
        for col in param_cols:
            value = row.get(col, '')
            
            if value is None or (isinstance(value, str) and value.strip() == ''):
                continue
            
            if value == "List":
                continue
            
            if col.startswith('Item-'):
                list_name, param_name = self.parse_list_column(col)
                if list_name and param_name:
                    if list_name not in lists:
                        lists[list_name] = {}
                    
                    # Split semicolon-separated values
                    if isinstance(value, str) and ';' in value:
                        values = value.split(';')
                    else:
                        values = [str(value)]
                    
                    for idx, val in enumerate(values):
                        if idx not in lists[list_name]:
                            lists[list_name][idx] = {}
                        lists[list_name][idx][param_name] = val.strip()
            else:
                params[col] = str(value)
        
        return params, lists
    
    def convert(self, progress_callback=None):
        """Main conversion method - returns (success, message)"""
        try:
            if progress_callback:
                progress_callback(10, "Reading Excel file...")
            
            wb = openpyxl.load_workbook(self.excel_file)
            
            # Build XML structure
            root = ET.Element("raml", {
                "version": "2.0",
                "xmlns": "raml21.xsd"
            })
            cm_data = ET.SubElement(root, "cmData", {
                "type": "plan",
                "name": "Nokia Config",
                "version": self.version
            })
            header = ET.SubElement(cm_data, "header")
            ET.SubElement(header, "log", {
                "dateTime": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
                "action": "created",
                "appInfo": "Nokia Configuration Manager"
            })
            
            if progress_callback:
                progress_callback(30, "Processing sheets...")
            
            # Process sheets
            mo_classes = self.discover_sheets()
            for mo_class in mo_classes:
                if mo_class not in self.operations:
                    continue
                
                operation = self.operations[mo_class]
                sheet_name = mo_class[:31]
                
                if sheet_name not in wb.sheetnames:
                    continue
                
                ws = wb[sheet_name]
                
                # Read metadata
                metadata = ws.cell(row=1, column=1).value
                if metadata and metadata.startswith("HIERARCHY_COLS:"):
                    hierarchy_count = int(metadata.split(':')[1])
                else:
                    hierarchy_count = 0
                
                # Read headers
                headers = []
                for col in range(1, ws.max_column + 1):
                    header_val = ws.cell(row=2, column=col).value
                    if header_val:
                        headers.append(header_val)
                
                hierarchy_cols = headers[:hierarchy_count]
                param_cols = headers[hierarchy_count:]
                
                # Process rows
                for row_idx in range(3, ws.max_row + 1):
                    row_data = {}
                    for col_idx, header in enumerate(headers, start=1):
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        row_data[header] = cell_value
                    
                    # Skip empty rows
                    if not any(row_data.values()):
                        continue
                    
                    distname = self.build_distname(row_data, hierarchy_cols)
                    params, lists = self.build_parameters(row_data, param_cols)
                    
                    # Create managedObject element
                    mo_elem = ET.SubElement(cm_data, "managedObject", {
                        "class": mo_class,
                        "distName": distname,
                        "operation": operation
                    })
                    
                    # Add parameters
                    for param_name, param_value in params.items():
                        ET.SubElement(mo_elem, "p", {"name": param_name}).text = param_value
                    
                    # Add lists
                    for list_name, items in lists.items():
                        list_elem = ET.SubElement(mo_elem, "list", {"name": list_name})
                        for item_idx in sorted(items.keys()):
                            item_elem = ET.SubElement(list_elem, "item")
                            for param_name, param_value in items[item_idx].items():
                                ET.SubElement(item_elem, "p", {"name": param_name}).text = param_value
            
            if progress_callback:
                progress_callback(80, "Writing XML file...")
            
            # Write XML
            tree = ET.ElementTree(root)
            ET.indent(tree, space="  ")
            tree.write(self.output_xml, encoding='utf-8', xml_declaration=True)
            
            if progress_callback:
                progress_callback(100, "Complete!")
            
            return True, "Conversion successful"
            
        except Exception as e:
            error_msg = f"Error during conversion: {str(e)}"
            print(error_msg)
            import traceback
            traceback.print_exc()
            return False, error_msg


class XMLComparator:
    """Compare two XML files and generate Excel report"""
    
    def __init__(self, xml1_path, xml2_path, output_excel):
        self.xml1_path = xml1_path
        self.xml2_path = xml2_path
        self.output_excel = output_excel
    
    def parse_xml_to_dict(self, xml_path):
        """Parse XML into dictionary keyed by distName"""
        data = {}
        
        for event, elem in ET.iterparse(xml_path, events=('end',)):
            if elem.tag.endswith('managedObject'):
                mo_class = elem.get('class', '')
                distname = elem.get('distName', '')
                
                params = {}
                for child in elem:
                    if child.tag.endswith('p'):
                        name = child.attrib.get("name", "")
                        value = child.text.strip() if child.text else ""
                        params[name] = value
                
                key = f"{mo_class}|{distname}"
                data[key] = {
                    'mo_class': mo_class,
                    'distname': distname,
                    'params': params
                }
                
                elem.clear()
        
        return data
    
    def compare(self, progress_callback=None):
        """Compare XML files - returns (success, diff_count)"""
        try:
            if progress_callback:
                progress_callback(10, "Parsing XML 1...")
            
            data1 = self.parse_xml_to_dict(self.xml1_path)
            
            if progress_callback:
                progress_callback(30, "Parsing XML 2...")
            
            data2 = self.parse_xml_to_dict(self.xml2_path)
            
            if progress_callback:
                progress_callback(50, "Comparing...")
            
            # Find differences grouped by MO class
            differences_by_mo = defaultdict(list)
            all_keys = set(data1.keys()) | set(data2.keys())

            for key in all_keys:
                mo_class = key.split('|')[0]
                distname = key.split('|', 1)[1]

                if key in data1 and key in data2:
                    # Compare parameters
                    params1 = data1[key]['params']
                    params2 = data2[key]['params']

                    all_params = set(params1.keys()) | set(params2.keys())

                    for param in all_params:
                        val1 = params1.get(param, '')
                        val2 = params2.get(param, '')

                        if val1 != val2:
                            differences_by_mo[mo_class].append({
                                'MO_Class': mo_class,
                                'distName': distname,
                                'Parameter': param,
                                'Value_XML1': val1,
                                'Value_XML2': val2,
                                'Status': 'MODIFIED' if (val1 and val2) else ('ADDED' if val2 else 'DELETED')
                            })

                elif key in data1:
                    # Element deleted
                    for param, value in data1[key]['params'].items():
                        differences_by_mo[mo_class].append({
                            'MO_Class': mo_class,
                            'distName': distname,
                            'Parameter': param,
                            'Value_XML1': value,
                            'Value_XML2': '',
                            'Status': 'DELETED'
                        })

                else:
                    # Element added
                    for param, value in data2[key]['params'].items():
                        differences_by_mo[mo_class].append({
                            'MO_Class': mo_class,
                            'distName': distname,
                            'Parameter': param,
                            'Value_XML1': '',
                            'Value_XML2': value,
                            'Status': 'ADDED'
                        })
            
            if progress_callback:
                progress_callback(70, "Writing Excel report...")

            # Calculate total differences
            total_differences = sum(len(diffs) for diffs in differences_by_mo.values())

            # Write to Excel with legacy format
            if total_differences > 0:
                wb = openpyxl.Workbook()

                # Remove default sheet
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])

                # Create Summary sheet
                summary_ws = wb.create_sheet('Summary', 0)

                # Add metadata
                import os
                summary_ws['A1'] = 'Comparison Report'
                summary_ws['A1'].font = Font(bold=True, size=16)

                summary_ws['A3'] = 'Date:'
                summary_ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                summary_ws['A4'] = 'File 1:'
                summary_ws['B4'] = os.path.basename(self.xml1_path)

                summary_ws['A5'] = 'File 2:'
                summary_ws['B5'] = os.path.basename(self.xml2_path)

                summary_ws['A6'] = 'Total Differences:'
                summary_ws['B6'] = total_differences

                # MO Class summary table
                summary_ws['A8'] = 'MO Class'
                summary_ws['B8'] = 'Total Differences'
                summary_ws['C8'] = 'Modified'
                summary_ws['D8'] = 'Added'
                summary_ws['E8'] = 'Deleted'

                # Format header
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                for cell in ['A8', 'B8', 'C8', 'D8', 'E8']:
                    summary_ws[cell].fill = header_fill
                    summary_ws[cell].font = header_font

                # Add MO class statistics
                row = 9
                for mo_class in sorted(differences_by_mo.keys()):
                    diffs = differences_by_mo[mo_class]
                    modified = sum(1 for d in diffs if d['Status'] == 'MODIFIED')
                    added = sum(1 for d in diffs if d['Status'] == 'ADDED')
                    deleted = sum(1 for d in diffs if d['Status'] == 'DELETED')

                    summary_ws[f'A{row}'] = mo_class
                    summary_ws[f'B{row}'] = len(diffs)
                    summary_ws[f'C{row}'] = modified
                    summary_ws[f'D{row}'] = added
                    summary_ws[f'E{row}'] = deleted
                    row += 1

                # Adjust column widths
                summary_ws.column_dimensions['A'].width = 20
                summary_ws.column_dimensions['B'].width = 20
                summary_ws.column_dimensions['C'].width = 15
                summary_ws.column_dimensions['D'].width = 15
                summary_ws.column_dimensions['E'].width = 15

                # Create separate sheet for each MO class
                for mo_class in sorted(differences_by_mo.keys()):
                    # Sanitize sheet name (Excel has 31 char limit and doesn't allow certain chars)
                    sheet_name = mo_class[:31].replace('/', '_').replace('\\', '_').replace('*', '_').replace('?', '_').replace(':', '_').replace('[', '_').replace(']', '_')

                    ws = wb.create_sheet(sheet_name)

                    # Add headers
                    ws.append(['MO_Class', 'distName', 'Parameter', 'Value_XML1', 'Value_XML2', 'Status'])

                    # Format header
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font

                    # Add differences
                    for diff in differences_by_mo[mo_class]:
                        ws.append([
                            diff['MO_Class'],
                            diff['distName'],
                            diff['Parameter'],
                            diff['Value_XML1'],
                            diff['Value_XML2'],
                            diff['Status']
                        ])

                    # Color-code rows by status
                    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                        status = row[5].value
                        if status == 'MODIFIED':
                            fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                        elif status == 'ADDED':
                            fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif status == 'DELETED':
                            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        else:
                            continue

                        for cell in row:
                            cell.fill = fill

                    # Adjust column widths
                    ws.column_dimensions['A'].width = 20
                    ws.column_dimensions['B'].width = 40
                    ws.column_dimensions['C'].width = 30
                    ws.column_dimensions['D'].width = 20
                    ws.column_dimensions['E'].width = 20
                    ws.column_dimensions['F'].width = 15

                # Save workbook
                wb.save(self.output_excel)
            else:
                # No differences - create empty file
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Summary"
                ws['A1'] = 'Comparison Report'
                ws['A1'].font = Font(bold=True, size=16)
                ws['A3'] = 'No differences found'
                wb.save(self.output_excel)

            if progress_callback:
                progress_callback(100, "Complete!")

            return True, total_differences
            
        except Exception as e:
            error_msg = f"Error during comparison: {str(e)}"
            print(error_msg)
            import traceback
            traceback.print_exc()
            return False, 0
