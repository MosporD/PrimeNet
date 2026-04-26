"""
Scheduled Report Generation Routes
Generates Excel reports on demand or on a schedule.
"""

from flask import Blueprint, request, jsonify, render_template, redirect, url_for, send_file
from functools import wraps
import sqlite3
import os
import io
import math
import re
import json
import urllib.request
import urllib.parse
from datetime import datetime, timedelta
from itertools import combinations
from xml.sax.saxutils import escape as xml_escape

from database_enhanced import get_user_by_session, log_activity
from sync_config import NCMUSERS_DB, METADATA_DB, PROJECT_ROOT

reports_bp = Blueprint(
    'reports', __name__,
    template_folder='templates',
    static_folder='static',
    static_url_path='/reports/static',
)

REPORTS_DIR = os.path.join(PROJECT_ROOT, 'generated_reports')
os.makedirs(REPORTS_DIR, exist_ok=True)
CONFLICT_CACHE_TTL = timedelta(days=1)
_CONFLICT_CACHE: dict[str, dict] = {}


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.cookies.get('session_token')
        if not token:
            return redirect(url_for('auth.login_page'))
        user = get_user_by_session(token)
        if not user:
            return redirect(url_for('auth.login_page'))
        request.current_user = user
        return f(*args, **kwargs)
    return decorated


def get_current_user():
    token = request.cookies.get('session_token')
    return get_user_by_session(token) if token else None


def format_user(user):
    if not user:
        return None
    return {'id': user.get('id'), 'username': user.get('username'), 'role': user.get('role')}


def _ncm():
    conn = sqlite3.connect(NCMUSERS_DB)
    conn.row_factory = sqlite3.Row
    conn.execute('''
        CREATE TABLE IF NOT EXISTS elevation_cache (
            coord_key TEXT PRIMARY KEY,
            lat REAL NOT NULL,
            lng REAL NOT NULL,
            elevation_m REAL,
            updated_at TEXT NOT NULL
        )
    ''')
    conn.commit()
    return conn


def _meta():
    conn = sqlite3.connect(METADATA_DB, timeout=15)
    conn.execute('PRAGMA journal_mode=WAL')
    conn.row_factory = sqlite3.Row
    return conn


def _sql_ident(name: str) -> str:
    return '"' + str(name).replace('"', '""') + '"'


def _pick_col(candidates: list[str], low_to_real: dict[str, str]) -> str | None:
    for c in candidates:
        real = low_to_real.get(str(c).strip().lower())
        if real:
            return real
    return None


def _safe_float(v):
    try:
        if v is None:
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def _extract_coband_key(cell_name: str) -> str:
    """
    Coband key from the last numeric token in the cell name.
    Examples: ...L9 -> "9", ...L18 -> "18".
    """
    s = str(cell_name or '').strip()
    if not s:
        return ''
    hits = re.findall(r'(\d+)', s)
    return hits[-1] if hits else ''


def _haversine_km(lat1, lon1, lat2, lon2) -> float | None:
    lat1 = _safe_float(lat1)
    lon1 = _safe_float(lon1)
    lat2 = _safe_float(lat2)
    lon2 = _safe_float(lon2)
    if None in (lat1, lon1, lat2, lon2):
        return None
    r = 6371.0
    p1 = math.radians(lat1)
    p2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlambda / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


def _bearing_deg(lat1, lon1, lat2, lon2) -> float | None:
    lat1 = _safe_float(lat1)
    lon1 = _safe_float(lon1)
    lat2 = _safe_float(lat2)
    lon2 = _safe_float(lon2)
    if None in (lat1, lon1, lat2, lon2):
        return None
    p1 = math.radians(lat1)
    p2 = math.radians(lat2)
    dlambda = math.radians(lon2 - lon1)
    y = math.sin(dlambda) * math.cos(p2)
    x = math.cos(p1) * math.sin(p2) - math.sin(p1) * math.cos(p2) * math.cos(dlambda)
    b = math.degrees(math.atan2(y, x))
    return (b + 360.0) % 360.0


def _az_diff_deg(a, b) -> float | None:
    a = _safe_float(a)
    b = _safe_float(b)
    if a is None or b is None:
        return None
    d = abs((a - b) % 360.0)
    return min(d, 360.0 - d)


def _coord_key(lat: float, lng: float) -> str:
    return f'{float(lat):.5f},{float(lng):.5f}'


def _fetch_elevation_remote(points: list[tuple[float, float]]) -> dict[str, float | None]:
    """
    Query elevation service for a list of (lat, lng).
    Uses OpenTopoData (SRTM90m). Returns map by coord_key.
    """
    if not points:
        return {}
    out: dict[str, float | None] = {}
    batch_size = 80
    for i in range(0, len(points), batch_size):
        batch = points[i:i + batch_size]
        loc = '|'.join(f'{lat:.6f},{lng:.6f}' for lat, lng in batch)
        url = 'https://api.opentopodata.org/v1/srtm90m?locations=' + urllib.parse.quote(loc, safe='|,')
        try:
            req = urllib.request.Request(url, headers={'User-Agent': 'PrimeNet/1.0'})
            with urllib.request.urlopen(req, timeout=12) as resp:
                payload = json.loads(resp.read().decode('utf-8'))
            results = payload.get('results') if isinstance(payload, dict) else None
            if not isinstance(results, list):
                results = []
            for idx, item in enumerate(results):
                if idx >= len(batch):
                    break
                lat, lng = batch[idx]
                elev = None
                if isinstance(item, dict):
                    v = item.get('elevation')
                    try:
                        elev = float(v) if v is not None else None
                    except (TypeError, ValueError):
                        elev = None
                out[_coord_key(lat, lng)] = elev
        except Exception:
            for lat, lng in batch:
                out[_coord_key(lat, lng)] = None
    return out


def _elevation_for_points(points: list[tuple[float, float]]) -> dict[str, float | None]:
    """
    Resolve elevations with DB cache first, then remote fetch for misses.
    """
    unique = []
    seen = set()
    for lat, lng in points:
        if lat is None or lng is None:
            continue
        k = _coord_key(lat, lng)
        if k in seen:
            continue
        seen.add(k)
        unique.append((float(lat), float(lng)))
    if not unique:
        return {}

    conn = _ncm()
    cached: dict[str, float | None] = {}
    misses: list[tuple[float, float]] = []
    for lat, lng in unique:
        k = _coord_key(lat, lng)
        row = conn.execute('SELECT elevation_m FROM elevation_cache WHERE coord_key = ?', (k,)).fetchone()
        if row is not None:
            cached[k] = row['elevation_m']
        else:
            misses.append((lat, lng))

    fetched = _fetch_elevation_remote(misses)
    now = datetime.utcnow().isoformat() + 'Z'
    for lat, lng in misses:
        k = _coord_key(lat, lng)
        elev = fetched.get(k)
        cached[k] = elev
        conn.execute(
            'INSERT OR REPLACE INTO elevation_cache (coord_key, lat, lng, elevation_m, updated_at) VALUES (?, ?, ?, ?, ?)',
            (k, float(lat), float(lng), elev, now),
        )
    conn.commit()
    conn.close()
    return cached


def _normalize_conflict_tech(technology: str = '4G') -> str:
    t = str(technology or '4G').strip().upper()
    return t if t in ('3G', '4G', '5G') else '4G'


def _metadata_inventory_union_sql(conn: sqlite3.Connection) -> str:
    """
    Build a normalized UNION over per-technology metadata tables.
    We do not rely on the legacy `cells` table because recent loaders may keep
    data only in canonical tables (`cells_2g`, `cells_3g`, ...).
    """
    specs = [
        ('cells_2g', '2G', ['site_id', 'bcf id', 'bts id']),
        ('cells_3g', '3G', ['site_id', 'nodeb_id']),
        ('cells_4g_fdd', '4G-FDD', ['site_id', 'enb_id_actual', 'enodeb id']),
        ('cells_4g_tdd', '4G-TDD', ['site_id', 'enb_id_actual', 'enodeb id']),
        ('cells_5g', '5G', ['site_id', 'gnb_id_actual', 'gnb id']),
    ]

    parts: list[str] = []
    for table, tech, site_aliases in specs:
        cols = conn.execute(f'PRAGMA table_info({_sql_ident(table)})').fetchall()
        if not cols:
            continue
        low_to_real = {str(r[1]).strip().lower(): r[1] for r in cols}
        site_col = _pick_col(site_aliases, low_to_real)
        cell_col = _pick_col(
            ['cell_name', 'cell name', 'wcel name', 'lncel name', 'nrcel name', 'bts name'],
            low_to_real,
        )
        vendor_col = _pick_col(['vendor'], low_to_real)
        band_col = _pick_col(
            [
                'frequency_band', 'band',
                # LTE/NR identifiers
                'earfcn', 'nrarfcn', 'arfcn',
                # 3G channel identifiers (common naming variants)
                'uarfcn', 'uarfcn_dl', 'uarfcn downlink', 'dl_uarfcn', 'downlink_uarfcn',
                'uarfcn_downlink', 'downlink uarfcn', 'channel', 'channel_number',
            ],
            low_to_real,
        )
        az_col = _pick_col(['azimuth', 'azimuth_deg', 'azimuth degree'], low_to_real)
        pci_col = _pick_col(['pci', 'psc', 'bcch'], low_to_real)
        area_col = _pick_col(['area', 'region', 'market'], low_to_real)
        et_col = _pick_col(['electrical_tilt', 'electrical tilt', 'e_tilt'], low_to_real)
        mt_col = _pick_col(['mechanical_tilt', 'mechanical tilt', 'm_tilt'], low_to_real)
        status_col = _pick_col(['status', 'activity_status'], low_to_real)

        site_expr = f'CAST({_sql_ident(site_col)} AS TEXT)' if site_col else 'NULL'
        cell_expr = _sql_ident(cell_col) if cell_col else 'NULL'
        vendor_expr = _sql_ident(vendor_col) if vendor_col else 'NULL'
        band_expr = _sql_ident(band_col) if band_col else 'NULL'
        az_expr = _sql_ident(az_col) if az_col else 'NULL'
        pci_expr = _sql_ident(pci_col) if pci_col else 'NULL'
        area_expr = _sql_ident(area_col) if area_col else 'NULL'
        et_expr = _sql_ident(et_col) if et_col else 'NULL'
        mt_expr = _sql_ident(mt_col) if mt_col else 'NULL'
        status_expr = _sql_ident(status_col) if status_col else "''"

        parts.append(
            f"""
            SELECT
                {site_expr} AS site_id,
                {cell_expr} AS cell_name,
                '{tech}' AS technology,
                {vendor_expr} AS vendor,
                {band_expr} AS frequency_band,
                {az_expr} AS azimuth,
                {pci_expr} AS pci,
                {area_expr} AS area,
                {et_expr} AS electrical_tilt,
                {mt_expr} AS mechanical_tilt,
                {status_expr} AS status
            FROM {_sql_ident(table)}
            """
        )

    if not parts:
        # Keep query valid even if tables are missing.
        return (
            "SELECT NULL AS site_id, NULL AS cell_name, NULL AS technology, NULL AS vendor, "
            "NULL AS frequency_band, NULL AS azimuth, NULL AS pci, NULL AS area, NULL AS electrical_tilt, "
            "NULL AS mechanical_tilt, NULL AS status WHERE 1=0"
        )
    return "\nUNION ALL\n".join(parts)


# ── Page ──────────────────────────────────────────────────────────────────────

@reports_bp.route('/reports')
@login_required
def reports_page():
    user = get_current_user()
    return render_template('reports.html', user=format_user(user))


@reports_bp.route('/conflict-map')
@login_required
def conflict_map_page():
    user = get_current_user()
    return render_template('conflict_map.html', user=format_user(user))


# ── Report generators ─────────────────────────────────────────────────────────

def _generate_site_inventory(technology: str = 'all'):
    """Excel: one row per cell with dynamic full metadata columns."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError('openpyxl is required for report generation')

    conn = _meta()
    tech_key = (technology or 'all').strip().upper()
    tech_filters = {
        'ALL': None,
        '2G': ['2G'],
        '3G': ['3G'],
        '4G': ['4G-FDD', '4G-TDD'],
        '5G': ['5G'],
    }
    selected_techs = tech_filters.get(tech_key)
    if selected_techs is None and tech_key != 'ALL':
        selected_techs = None

    specs = [
        ('cells_2g', '2G', ['site_id', 'bcf id', 'bts id']),
        ('cells_3g', '3G', ['site_id', 'nodeb_id']),
        ('cells_4g_fdd', '4G-FDD', ['site_id', 'enb_id_actual', 'enodeb id']),
        ('cells_4g_tdd', '4G-TDD', ['site_id', 'enb_id_actual', 'enodeb id']),
        ('cells_5g', '5G', ['site_id', 'gnb_id_actual', 'gnb id']),
    ]

    site_cols_info = conn.execute('PRAGMA table_info("sites")').fetchall()
    site_cols = [r[1] for r in site_cols_info]
    site_rows = conn.execute('SELECT * FROM "sites"').fetchall() if site_cols else []
    site_id_col = _pick_col(['site_id'], {c.strip().lower(): c for c in site_cols}) if site_cols else None
    site_by_id: dict[str, dict] = {}
    if site_id_col:
        for sr in site_rows:
            sid = sr[site_id_col]
            if sid is None:
                continue
            site_by_id[str(sid)] = dict(sr)

    tech_records: dict[str, list[dict]] = {}
    tech_columns: dict[str, list[str]] = {}
    tech_seen_columns: dict[str, set[str]] = {}

    def _register_col(tech_name: str, col_name: str):
        if tech_name not in tech_columns:
            tech_columns[tech_name] = []
            tech_seen_columns[tech_name] = set()
        if col_name not in tech_seen_columns[tech_name]:
            tech_seen_columns[tech_name].add(col_name)
            tech_columns[tech_name].append(col_name)

    for table, tech, site_aliases in specs:
        if selected_techs and tech not in selected_techs:
            continue

        cols_info = conn.execute(f'PRAGMA table_info({_sql_ident(table)})').fetchall()
        if not cols_info:
            continue

        table_cols = [r[1] for r in cols_info]
        low_to_real = {str(c).strip().lower(): c for c in table_cols}
        site_col = _pick_col(site_aliases, low_to_real)
        cell_col = _pick_col(
            ['cell_name', 'cell name', 'wcel name', 'lncel name', 'nrcel name', 'bts name'],
            low_to_real,
        )
        status_col = _pick_col(['status', 'activity_status'], low_to_real)

        table_rows = conn.execute(f'SELECT * FROM {_sql_ident(table)}').fetchall()
        for tr in table_rows:
            if cell_col:
                cell_val = tr[cell_col]
                if cell_val is None or str(cell_val).strip() == '':
                    continue
            else:
                continue

            if status_col:
                status_val = tr[status_col]
                status_txt = '' if status_val is None else str(status_val).strip().lower()
                if status_txt not in ('', 'active'):
                    continue

            row_dict = dict(tr)
            record: dict = {'technology': tech}

            sid = str(row_dict.get(site_col)) if site_col and row_dict.get(site_col) is not None else None
            if sid is not None:
                record['site_id'] = sid

            if cell_col:
                record['cell_name'] = row_dict.get(cell_col)

            # Attach all site columns (if found) to enrich output.
            if sid is not None and sid in site_by_id:
                for sc, sv in site_by_id[sid].items():
                    if sc not in record:
                        record[sc] = sv
                    else:
                        record[f'site_{sc}'] = sv

            # Attach every metadata column from the selected tech table.
            for tc, tv in row_dict.items():
                if tc not in record:
                    record[tc] = tv
                else:
                    record[f'cell_{tc}'] = tv

            if tech not in tech_records:
                tech_records[tech] = []
                for base_col in ['site_id', 'site_name', 'technology', 'cell_name']:
                    _register_col(tech, base_col)

            for key in record.keys():
                _register_col(tech, key)
            tech_records[tech].append(record)

    conn.close()

    tech_label = tech_key if tech_key in tech_filters else 'ALL'
    pull_date = datetime.now().strftime('%Y-%m-%d')

    wb = Workbook()
    hdr_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True)
    ordered_techs = [t for _, t, _ in specs if t in tech_records]

    if not ordered_techs:
        ws = wb.active
        ws.title = 'Site Inventory'
        ws.cell(row=1, column=1, value='No inventory data found')
    else:
        # Use one sheet per technology when multiple technologies are included.
        for idx, tech in enumerate(ordered_techs):
            ws = wb.active if idx == 0 else wb.create_sheet()
            ws.title = tech[:31]

            cols_for_tech = tech_columns.get(tech, [])
            lead_cols = [c for c in ['site_id', 'site_name', 'technology', 'cell_name'] if c in cols_for_tech]
            other_cols = [c for c in cols_for_tech if c not in lead_cols]
            headers = lead_cols + other_cols

            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h.replace('_', ' ').title())
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal='center')

            for rec in tech_records.get(tech, []):
                ws.append([rec.get(h) for h in headers])

            # Auto-width per sheet
            for col in ws.columns:
                max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    total_rows = sum(len(v) for v in tech_records.values())
    return buf, f"Cells_{tech_label}_{pull_date}.xlsx", total_rows


def _generate_pci_conflicts(technology: str = '4G'):
    """Excel: co-band PCI conflict candidates using distance + azimuth logic (3G/4G/5G)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise RuntimeError('openpyxl required')

    tech_req, pair_rows = _build_conflict_pairs(technology)

    wb = Workbook()
    ws = wb.active
    ws.title = 'PCI Conflicts'

    headers = [
        'Risk',
        'PCI',
        'CoBand',
        'Distance_km',
        'Bearing_A_to_B_deg',
        'Bearing_B_to_A_deg',
        'Cell_A',
        'Site_A',
        'Area_A',
        'Cluster_A',
        'Azimuth_A',
        'A_to_B_Azimuth_Diff_deg',
        'Band_A',
        'Cell_B',
        'Site_B',
        'Area_B',
        'Cluster_B',
        'Azimuth_B',
        'B_to_A_Azimuth_Diff_deg',
        'Band_B',
    ]
    hdr_fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font

    for r in pair_rows:
        ws.append([
            r['risk'],
            r['pci'],
            r['coband'],
            r['distance_km'],
            r['bearing_ab'],
            r['bearing_ba'],
            r['a_name'],
            r['a_site'],
            r['a_area'],
            r['a_cluster'],
            r['a_az'],
            r['a_to_b_diff'],
            r['a_band'],
            r['b_name'],
            r['b_site'],
            r['b_area'],
            r['b_cluster'],
            r['b_az'],
            r['b_to_a_diff'],
            r['b_band'],
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, f"PCI_Conflicts_{tech_req}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx", len(pair_rows)


def _build_conflict_pairs(technology: str = '4G'):
    """Return normalized conflict pair rows for map/report usage."""
    conn = _meta()
    inv_union = _metadata_inventory_union_sql(conn)
    site_cols = conn.execute("PRAGMA table_info('sites')").fetchall()
    site_low_to_real = {str(r[1]).strip().lower(): r[1] for r in site_cols}
    site_area_col = _pick_col(['area', 'region', 'market'], site_low_to_real)
    site_cluster_col = _pick_col(['cluster', 'cluster_name'], site_low_to_real)
    site_area_expr = f"s.{_sql_ident(site_area_col)}" if site_area_col else "NULL"
    site_cluster_expr = f"s.{_sql_ident(site_cluster_col)}" if site_cluster_col else "NULL"

    tech_req = _normalize_conflict_tech(technology)
    if tech_req == '3G':
        tech_filter = ("3G",)
    elif tech_req == '4G':
        tech_filter = ("4G-FDD", "4G-TDD")
    else:
        tech_filter = ("5G",)
    filter_sql = ", ".join(["?"] * len(tech_filter))

    rows = conn.execute(f'''
        SELECT v.cell_name, v.technology, v.vendor, v.pci, v.azimuth, v.frequency_band,
               v.area AS cell_area,
               s.site_id, s.site_name, s.latitude, s.longitude,
               {site_area_expr} AS site_area,
               {site_cluster_expr} AS site_cluster
        FROM ({inv_union}) v
        LEFT JOIN sites s ON CAST(s.site_id AS TEXT) = CAST(v.site_id AS TEXT)
        WHERE v.cell_name IS NOT NULL
          AND TRIM(CAST(v.cell_name AS TEXT)) <> ''
          AND v.pci IS NOT NULL
          AND TRIM(CAST(v.pci AS TEXT)) <> ''
          AND (
                v.status IS NULL
                OR TRIM(COALESCE(v.status, '')) = ''
                OR LOWER(TRIM(COALESCE(v.status, ''))) = 'active'
          )
          AND v.technology IN ({filter_sql})
        ORDER BY v.pci, s.site_name
    ''', tech_filter).fetchall()
    conn.close()

    from collections import defaultdict
    groups = defaultdict(list)
    for r in rows:
        rd = dict(r)
        coband = _extract_coband_key(rd.get('cell_name'))
        if not coband:
            continue
        groups[(str(rd.get('pci')).strip(), coband)].append(rd)

    DIST_MAX_KM = 6.0
    DIST_HIGH_KM = 4.0
    AZ_NEAR_DEG = 50.0
    pair_rows = []
    for (pci, coband), grp in groups.items():
        if len({g['site_id'] for g in grp}) < 2:
            continue
        for a, b in combinations(grp, 2):
            if str(a.get('site_id') or '') == str(b.get('site_id') or ''):
                continue
            dist_km = _haversine_km(a.get('latitude'), a.get('longitude'), b.get('latitude'), b.get('longitude'))
            if dist_km is None or dist_km > DIST_MAX_KM:
                continue
            brg_ab = _bearing_deg(a.get('latitude'), a.get('longitude'), b.get('latitude'), b.get('longitude'))
            brg_ba = _bearing_deg(b.get('latitude'), b.get('longitude'), a.get('latitude'), a.get('longitude'))
            d_a = _az_diff_deg(a.get('azimuth'), brg_ab)
            d_b = _az_diff_deg(b.get('azimuth'), brg_ba)
            both_aligned = d_a is not None and d_b is not None and d_a <= AZ_NEAR_DEG and d_b <= AZ_NEAR_DEG
            one_aligned = (d_a is not None and d_a <= AZ_NEAR_DEG) or (d_b is not None and d_b <= AZ_NEAR_DEG)
            if both_aligned and dist_km <= DIST_HIGH_KM:
                risk = 'High'
            elif both_aligned or one_aligned:
                risk = 'Medium'
            else:
                risk = 'Low'

            pair_rows.append({
                'risk': risk,
                'pci': pci,
                'coband': coband,
                'technology': a.get('technology') or b.get('technology'),
                'distance_km': round(dist_km, 3),
                'bearing_ab': None if brg_ab is None else round(brg_ab, 1),
                'bearing_ba': None if brg_ba is None else round(brg_ba, 1),
                'a_name': a.get('cell_name'),
                'a_site': a.get('site_name') or a.get('site_id'),
                'a_az': a.get('azimuth'),
                'a_band': a.get('frequency_band'),
                'a_area': a.get('cell_area') or a.get('site_area'),
                'a_cluster': a.get('site_cluster'),
                'a_lat': _safe_float(a.get('latitude')),
                'a_lng': _safe_float(a.get('longitude')),
                'b_name': b.get('cell_name'),
                'b_site': b.get('site_name') or b.get('site_id'),
                'b_az': b.get('azimuth'),
                'b_band': b.get('frequency_band'),
                'b_area': b.get('cell_area') or b.get('site_area'),
                'b_cluster': b.get('site_cluster'),
                'b_lat': _safe_float(b.get('latitude')),
                'b_lng': _safe_float(b.get('longitude')),
                'a_to_b_diff': None if d_a is None else round(d_a, 1),
                'b_to_a_diff': None if d_b is None else round(d_b, 1),
            })

    risk_rank = {'High': 0, 'Medium': 1, 'Low': 2}
    pair_rows.sort(key=lambda r: (risk_rank.get(r['risk'], 9), r['distance_km'], r['pci']))
    return tech_req, pair_rows


def _get_cached_conflict_pairs(technology: str, force_refresh: bool = False):
    tech = _normalize_conflict_tech(technology)
    now = datetime.utcnow()
    cached = _CONFLICT_CACHE.get(tech)
    if (not force_refresh) and cached:
        gen = cached.get('generated_at')
        if isinstance(gen, datetime) and now - gen <= CONFLICT_CACHE_TTL:
            return tech, cached.get('rows', []), gen, False
    tech_req, rows = _build_conflict_pairs(tech)
    generated_at = datetime.utcnow()
    _CONFLICT_CACHE[tech_req] = {'rows': rows, 'generated_at': generated_at}
    return tech_req, rows, generated_at, True


def _kmlline_style_id(risk: str) -> str:
    r = str(risk or '').lower()
    if r == 'high':
        return 'risk-high'
    if r == 'medium':
        return 'risk-medium'
    return 'risk-low'


def _destination_point(lat: float, lng: float, bearing_deg: float, distance_km: float):
    r = 6371.0
    br = math.radians(bearing_deg)
    p1 = math.radians(lat)
    l1 = math.radians(lng)
    d = distance_km / r
    p2 = math.asin(math.sin(p1) * math.cos(d) + math.cos(p1) * math.sin(d) * math.cos(br))
    l2 = l1 + math.atan2(math.sin(br) * math.sin(d) * math.cos(p1), math.cos(d) - math.sin(p1) * math.sin(p2))
    return math.degrees(p2), math.degrees(l2)


def _wedge_polygon_coords(lat, lng, azimuth, width_deg=40.0, distance_km=0.8, segments=8):
    lat = _safe_float(lat)
    lng = _safe_float(lng)
    az = _safe_float(azimuth)
    if None in (lat, lng, az):
        return []
    start = az - (width_deg / 2.0)
    step = width_deg / max(1, segments)
    pts = [(lat, lng)]
    for i in range(segments + 1):
        b = start + (i * step)
        pts.append(_destination_point(lat, lng, b, distance_km))
    pts.append((lat, lng))
    return pts


def _generate_config_versions_report():
    """Excel: all config versions uploaded."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise RuntimeError('openpyxl required')

    conn = _ncm()
    rows = conn.execute('''
        SELECT cv.ne_name, cv.version_num, cv.file_name, cv.comment,
               cv.created_at, u.username as uploaded_by,
               LENGTH(cv.xml_content) as file_size_bytes
        FROM config_versions cv
        LEFT JOIN users u ON cv.uploaded_by = u.id
        ORDER BY cv.ne_name, cv.version_num
    ''').fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Config Versions'

    headers = ['NE Name', 'Version', 'File Name', 'Comment', 'Uploaded At', 'Uploaded By', 'Size (bytes)']
    hdr_fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font

    for r in rows:
        ws.append([r['ne_name'], r['version_num'], r['file_name'],
                   r['comment'], r['created_at'], r['uploaded_by'], r['file_size_bytes']])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, f"Config_Versions_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx", len(rows)


REPORT_TYPES = {
    'site_inventory':     ('Site Inventory',       _generate_site_inventory),
    'pci_conflicts':      ('Conflict Report',      _generate_pci_conflicts),
    'config_versions':    ('Configuration Log',   _generate_config_versions_report),
}


# ── API: generate on-demand ───────────────────────────────────────────────────

@reports_bp.route('/api/reports/generate', methods=['POST'])
@login_required
def generate_report():
    user = get_current_user()
    data = request.get_json()
    report_type = data.get('report_type', '')

    if report_type not in REPORT_TYPES:
        return jsonify({'error': f'Unknown report type: {report_type}'}), 400

    label, generator = REPORT_TYPES[report_type]
    try:
        if report_type == 'site_inventory':
            technology = str(data.get('technology', 'all') or 'all')
            buf, filename, row_count = generator(technology)
        elif report_type == 'pci_conflicts':
            technology = str(data.get('technology', '4G') or '4G')
            buf, filename, row_count = generator(technology)
        else:
            buf, filename, row_count = generator()
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    # Save to disk and record in archive
    file_path = os.path.join(REPORTS_DIR, filename)
    with open(file_path, 'wb') as fh:
        fh.write(buf.getvalue())

    conn = _ncm()
    conn.execute('''
        INSERT INTO report_archive (report_name, report_type, file_path, file_size, generated_by)
        VALUES (?, ?, ?, ?, ?)
    ''', (filename, report_type, file_path, os.path.getsize(file_path), user['id']))
    archive_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
    conn.commit()
    conn.close()

    log_activity(user['id'], 'report_generated', f'Generated {label} report ({row_count} rows)')
    return jsonify({'success': True, 'archive_id': archive_id, 'filename': filename, 'rows': row_count})


# ── API: download report ──────────────────────────────────────────────────────

@reports_bp.route('/api/reports/download/<int:report_id>')
@login_required
def download_report(report_id):
    conn = _ncm()
    row = conn.execute('SELECT * FROM report_archive WHERE id = ?', (report_id,)).fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'Report not found'}), 404
    row = dict(row)
    if not os.path.exists(row['file_path']):
        return jsonify({'error': 'Report file no longer exists on disk'}), 404
    return send_file(row['file_path'], as_attachment=True, download_name=row['report_name'],
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── API: report archive ───────────────────────────────────────────────────────

@reports_bp.route('/api/reports/archive')
@login_required
def report_archive():
    conn = _ncm()
    rows = conn.execute('''
        SELECT ra.id, ra.report_name, ra.report_type, ra.file_size, ra.generated_at,
               u.username as generated_by_name
        FROM report_archive ra
        LEFT JOIN users u ON ra.generated_by = u.id
        ORDER BY ra.generated_at DESC
        LIMIT 100
    ''').fetchall()
    conn.close()
    return jsonify({'success': True, 'reports': [dict(r) for r in rows]})


# ── API: delete report from archive ──────────────────────────────────────────

@reports_bp.route('/api/reports/archive/<int:report_id>', methods=['DELETE'])
@login_required
def delete_report(report_id):
    user = get_current_user()
    if not user or str(user.get('role', '')).strip().lower() != 'admin':
        return jsonify({'error': 'Admin privileges required to delete archived reports'}), 403

    conn = _ncm()
    row = conn.execute('SELECT * FROM report_archive WHERE id = ?', (report_id,)).fetchone()
    if row:
        try:
            if os.path.exists(row['file_path']):
                os.remove(row['file_path'])
        except OSError:
            pass
        conn.execute('DELETE FROM report_archive WHERE id = ?', (report_id,))
        conn.commit()
    conn.close()
    return jsonify({'success': True})


# ── API: list report types ────────────────────────────────────────────────────

@reports_bp.route('/api/reports/types')
@login_required
def report_types():
    return jsonify({
        'success': True,
        'types': [{'id': k, 'label': v[0]} for k, v in REPORT_TYPES.items()]
    })


@reports_bp.route('/api/conflict-map/data')
@login_required
def pci_conflicts_map_data():
    technology = str(request.args.get('technology', '4G') or '4G')
    risk = str(request.args.get('risk', 'all') or 'all').strip().lower()
    area_values = [str(v).strip() for v in request.args.getlist('area') if str(v).strip()]
    if not area_values:
        single = str(request.args.get('area', 'all') or 'all').strip()
        if single:
            area_values = [single]
    area_set = {v for v in area_values if v.lower() != 'all'}
    band = str(request.args.get('band', 'all') or 'all').strip()
    include_elevation = str(request.args.get('include_elevation', '0')).strip().lower() in ('1', 'true', 'yes', 'on')
    tech_req, rows, generated_at, refreshed = _get_cached_conflict_pairs(technology, force_refresh=False)

    def _row_match(r):
        if risk in ('high', 'medium', 'low') and str(r.get('risk', '')).lower() != risk:
            return False
        if area_set:
            area_a = str(r.get('a_area') or '')
            area_b = str(r.get('b_area') or '')
            if (area_a not in area_set) and (area_b not in area_set):
                return False
        if band.lower() != 'all':
            band_a = str(r.get('a_band') or '')
            band_b = str(r.get('b_band') or '')
            if band not in (band_a, band_b):
                return False
        return True

    filtered = [dict(r) for r in rows if _row_match(r)]
    areas = sorted({str(v) for r in rows for v in (r.get('a_area'), r.get('b_area')) if v})
    bands = sorted({str(v) for r in rows for v in (r.get('a_band'), r.get('b_band')) if v})

    if include_elevation and filtered:
        pts = []
        for r in filtered:
            if r.get('a_lat') is not None and r.get('a_lng') is not None:
                pts.append((float(r['a_lat']), float(r['a_lng'])))
            if r.get('b_lat') is not None and r.get('b_lng') is not None:
                pts.append((float(r['b_lat']), float(r['b_lng'])))
        elev_map = _elevation_for_points(pts)
        for r in filtered:
            a_e = elev_map.get(_coord_key(float(r['a_lat']), float(r['a_lng']))) if r.get('a_lat') is not None and r.get('a_lng') is not None else None
            b_e = elev_map.get(_coord_key(float(r['b_lat']), float(r['b_lng']))) if r.get('b_lat') is not None and r.get('b_lng') is not None else None
            r['a_elevation_m'] = a_e
            r['b_elevation_m'] = b_e
            if a_e is not None and b_e is not None:
                r['elevation_delta_m'] = round(float(a_e) - float(b_e), 1)
            else:
                r['elevation_delta_m'] = None

    return jsonify({
        'success': True,
        'technology': tech_req,
        'total': len(rows),
        'filtered_total': len(filtered),
        'filters': {
            'areas': areas,
            'bands': bands,
            'risk': ['High', 'Medium', 'Low'],
        },
        'cache': {
            'generated_at': generated_at.isoformat() + 'Z' if generated_at else None,
            'refreshed': refreshed,
        },
        'include_elevation': include_elevation,
        'rows': filtered,
    })


@reports_bp.route('/api/conflict-map/refresh', methods=['POST'])
@login_required
def refresh_conflict_map_data():
    payload = request.get_json(silent=True) or {}
    tech_req = str(payload.get('technology', 'all') or 'all').strip().upper()
    targets = ['3G', '4G', '5G'] if tech_req == 'ALL' else [_normalize_conflict_tech(tech_req)]
    out = {}
    for t in targets:
        _, rows, generated_at, _ = _get_cached_conflict_pairs(t, force_refresh=True)
        out[t] = {
            'rows': len(rows),
            'generated_at': generated_at.isoformat() + 'Z',
        }
    return jsonify({'success': True, 'refreshed': out})


@reports_bp.route('/api/conflict-map/export-kml')
@login_required
def export_conflict_map_kml():
    technology = str(request.args.get('technology', '4G') or '4G')
    risk = str(request.args.get('risk', 'all') or 'all').strip().lower()
    area_values = [str(v).strip() for v in request.args.getlist('area') if str(v).strip()]
    if not area_values:
        single = str(request.args.get('area', 'all') or 'all').strip()
        if single:
            area_values = [single]
    area_set = {v for v in area_values if v.lower() != 'all'}
    band = str(request.args.get('band', 'all') or 'all').strip()
    tech_req, rows, _, _ = _get_cached_conflict_pairs(technology, force_refresh=False)

    def _row_match(r):
        if risk in ('high', 'medium', 'low') and str(r.get('risk', '')).lower() != risk:
            return False
        if area_set:
            area_a = str(r.get('a_area') or '')
            area_b = str(r.get('b_area') or '')
            if (area_a not in area_set) and (area_b not in area_set):
                return False
        if band.lower() != 'all':
            band_a = str(r.get('a_band') or '')
            band_b = str(r.get('b_band') or '')
            if band not in (band_a, band_b):
                return False
        return True

    filtered = [r for r in rows if _row_match(r)]
    parts = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<kml xmlns="http://www.opengis.net/kml/2.2"><Document>',
        f'<name>{xml_escape(f"Conflict_Map_{tech_req}")}</name>',
        '<Style id="risk-high"><LineStyle><color>ff2b2bc0</color><width>3</width></LineStyle><PolyStyle><color>552b2bc0</color></PolyStyle></Style>',
        '<Style id="risk-medium"><LineStyle><color>ff12a3f3</color><width>3</width></LineStyle><PolyStyle><color>5512a3f3</color></PolyStyle></Style>',
        '<Style id="risk-low"><LineStyle><color>ffb98029</color><width>3</width></LineStyle><PolyStyle><color>55b98029</color></PolyStyle></Style>',
    ]

    for idx, r in enumerate(filtered, start=1):
        a_lat, a_lng = _safe_float(r.get('a_lat')), _safe_float(r.get('a_lng'))
        b_lat, b_lng = _safe_float(r.get('b_lat')), _safe_float(r.get('b_lng'))
        if None in (a_lat, a_lng, b_lat, b_lng):
            continue
        style_id = _kmlline_style_id(r.get('risk'))
        name = xml_escape(f"{r.get('risk', 'Risk')} | {r.get('pci', '')} | {r.get('a_site', '')} -> {r.get('b_site', '')}")
        desc = xml_escape(
            f"Technology: {r.get('technology', '-')}\n"
            f"Distance(km): {r.get('distance_km', '-')}\n"
            f"A: {r.get('a_name', '')} ({r.get('a_site', '')})\n"
            f"B: {r.get('b_name', '')} ({r.get('b_site', '')})"
        )
        parts.append(
            f'<Placemark><name>{name}</name><description>{desc}</description>'
            f'<styleUrl>#{style_id}</styleUrl><LineString><tessellate>1</tessellate><coordinates>{a_lng},{a_lat},0 {b_lng},{b_lat},0</coordinates></LineString></Placemark>'
        )

        for side in ('a', 'b'):
            wpts = _wedge_polygon_coords(r.get(f'{side}_lat'), r.get(f'{side}_lng'), r.get(f'{side}_az'))
            if not wpts:
                continue
            coords = ' '.join(f'{lng},{lat},0' for lat, lng in wpts)
            pname = xml_escape(f"Wedge {idx}-{side.upper()} {r.get(f'{side}_name', '')}")
            parts.append(
                f'<Placemark><name>{pname}</name><styleUrl>#{style_id}</styleUrl>'
                f'<Polygon><outerBoundaryIs><LinearRing><coordinates>{coords}</coordinates></LinearRing></outerBoundaryIs></Polygon></Placemark>'
            )

    parts.append('</Document></kml>')
    payload = '\n'.join(parts).encode('utf-8')
    return send_file(
        io.BytesIO(payload),
        as_attachment=True,
        download_name=f'Conflict_Map_{tech_req}_{datetime.now().strftime("%Y%m%d_%H%M")}.kml',
        mimetype='application/vnd.google-earth.kml+xml',
    )
