"""
Network Map Routes
Handles network visualization and KPI display
"""

from flask import Blueprint, request, jsonify, render_template, redirect, url_for
from functools import wraps
from collections import defaultdict
import math
import sqlite3
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from sync_config import NOKIA_PM_DB, HUAWEI_PM_DB, METADATA_DB, pm_table_name
from database_enhanced import get_user_by_session, log_activity
from sync.metadata_active_sql import (
    _STATUS_2G,
    _STATUS_3G_FDD,
    _STATUS_4G_FDD,
    _STATUS_4G_TDD,
    _STATUS_5G,
)


def _per_tech_union_sql(tech: str | None = None) -> tuple[str, list]:
    """
    Build a UNION ALL subquery over per-technology tables in metadata.db.
    Returns (sql, params) where sql is a SELECT ... UNION ALL ... string producing:
      cell_name, site_id, site_name, technology, vendor,
      latitude, longitude, azimuth, mechanical_tilt, electrical_tilt,
      frequency_band, pci, activity_status, status
    ``activity_status`` / ``status``: Active | Inactive from sync/metadata_active_sql.py
    (2G: Huawei active_state, Nokia admin_state; 3G/5G: Huawei activated / Nokia unlocked;
     4G-FDD: Nokia Unlocked or CELL_ACTIVE on active_state, Huawei activated or CELL_ACTIVE;
     4G-TDD: Huawei CELL_ACTIVE, Nokia admin_state Unlocked).
    """
    parts = {
        '2G': f"""
            SELECT
                cell_name,
                site_id      AS site_id,
                site_name    AS site_name,
                '2G'         AS technology,
                vendor,
                CAST(lat  AS REAL) AS latitude,
                CAST(long AS REAL) AS longitude,
                CAST(azimuth AS REAL) AS azimuth,
                CAST(mtilt   AS REAL) AS mechanical_tilt,
                CAST(etilt   AS REAL) AS electrical_tilt,
                frequency_band AS frequency_band,
                CAST(COALESCE(NULLIF(TRIM(bcch), ''), NULLIF(TRIM(bcc), '')) AS INTEGER) AS pci,
                {_STATUS_2G} AS activity_status,
                {_STATUS_2G} AS status
            FROM cells_2g
        """,
        '3G': f"""
            SELECT
                cell_name,
                nodeb_id     AS site_id,
                nodeb_name   AS site_name,
                '3G'         AS technology,
                vendor,
                CAST(lat  AS REAL) AS latitude,
                CAST(long AS REAL) AS longitude,
                CAST(azimuth AS REAL) AS azimuth,
                CAST(mtilt   AS REAL) AS mechanical_tilt,
                CAST(etilt   AS REAL) AS electrical_tilt,
                dl_uarfcn     AS frequency_band,
                CAST(psc AS INTEGER) AS pci,
                {_STATUS_3G_FDD} AS activity_status,
                {_STATUS_3G_FDD} AS status
            FROM cells_3g
        """,
        '4G-FDD': f"""
            SELECT
                cell_name,
                enb_id_actual AS site_id,
                enb_name      AS site_name,
                '4G-FDD'      AS technology,
                vendor,
                CAST(lat  AS REAL) AS latitude,
                CAST(long AS REAL) AS longitude,
                CAST(azimuth AS REAL) AS azimuth,
                CAST(mtilt   AS REAL) AS mechanical_tilt,
                CAST(etilt   AS REAL) AS electrical_tilt,
                band          AS frequency_band,
                CAST(pci AS INTEGER) AS pci,
                {_STATUS_4G_FDD} AS activity_status,
                {_STATUS_4G_FDD} AS status
            FROM cells_4g_fdd
        """,
        '4G-TDD': f"""
            SELECT
                cell_name,
                enb_id_actual AS site_id,
                enb_name      AS site_name,
                '4G-TDD'      AS technology,
                vendor,
                CAST(lat  AS REAL) AS latitude,
                CAST(long AS REAL) AS longitude,
                CAST(azimuth AS REAL) AS azimuth,
                CAST(mtilt   AS REAL) AS mechanical_tilt,
                CAST(etilt   AS REAL) AS electrical_tilt,
                band          AS frequency_band,
                CAST(pci AS INTEGER) AS pci,
                {_STATUS_4G_TDD} AS activity_status,
                {_STATUS_4G_TDD} AS status
            FROM cells_4g_tdd
        """,
        '5G': f"""
            SELECT
                cell_name,
                gnb_id_actual AS site_id,
                gnb_name      AS site_name,
                '5G'          AS technology,
                vendor,
                CAST(lat  AS REAL) AS latitude,
                CAST(long AS REAL) AS longitude,
                CAST(azimuth AS REAL) AS azimuth,
                CAST(mtilt   AS REAL) AS mechanical_tilt,
                CAST(etilt   AS REAL) AS electrical_tilt,
                bw            AS frequency_band,
                CAST(pci AS INTEGER) AS pci,
                {_STATUS_5G} AS activity_status,
                {_STATUS_5G} AS status
            FROM cells_5g
        """,
    }

    if tech and tech != 'all':
        # Tech can come from UI as '2G','3G','4G-FDD','4G-TDD','5G'
        if tech not in parts:
            return "SELECT NULL AS cell_name, NULL AS site_id, NULL AS site_name, NULL AS technology, NULL AS vendor, NULL AS latitude, NULL AS longitude, NULL AS azimuth, NULL AS mechanical_tilt, NULL AS electrical_tilt, NULL AS frequency_band, NULL AS pci, NULL AS activity_status, NULL AS status WHERE 1=0", []
        return parts[tech], []

    # All technologies
    return " UNION ALL ".join(parts[t] for t in ['2G', '3G', '4G-FDD', '4G-TDD', '5G']), []


# KML export: match map.js sector geometry (wedges) and tech colours
_KML_TECH_COLORS = {
    '2G': '#7f8c8d',
    '3G': '#27ae60',
    '4G-FDD': '#1a5276',
    '4G-TDD': '#148f77',
    '5G': '#9b59b6',
}
_KML_SECTOR_RADIUS_M = 600.0
_KML_SECTOR_BEAMWIDTH = 65.0


def _kml_color_aabbggrr(hex6: str, alpha: int = 255) -> str:
    """KML colour: aabbggrr (alpha, blue, green, red)."""
    h = hex6.lstrip('#')
    if len(h) != 6:
        return 'FF000000'
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return f'{max(0, min(255, alpha)):02X}{b:02X}{g:02X}{r:02X}'


def _kml_cell_active(row) -> bool:
    raw = (row['activity_status'] if row['activity_status'] is not None else row['status']) or ''
    return str(raw).strip().lower() == 'active'


def _kml_wedge_ring_coords(site_lat: float, site_lng: float, azimuth: float,
                           half_beam: float = _KML_SECTOR_BEAMWIDTH / 2.0,
                           radius_m: float = _KML_SECTOR_RADIUS_M,
                           step_deg: float = 3.0) -> list[tuple[float, float]]:
    """Closed (lon, lat) ring, same construction as network_map/static/map.js."""
    r_lat = radius_m / 111320.0
    c = math.cos(math.radians(site_lat))
    denom = 111320.0 * (c if abs(c) > 1e-9 else 1e-9)
    r_lng = radius_m / denom
    coords: list[tuple[float, float]] = [(site_lng, site_lat)]
    a = azimuth - half_beam
    end = azimuth + half_beam + 0.001
    while a <= end:
        rad = math.radians(a)
        lat = site_lat + r_lat * math.cos(rad)
        lng = site_lng + r_lng * math.sin(rad)
        coords.append((lng, lat))
        a += step_deg
    coords.append((site_lng, site_lat))
    return coords


def _kml_circle_ring_coords(lat0: float, lng0: float, radius_m: float, n: int = 36,
                            reverse: bool = False) -> list[tuple[float, float]]:
    """One closed ring (lon, lat); reverse=True for inner donut hole winding."""
    r_lat = radius_m / 111320.0
    c = math.cos(math.radians(lat0))
    r_lng = radius_m / (111320.0 * (c if abs(c) > 1e-9 else 1e-9))
    idx = range(n)
    if reverse:
        idx = range(n - 1, -1, -1)
    pts: list[tuple[float, float]] = []
    for i in idx:
        ang = 2 * math.pi * (i / n)
        lat = lat0 + r_lat * math.cos(ang)
        lng = lng0 + r_lng * math.sin(ang)
        pts.append((lng, lat))
    pts.append(pts[0])
    return pts


def _kml_ring_to_coordinates(coords: list[tuple[float, float]]) -> str:
    return ' '.join(f'{lng},{lat},0' for lng, lat in coords)


def _kml_cdata_fragment(t: str) -> str:
    """CDATA cannot contain the literal sequence ]]> ."""
    return (t or '').replace(']]>', ']] >')


def _kml_group_cells_into_wedges(cells: list) -> dict[tuple[str, float], list]:
    """Technology + 5° azimuth bucket → cells (same keying as map.js)."""
    by_key: dict[tuple[str, float], list] = defaultdict(list)
    for c in cells:
        if not _kml_cell_active(c):
            continue
        az = c['azimuth']
        if az is None:
            continue
        try:
            a = float(az)
        except (TypeError, ValueError):
            continue
        if not math.isfinite(a):
            continue
        bucket = round(a / 5.0) * 5.0
        tech = c['technology'] or 'Unknown'
        by_key[(tech, bucket)].append(c)
    return by_key


def _metadata_table_for_tech(tech: str) -> str | None:
    mapping = {
        '2G': 'cells_2g',
        '3G': 'cells_3g',
        '4G-FDD': 'cells_4g_fdd',
        '4G-TDD': 'cells_4g_tdd',
        '5G': 'cells_5g',
    }
    return mapping.get((tech or '').strip())

network_map_bp = Blueprint(
    'network_map', __name__,
    template_folder='templates',
    static_folder='static',
    static_url_path='/network_map/static',
)


def login_required(f):
    """Decorator to require login"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        session_token = request.cookies.get('session_token')
        if not session_token:
            return redirect(url_for('auth.login_page'))

        user = get_user_by_session(session_token)
        if not user:
            return redirect(url_for('auth.login_page'))

        request.current_user = user
        return f(*args, **kwargs)

    return decorated_function

def get_current_user():
    """Get current logged-in user"""
    session_token = request.cookies.get('session_token')
    if session_token:
        return get_user_by_session(session_token)
    return None

def format_user_data(user):
    """Format user data for templates"""
    if not user:
        return None
    if isinstance(user, dict):
        return {'username': user.get('username'), 'email': user.get('email'), 'role': user.get('role'), 'id': user.get('id')}
    return {'username': (user.get('username') if isinstance(user, dict) else user[1]), 'email': (user.get('email') if isinstance(user, dict) else user[2]), 'role': (user.get('role') if isinstance(user, dict) else user[6]), 'id': (user.get('id') if isinstance(user, dict) else user[0])}


@network_map_bp.route('/network-map')
@login_required
def network_map_page():
    """Render Network Map page"""
    user = get_current_user()
    return render_template('network_map.html', user=format_user_data(user))

@network_map_bp.route('/api/map/sites', methods=['GET'])
def get_all_sites():
    """Get all network sites, optionally filtered by technology.

    IMPORTANT: For Network Map we use the per-technology tables (cells_2g/3g/4g_fdd/4g_tdd/5g)
    so counts match the CSV snapshots (no de-duplication via the unified cells table).
    """
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    tech = request.args.get('tech', '').strip()
    tech_value = request.args.get('tech_value', '').strip()

    try:
        conn = sqlite3.connect(METADATA_DB)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        union_sql, params = _per_tech_union_sql(tech if tech else None)
        scope_clause = ''
        sql_params = list(params)
        if tech_value:
            if tech == '2G':
                scope_clause = ' AND CAST(v.pci AS TEXT) = ?'
            else:
                scope_clause = ' AND CAST(v.frequency_band AS TEXT) = ?'
            sql_params.append(tech_value)

        cursor.execute(f'''
            WITH v AS (
                {union_sql}
            ),
            filtered AS (
                SELECT *
                FROM v
                WHERE 1=1
                  {scope_clause}
            ),
            offline AS (
                SELECT
                    site_id,
                    SUM(
                        CASE WHEN LOWER(TRIM(COALESCE(activity_status, status, ''))) = 'inactive'
                             THEN 1 ELSE 0 END
                    ) AS offline_cell_count
                FROM filtered
                GROUP BY site_id
            ),
            wedge_agg AS (
                SELECT
                    site_id,
                    technology,
                    ROUND(CAST(azimuth AS REAL) / 5.0) * 5 AS az_bin,
                    COUNT(*) AS cell_count,
                    SUM(
                        CASE WHEN LOWER(TRIM(COALESCE(activity_status, status, ''))) = 'inactive'
                             THEN 1 ELSE 0 END
                    ) AS offline_count
                FROM filtered
                WHERE azimuth IS NOT NULL
                  AND TRIM(CAST(azimuth AS TEXT)) <> ''
                GROUP BY site_id, technology,
                         ROUND(CAST(azimuth AS REAL) / 5.0) * 5
            ),
            full_sector_off AS (
                SELECT site_id, COUNT(*) AS full_sector_offline_count
                FROM wedge_agg
                WHERE cell_count > 0 AND offline_count = cell_count
                GROUP BY site_id
            )
            SELECT DISTINCT
                s.site_id, s.site_name, s.latitude, s.longitude,
                s.region, s.site_type, s.vendor, s.status,
                COALESCE(o.offline_cell_count, 0) AS offline_cell_count,
                COALESCE(fs.full_sector_offline_count, 0) AS full_sector_offline_count
            FROM filtered f
            JOIN sites s ON s.site_id = f.site_id
            LEFT JOIN offline o ON o.site_id = s.site_id
            LEFT JOIN full_sector_off fs ON fs.site_id = s.site_id
            WHERE s.latitude IS NOT NULL AND s.longitude IS NOT NULL
            ORDER BY s.site_name
        ''', sql_params)

        sites = [dict(row) for row in cursor.fetchall()]
        conn.close()

        log_activity((user.get('id') if isinstance(user, dict) else user[0]), 'map_view', 'Viewed network map sites')
        return jsonify({'success': True, 'sites': sites})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@network_map_bp.route('/api/map/tech-filter-options', methods=['GET'])
def get_tech_filter_options():
    """Get dynamic dropdown values for tech-specific filters (UARFCN/Band/BCCH)."""
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    tech = request.args.get('tech', '').strip()
    if tech not in ('2G', '3G', '4G-FDD', '4G-TDD'):
        return jsonify({'success': True, 'label': '', 'values': []})

    try:
        conn = sqlite3.connect(METADATA_DB)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        union_sql, params = _per_tech_union_sql(tech)

        if tech == '2G':
            rows = cursor.execute(f'''
                SELECT DISTINCT CAST(v.pci AS TEXT) AS val
                FROM ({union_sql}) v
                WHERE v.pci IS NOT NULL AND TRIM(CAST(v.pci AS TEXT)) <> ''
                ORDER BY CAST(v.pci AS INTEGER)
            ''', params).fetchall()
            label = 'BCCH'
        elif tech == '3G':
            rows = cursor.execute(f'''
                SELECT DISTINCT CAST(v.frequency_band AS TEXT) AS val
                FROM ({union_sql}) v
                WHERE v.frequency_band IS NOT NULL AND TRIM(CAST(v.frequency_band AS TEXT)) <> ''
                ORDER BY CAST(v.frequency_band AS INTEGER), v.frequency_band
            ''', params).fetchall()
            label = 'UARFCN'
        else:
            rows = cursor.execute(f'''
                SELECT DISTINCT CAST(v.frequency_band AS TEXT) AS val
                FROM ({union_sql}) v
                WHERE v.frequency_band IS NOT NULL AND TRIM(CAST(v.frequency_band AS TEXT)) <> ''
                ORDER BY v.frequency_band
            ''', params).fetchall()
            label = 'Band'

        conn.close()
        values = [r['val'] for r in rows if r['val'] is not None]
        return jsonify({'success': True, 'label': label, 'values': values})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@network_map_bp.route('/api/map/site/<site_id>', methods=['GET'])
def get_site_details(site_id):
    """Get detailed information about a specific site"""
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        conn = sqlite3.connect(METADATA_DB)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute('''
            SELECT site_id, site_name, latitude, longitude, region, site_type, vendor, status
            FROM sites
            WHERE site_id = ?
        ''', (site_id,))

        site = cursor.fetchone()
        if not site:
            conn.close()
            return jsonify({'error': 'Site not found'}), 404

        site_data = dict(site)

        # Cells are the sectors — pull from per-technology tables (no de-dupe).
        union_sql, params = _per_tech_union_sql(None)
        cursor.execute(f'''
            SELECT
                cell_name,
                technology,
                vendor,
                frequency_band,
                azimuth,
                mechanical_tilt,
                electrical_tilt,
                pci,
                activity_status,
                status
            FROM ({union_sql}) v
            WHERE v.site_id = ?
            ORDER BY technology, cell_name
        ''', params + [site_id])
        cells = []
        for row in cursor.fetchall():
            d = dict(row)
            # Map JS expects an identifier; per-tech tables don't have integer ids.
            d['cell_id'] = d.get('cell_name')  # backward-compatible field name
            cells.append(d)
        site_data['cells'] = cells

        conn.close()

        log_activity((user.get('id') if isinstance(user, dict) else user[0]), 'site_view', f'Viewed site {site_id}')
        return jsonify({'success': True, 'site': site_data})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@network_map_bp.route('/api/map/cell/<int:cell_id>/kpis', methods=['GET'])
def get_cell_kpis(cell_id):
    """Get the latest KPI snapshot for a cell from the appropriate PM database."""
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        meta_conn = sqlite3.connect(METADATA_DB)
        meta_conn.row_factory = sqlite3.Row
        cell = meta_conn.execute('''
            SELECT c.cell_id, c.cell_name, c.technology, c.vendor,
                   c.azimuth, c.mechanical_tilt, c.electrical_tilt, c.pci,
                   st.site_id, st.site_name, st.region
            FROM cells c
            LEFT JOIN sites st ON c.site_id = st.site_id
            WHERE c.cell_id = ?
        ''', (cell_id,)).fetchone()
        meta_conn.close()

        if not cell:
            return jsonify({'error': 'Cell not found'}), 404

        cell_data = dict(cell)
        vendor    = cell_data.get('vendor', '')
        cell_name = cell_data['cell_name']
        cell_tech = cell_data.get('technology', '4G')
        pm_db     = NOKIA_PM_DB if vendor == 'Nokia' else HUAWEI_PM_DB
        if vendor == 'Huawei':
            from sync.pm_processor import huawei_pm_table_for_cell

            table = huawei_pm_table_for_cell(cell_name, cell_tech, pm_db)
        else:
            table = pm_table_name(cell_tech)

        try:
            if not table:
                cell_data['kpis'] = None
            else:
                pm_conn = sqlite3.connect(pm_db)
                pm_conn.row_factory = sqlite3.Row
                kpi = pm_conn.execute(f'''
                    SELECT *
                    FROM "{table}"
                    WHERE cell_name = ?
                    ORDER BY timestamp DESC
                    LIMIT 1
                ''', (cell_name,)).fetchone()
                pm_conn.close()
                cell_data['kpis'] = dict(kpi) if kpi else None
        except sqlite3.OperationalError:
            cell_data['kpis'] = None

        log_activity((user.get('id') if isinstance(user, dict) else user[0]),
                     'cell_kpi_view', f'Viewed KPIs for cell {cell_name}')
        return jsonify({'success': True, 'cell': cell_data})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@network_map_bp.route('/api/map/cell/kpis', methods=['GET'])
def get_cell_kpis_by_name():
    """
    KPI endpoint keyed by cell_name (used by the Network Map per-tech tables).
    Query param: ?cell_name=<name>
    """
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    cell_name = (request.args.get('cell_name') or '').strip()
    if not cell_name:
        return jsonify({'error': 'cell_name is required'}), 400

    try:
        conn = sqlite3.connect(METADATA_DB)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        union_sql, params = _per_tech_union_sql(None)
        row = cur.execute(f'''
            SELECT
                v.cell_name, v.technology, v.vendor,
                v.frequency_band, v.azimuth, v.mechanical_tilt, v.electrical_tilt,
                v.pci, v.activity_status, v.status,
                s.site_id, s.site_name, s.region
            FROM ({union_sql}) v
            LEFT JOIN sites s ON s.site_id = v.site_id
            WHERE v.cell_name = ?
            LIMIT 1
        ''', params + [cell_name]).fetchone()
        conn.close()

        if not row:
            return jsonify({'error': 'Cell not found'}), 404

        cell_data = dict(row)

        vendor = cell_data.get('vendor', '')
        tech   = cell_data.get('technology', '4G-FDD')
        pm_db  = NOKIA_PM_DB if vendor == 'Nokia' else HUAWEI_PM_DB
        if vendor == 'Huawei':
            from sync.pm_processor import huawei_pm_table_for_cell

            table = huawei_pm_table_for_cell(cell_name, tech, pm_db)
        else:
            table = pm_table_name(tech)

        # Add full metadata row from the technology-specific staging table so the
        # wedge-cell popup can show all available fields for that exact cell.
        cell_data['metadata'] = None
        meta_table = _metadata_table_for_tech(tech)
        if meta_table:
            try:
                meta_conn = sqlite3.connect(METADATA_DB)
                meta_conn.row_factory = sqlite3.Row
                md = meta_conn.execute(
                    f'SELECT * FROM "{meta_table}" WHERE cell_name = ? LIMIT 1',
                    (cell_name,)
                ).fetchone()
                meta_conn.close()
                cell_data['metadata'] = dict(md) if md else None
            except sqlite3.OperationalError:
                cell_data['metadata'] = None

        try:
            if not table:
                cell_data['kpis'] = None
            else:
                pm_conn = sqlite3.connect(pm_db)
                pm_conn.row_factory = sqlite3.Row
                kpi = pm_conn.execute(f'''
                    SELECT *
                    FROM "{table}"
                    WHERE cell_name = ?
                    ORDER BY timestamp DESC
                    LIMIT 1
                ''', (cell_name,)).fetchone()
                pm_conn.close()
                cell_data['kpis'] = dict(kpi) if kpi else None
        except sqlite3.OperationalError:
            cell_data['kpis'] = None

        log_activity((user.get('id') if isinstance(user, dict) else user[0]),
                     'cell_kpi_view', f'Viewed KPIs for cell {cell_name}')
        return jsonify({'success': True, 'cell': cell_data})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@network_map_bp.route('/api/map/stats', methods=['GET'])
def get_network_stats():
    """Get overall network statistics for Network Map (per-tech tables)."""
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        conn = sqlite3.connect(METADATA_DB)
        cursor = conn.cursor()

        # Count per-tech rows directly so numbers match CSV snapshots.
        tech_counts = {}
        for tech, table in [
            ('2G', 'cells_2g'),
            ('3G', 'cells_3g'),
            ('4G-FDD', 'cells_4g_fdd'),
            ('4G-TDD', 'cells_4g_tdd'),
            ('5G', 'cells_5g'),
        ]:
            try:
                tech_counts[tech] = cursor.execute(f'SELECT COUNT(*) FROM {table}').fetchone()[0]
            except sqlite3.OperationalError:
                tech_counts[tech] = 0

        total_cells = sum(tech_counts.values())

        # Sites shown on map are those with coordinates AND at least one cell row in the per-tech tables.
        union_sql, params = _per_tech_union_sql(None)
        total_sites = cursor.execute(f'''
            SELECT COUNT(DISTINCT s.site_id)
            FROM ({union_sql}) v
            JOIN sites s ON s.site_id = v.site_id
            WHERE s.latitude IS NOT NULL AND s.longitude IS NOT NULL
        ''', params).fetchone()[0]

        conn.close()

        return jsonify({'success': True, 'stats': {
            'total_sites':   total_sites,
            'total_sectors': total_cells,
            'total_cells':   total_cells,
            'tech_counts':   tech_counts,
        }})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@network_map_bp.route('/api/map/search/cell-code', methods=['GET'])
def search_by_cell_code():
    """Search cells by PCI (4G/5G), PSC (3G), or BCCH (2G) using per-tech tables."""
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    code = request.args.get('code', '').strip()
    tech = request.args.get('tech', '').strip()

    if not code or not code.lstrip('-').isdigit():
        return jsonify({'success': True, 'matches': []})

    try:
        conn = sqlite3.connect(METADATA_DB)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        union_sql, u_params = _per_tech_union_sql(tech if tech else None)
        cursor.execute(f'''
            SELECT
                v.cell_name,
                v.technology,
                v.vendor,
                v.frequency_band,
                v.azimuth,
                v.mechanical_tilt,
                v.electrical_tilt,
                v.pci,
                v.activity_status,
                v.status,
                s.site_id,
                s.site_name,
                s.latitude,
                s.longitude
            FROM ({union_sql}) v
            JOIN sites s ON s.site_id = v.site_id
            WHERE v.pci = ?
              AND s.latitude IS NOT NULL AND s.longitude IS NOT NULL
            ORDER BY s.site_name, v.technology, v.cell_name
        ''', u_params + [int(code)])

        matches = []
        for row in cursor.fetchall():
            d = dict(row)
            d['cell_id'] = d.get('cell_name')  # map.js expects an id-like field
            matches.append(d)
        conn.close()

        return jsonify({'success': True, 'matches': matches})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@network_map_bp.route('/api/map/export/cell-code', methods=['GET'])
@login_required
def export_cell_code():
    """Export cells matching a PCI / SC / BCCH code as an Excel file."""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import send_file

    code = request.args.get('code', '').strip()
    tech = request.args.get('tech', '').strip()

    if not code or not code.lstrip('-').isdigit():
        return jsonify({'error': 'Invalid code'}), 400

    try:
        conn = sqlite3.connect(METADATA_DB)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        union_sql, u_params = _per_tech_union_sql(tech if tech else None)
        rows = cursor.execute(f'''
            SELECT
                v.cell_name,
                s.site_name,
                s.site_id,
                v.technology,
                v.vendor,
                v.frequency_band,
                v.pci,
                v.azimuth,
                v.mechanical_tilt,
                v.electrical_tilt,
                v.activity_status,
                s.latitude,
                s.longitude
            FROM ({union_sql}) v
            JOIN sites s ON s.site_id = v.site_id
            WHERE v.pci = ?
              AND s.latitude IS NOT NULL AND s.longitude IS NOT NULL
            ORDER BY s.site_name, v.technology, v.cell_name
        ''', u_params + [int(code)]).fetchall()
        conn.close()

        code_label = 'PSC' if tech == '3G' else ('BCCH' if tech == '2G' else 'PCI')

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = f'{code_label}_{code}'

        headers = ['Cell Name', 'Site Name', 'Site ID', 'Technology', 'Vendor',
                   'Band', code_label, 'Azimuth (°)', 'M.Tilt (°)', 'E.Tilt (°)',
                   'Activity status', 'Latitude', 'Longitude']
        ws.append(headers)

        hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        for cell in ws[1]:
            cell.font      = Font(bold=True, color='FFFFFF')
            cell.fill      = hdr_fill
            cell.alignment = Alignment(horizontal='center')

        for row in rows:
            ws.append([
                row['cell_name'], row['site_name'], row['site_id'],
                row['technology'], row['vendor'], row['frequency_band'],
                row['pci'], row['azimuth'], row['mechanical_tilt'],
                row['electrical_tilt'], row['activity_status'],
                row['latitude'], row['longitude']
            ])

        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        uid = (request.current_user.get('id')
               if isinstance(request.current_user, dict)
               else request.current_user[0])
        log_activity(uid, 'export', f'Exported {len(rows)} cells with {code_label}={code}')

        return send_file(
            buf,
            download_name=f'{code_label}_{code}_cells.xlsx',
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@network_map_bp.route('/api/map/export/sites', methods=['GET'])
@login_required
def export_sites_excel():
    """Export all sites + cells for the current filter as an Excel file (two sheets)."""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import send_file
    from datetime import datetime as dt

    tech       = request.args.get('tech',       '').strip()
    tech_value = request.args.get('tech_value', '').strip()
    vendor     = request.args.get('vendor',     '').strip()
    search     = request.args.get('search',     '').strip()

    try:
        conn = sqlite3.connect(METADATA_DB)
        conn.row_factory = sqlite3.Row

        # ── Sites ─────────────────────────────────────────────────────────────
        s_cond   = ['s.latitude IS NOT NULL', 's.longitude IS NOT NULL']
        s_params = []
        if vendor:
            s_cond.append('s.vendor = ?');  s_params.append(vendor)
        if search:
            s_cond.append('(s.site_name LIKE ? OR CAST(s.site_id AS TEXT) LIKE ?)')
            s_params += [f'%{search}%', f'%{search}%']
        if tech_value:
            if tech == '2G':
                s_cond.append('CAST(v.pci AS TEXT) = ?')
            else:
                s_cond.append('CAST(v.frequency_band AS TEXT) = ?')
            s_params.append(tech_value)

        union_sql, u_params = _per_tech_union_sql(tech if tech else None)
        # Sites are those matching the filter and having at least one cell row.
        sites = conn.execute(f'''
            SELECT DISTINCT
                s.site_id, s.site_name, s.latitude, s.longitude,
                s.region, s.site_type, s.vendor, s.status
            FROM ({union_sql}) v
            JOIN sites s ON s.site_id = v.site_id
            WHERE {" AND ".join(s_cond)}
            ORDER BY s.site_name
        ''', u_params + s_params).fetchall()

        # ── Cells ─────────────────────────────────────────────────────────────
        c_cond, c_params = [], []
        if vendor:
            c_cond.append('s.vendor = ?');      c_params.append(vendor)
        if search:
            c_cond.append('(s.site_name LIKE ? OR CAST(s.site_id AS TEXT) LIKE ?)')
            c_params += [f'%{search}%', f'%{search}%']
        if tech_value:
            if tech == '2G':
                c_cond.append('CAST(v.pci AS TEXT) = ?')
            else:
                c_cond.append('CAST(v.frequency_band AS TEXT) = ?')
            c_params.append(tech_value)
        c_where = ('WHERE ' + ' AND '.join(c_cond)) if c_cond else ''
        union_sql, u_params = _per_tech_union_sql(tech if tech else None)
        cells = conn.execute(f'''
            SELECT
                v.cell_name,
                v.site_id,
                s.site_name,
                v.technology,
                v.frequency_band,
                v.azimuth,
                v.mechanical_tilt,
                v.electrical_tilt,
                v.pci,
                v.activity_status,
                v.status,
                v.vendor,
                s.latitude,
                s.longitude
            FROM ({union_sql}) v
            JOIN sites s ON s.site_id = v.site_id
            {c_where}
            ORDER BY s.site_name, v.technology, v.cell_name
        ''', u_params + c_params).fetchall()
        conn.close()

        # ── Build workbook ────────────────────────────────────────────────────
        HDR_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        HDR_FONT = Font(bold=True, color='FFFFFF')
        CTR      = Alignment(horizontal='center')

        def _style_header(ws):
            for cell in ws[1]:
                cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.alignment = CTR

        def _autofit(ws):
            for col in ws.columns:
                w = max((len(str(c.value or '')) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(w + 4, 45)

        wb   = openpyxl.Workbook()
        ws_s = wb.active
        ws_s.title = 'Sites'
        ws_s.append(['Site ID', 'Site Name', 'Latitude', 'Longitude',
                     'Region', 'Site Type', 'Vendor', 'Status'])
        _style_header(ws_s)
        for r in sites:
            ws_s.append([r['site_id'], r['site_name'], r['latitude'], r['longitude'],
                         r['region'], r['site_type'], r['vendor'], r['status']])
        _autofit(ws_s)

        ws_c = wb.create_sheet('Cells')
        ws_c.append(['Cell Name', 'Site ID', 'Site Name', 'Technology', 'Frequency Band',
                     'Azimuth (°)', 'Mech. Tilt (°)', 'Elec. Tilt (°)',
                     'PCI / SC / BCCH', 'Activity status', 'Vendor', 'Latitude', 'Longitude'])
        _style_header(ws_c)
        for r in cells:
            ws_c.append([r['cell_name'], r['site_id'], r['site_name'], r['technology'],
                         r['frequency_band'], r['azimuth'], r['mechanical_tilt'],
                         r['electrical_tilt'], r['pci'],
                         r['activity_status'] or r['status'], r['vendor'],
                         r['latitude'], r['longitude']])
        _autofit(ws_c)

        buf = BytesIO(); wb.save(buf); buf.seek(0)
        tech_label = (tech or 'All').replace('-', '_')
        fname = f'network_export_{tech_label}_{dt.now().strftime("%Y%m%d_%H%M")}.xlsx'
        uid = (request.current_user.get('id')
               if isinstance(request.current_user, dict) else request.current_user[0])
        log_activity(uid, 'export',
                     f'Excel export: tech={tech or "All"}, {len(sites)} sites, {len(cells)} cells')
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@network_map_bp.route('/api/map/export/kml', methods=['GET'])
@login_required
def export_sites_kml():
    """Export sites as a KML file for Google Earth."""
    from flask import Response
    from datetime import datetime as dt
    import xml.sax.saxutils as sx

    tech       = request.args.get('tech',       '').strip()
    tech_value = request.args.get('tech_value', '').strip()
    vendor     = request.args.get('vendor',     '').strip()
    search     = request.args.get('search',     '').strip()

    try:
        conn = sqlite3.connect(METADATA_DB)
        conn.row_factory = sqlite3.Row

        s_cond   = ['s.latitude IS NOT NULL', 's.longitude IS NOT NULL']
        s_params = []
        if vendor:
            s_cond.append('s.vendor = ?');  s_params.append(vendor)
        if search:
            s_cond.append('(s.site_name LIKE ? OR CAST(s.site_id AS TEXT) LIKE ?)')
            s_params += [f'%{search}%', f'%{search}%']
        if tech_value:
            if tech == '2G':
                s_cond.append('CAST(v.pci AS TEXT) = ?')
            else:
                s_cond.append('CAST(v.frequency_band AS TEXT) = ?')
            s_params.append(tech_value)
        union_sql, u_params = _per_tech_union_sql(tech if tech else None)
        sites = conn.execute(f'''
            SELECT DISTINCT
                s.site_id, s.site_name, s.latitude, s.longitude,
                s.region, s.vendor, s.status
            FROM ({union_sql}) v
            JOIN sites s ON s.site_id = v.site_id
            WHERE {" AND ".join(s_cond)}
            ORDER BY s.site_name
        ''', u_params + s_params).fetchall()

        # Fetch cells for each site to populate the description balloon
        cells_by_site = {}
        if sites:
            ids = [r['site_id'] for r in sites]
            ph  = ','.join('?' * len(ids))
            union_sql, u_params = _per_tech_union_sql(tech if tech else None)
            for r in conn.execute(f'''
                SELECT
                    v.site_id,
                    v.cell_name,
                    v.technology,
                    v.azimuth,
                    v.frequency_band,
                    v.pci,
                    v.activity_status,
                    v.status
                FROM ({union_sql}) v
                WHERE v.site_id IN ({ph})
                  {("AND CAST(v.pci AS TEXT) = ?" if tech_value and tech == "2G" else "")}
                  {("AND CAST(v.frequency_band AS TEXT) = ?" if tech_value and tech != "2G" else "")}
                ORDER BY v.technology, v.cell_name
            ''', u_params + ids + ([tech_value] if tech_value else [])).fetchall():
                cells_by_site.setdefault(r['site_id'], []).append(r)
        conn.close()

        # ── Build KML (site = donut polygon; active cells = sector wedges) ───
        tech_label = tech or 'All Technologies'
        lines = [
            '<?xml version="1.0" encoding="UTF-8"?>',
            '<kml xmlns="http://www.opengis.net/kml/2.2">',
            '<Document>',
            f'<name>{sx.escape(tech_label)} Network Export</name>',
            f'<description>Exported {dt.now().strftime("%Y-%m-%d %H:%M")} — sites as rings, active sectors as wedges.</description>',
            '<Style id="site-donut">',
            '  <LineStyle><color>FF333333</color><width>2</width></LineStyle>',
            f'  <PolyStyle><color>{_kml_color_aabbggrr("#ffffff", 140)}</color><outline>1</outline></PolyStyle>',
            '</Style>',
        ]
        for tname, hx in _KML_TECH_COLORS.items():
            sid = f'wedge-{tname.replace(" ", "_")}'
            fill_a = int(0.35 * 255)
            lines += [
                f'<Style id="{sid}">',
                f'  <LineStyle><color>{_kml_color_aabbggrr(hx, 255)}</color><width>2</width></LineStyle>',
                f'  <PolyStyle><color>{_kml_color_aabbggrr(hx, fill_a)}</color><outline>1</outline></PolyStyle>',
                '</Style>',
            ]
        lines += [
            '<Style id="wedge-default">',
            f'  <LineStyle><color>{_kml_color_aabbggrr("#34495e", 255)}</color><width>2</width></LineStyle>',
            f'  <PolyStyle><color>{_kml_color_aabbggrr("#34495e", int(0.35 * 255))}</color><outline>1</outline></PolyStyle>',
            '</Style>',
        ]

        lines += ['<Folder>', f'<name>{sx.escape("Sites")}</name>']
        wedge_count = 0
        for site in sites:
            site_cells = cells_by_site.get(site['site_id'], [])
            rows_html = ''.join(
                f'<tr><td>{sx.escape(c["cell_name"])}</td><td>{sx.escape(str(c["technology"] or ""))}</td>'
                f'<td>{c["azimuth"] if c["azimuth"] is not None else "—"}°</td>'
                f'<td>{sx.escape(str(c["frequency_band"] or "—"))}</td>'
                f'<td>{sx.escape(str((c["activity_status"] or c["status"]) or "—"))}</td></tr>'
                for c in site_cells
            )
            table = (
                '<table border="1" cellpadding="3">'
                '<tr><th>Cell</th><th>Tech</th><th>Azimuth</th><th>Band</th><th>Activity</th></tr>'
                f'{rows_html}</table>'
            ) if rows_html else ''
            desc = (
                f'<b>Site ID:</b> {sx.escape(str(site["site_id"]))}<br/>'
                f'<b>Vendor:</b> {sx.escape(site["vendor"] or "")}<br/>'
                f'<b>Region:</b> {sx.escape(site["region"] or "")}<br/>'
                f'<b>Status:</b> {sx.escape(site["status"] or "")}<br/><br/>{table}'
            )
            try:
                slat = float(site['latitude'])
                slng = float(site['longitude'])
            except (TypeError, ValueError):
                continue
            outer = _kml_circle_ring_coords(slat, slng, 24.0, 36, reverse=False)
            inner = _kml_circle_ring_coords(slat, slng, 12.0, 36, reverse=True)
            lines += [
                '<Placemark>',
                f'<name>{sx.escape(site["site_name"])}</name>',
                f'<description><![CDATA[{desc}]]></description>',
                '<styleUrl>#site-donut</styleUrl>',
                '<Polygon>',
                '  <extrude>0</extrude><altitudeMode>clampToGround</altitudeMode>',
                '  <outerBoundaryIs><LinearRing>',
                f'    <coordinates>{_kml_ring_to_coordinates(outer)}</coordinates>',
                '  </LinearRing></outerBoundaryIs>',
                '  <innerBoundaryIs><LinearRing>',
                f'    <coordinates>{_kml_ring_to_coordinates(inner)}</coordinates>',
                '  </LinearRing></innerBoundaryIs>',
                '</Polygon>',
                '</Placemark>',
            ]
        lines += ['</Folder>', '<Folder>', f'<name>{sx.escape("Sectors (wedges)")}</name>']

        for site in sites:
            site_cells = cells_by_site.get(site['site_id'], [])
            if not site_cells:
                continue
            try:
                slat = float(site['latitude'])
                slng = float(site['longitude'])
            except (TypeError, ValueError):
                continue
            wedges = _kml_group_cells_into_wedges(site_cells)
            for (wtech, bucket), group in sorted(wedges.items(), key=lambda x: (x[0][0], x[0][1])):
                ring = _kml_wedge_ring_coords(slat, slng, float(bucket))
                if len(ring) < 4:
                    continue
                sid = f'wedge-{wtech.replace(" ", "_")}'
                if wtech not in _KML_TECH_COLORS:
                    sid = 'wedge-default'

                names = ', '.join(_kml_cdata_fragment(str(c['cell_name'])) for c in group[:6])
                if len(group) > 6:
                    names += f' (+{len(group) - 6} more)'
                wdesc = (
                    f'<b>Site:</b> {_kml_cdata_fragment(site["site_name"])} (ID {_kml_cdata_fragment(str(site["site_id"]))})<br/>'
                    f'<b>Technology:</b> {_kml_cdata_fragment(wtech)}<br/>'
                    f'<b>Azimuth (bin centre):</b> {bucket}°<br/>'
                    f'<b>Cells:</b> {names}'
                )
                lines += [
                    '<Placemark>',
                    f'<name>{sx.escape(site["site_name"])} — {wtech} {int(bucket)}°</name>',
                    f'<description><![CDATA[{wdesc}]]></description>',
                    f'<styleUrl>#{sid}</styleUrl>',
                    '<Polygon>',
                    '  <extrude>0</extrude><altitudeMode>clampToGround</altitudeMode>',
                    '  <outerBoundaryIs><LinearRing>',
                    f'    <coordinates>{_kml_ring_to_coordinates(ring)}</coordinates>',
                    '  </LinearRing></outerBoundaryIs>',
                    '</Polygon>',
                    '</Placemark>',
                ]
                wedge_count += 1

        lines += ['</Folder>', '</Document>', '</kml>']

        tech_fn = (tech or 'All').replace('-', '_')
        fname   = f'network_export_{tech_fn}_{dt.now().strftime("%Y%m%d_%H%M")}.kml'
        uid = (request.current_user.get('id')
               if isinstance(request.current_user, dict) else request.current_user[0])
        log_activity(uid, 'export',
                     f'KML export: tech={tech or "All"}, {len(sites)} sites, {wedge_count} sector wedges')
        return Response('\n'.join(lines),
                        mimetype='application/vnd.google-earth.kml+xml',
                        headers={'Content-Disposition': f'attachment; filename="{fname}"'})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@network_map_bp.route('/api/map/refresh', methods=['POST'])
@login_required
def refresh_metadata():
    """Trigger a metadata sync in the background and return immediately."""
    import threading
    try:
        from sync.scheduler import trigger_metadata_now
        t = threading.Thread(target=trigger_metadata_now, daemon=True)
        t.start()
        uid = (request.current_user.get('id')
               if isinstance(request.current_user, dict)
               else request.current_user[0])
        log_activity(uid, 'metadata_refresh', 'Triggered metadata refresh from network map')
        return jsonify({'success': True, 'message': 'Metadata sync started'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
