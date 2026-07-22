"""Excel export for RET Management table views."""

from __future__ import annotations

import io
import re
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _safe_filename_part(value: str, *, fallback: str = 'export') -> str:
    cleaned = re.sub(r'[^\w\-]+', '_', (value or '').strip())
    return cleaned.strip('_')[:40] or fallback


def build_ret_workbook(payload: dict[str, Any]) -> tuple[io.BytesIO, str]:
    vendor = str(payload.get('vendor') or 'nokia').strip().lower()
    site_id = str(payload.get('site_id') or '').strip()
    ne_label = str(payload.get('ne_label') or '').strip()
    ne_name = str(payload.get('ne_name') or '').strip()
    mo_class = str(payload.get('mo_class') or '').strip()
    username = str(payload.get('username') or '').strip()
    table_filter = str(payload.get('table_filter') or '').strip()
    columns = [str(c) for c in (payload.get('columns') or []) if str(c).strip()]
    column_labels = payload.get('column_labels') or {}
    rows = payload.get('rows') or []
    total_rows = int(payload.get('total_rows') or len(rows))
    exported_rows = int(payload.get('exported_rows') or len(rows))

    if not columns:
        raise ValueError('No columns to export')
    if not rows:
        raise ValueError('No rows to export')

    vendor_label = 'Huawei U2020' if vendor == 'huawei' else 'Nokia NetAct'
    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    site_part = _safe_filename_part(site_id or ne_label or ne_name, fallback='site')
    filename = f'RET_{vendor}_{site_part}_{stamp}.xlsx'

    wb = Workbook()
    hdr_fill = PatternFill(start_color='1F6FEB', end_color='1F6FEB', fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True)

    ws_meta = wb.active
    ws_meta.title = 'Report Info'
    ws_meta.column_dimensions['A'].width = 24
    ws_meta.column_dimensions['B'].width = 52
    generated = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    meta_rows = [
        ('Generated (UTC)', generated),
        ('Exported By', username),
        ('Vendor', vendor_label),
        ('Site ID', site_id),
        ('Network Element', ne_label or ne_name),
        ('NE Name (U2020)', ne_name),
        ('MO Class', mo_class),
        ('Table Filter', table_filter or '(none)'),
        ('Rows In View', exported_rows),
        ('Total Rows Loaded', total_rows),
    ]
    ws_meta['A1'] = 'Field'
    ws_meta['B1'] = 'Value'
    for col, label in enumerate(('Field', 'Value'), 1):
        cell = ws_meta.cell(row=1, column=col, value=label)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')
    for row_idx, (field, value) in enumerate(meta_rows, start=2):
        ws_meta.cell(row=row_idx, column=1, value=field)
        ws_meta.cell(row=row_idx, column=2, value=value)

    sheet_title = 'Huawei RETSUBUNIT' if vendor == 'huawei' else 'Nokia RETU_R'
    ws_data = wb.create_sheet(sheet_title[:31])
    headers = [str(column_labels.get(col) or col) for col in columns]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_data.cell(row=1, column=col_idx, value=header)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for row_idx, item in enumerate(rows, start=2):
        if not isinstance(item, dict):
            continue
        for col_idx, col in enumerate(columns, start=1):
            value = item.get(col)
            if value is None:
                value = ''
            ws_data.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, col in enumerate(columns, start=1):
        max_len = len(str(column_labels.get(col) or col))
        for row_idx in range(2, min(exported_rows + 2, 102)):
            cell_val = ws_data.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(cell_val or '')))
        ws_data.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 48)

    ws_data.freeze_panes = 'A2'
    ws_data.auto_filter.ref = f'A1:{get_column_letter(len(columns))}{exported_rows + 1}'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, filename
