"""RAML/XML export for proposed AMLE parameter changes."""

from __future__ import annotations

import os
import xml.etree.ElementTree as ET
from datetime import datetime
from io import BytesIO
from typing import Any


REVIEW_HEADERS = [
    "MRBTS",
    "LNBTS",
    "LNCEL",
    "AMLEPR",
    "Max Layer",
    "Min Layer",
    "cacHeadroom",
    "deltaCac",
    "maxCacThreshold",
    "targetCarrierFreq",
    "Proposed cacHeadroom",
    "Proposed deltaCac",
    "Proposed maxCacThreshold",
    "Proposed targetCarrierFreq",
]

_ID_FILL = "FFF2CC"       # yellow — identification columns
_CURRENT_FILL = "E2EFDA"  # light green — current values
_PROPOSED_FILL = "E4DFEC" # light purple — proposed values
_GAP_FILL = "FCE4D6"      # light orange — config gap sheet


def _cell_value(value: Any) -> Any:
    if value is None or value == "":
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def build_review_excel(
    review_rows: list[dict[str, Any]],
    *,
    config_gaps: list[str] | None = None,
) -> bytes:
    """Build human-review Excel workbook matching the AMLEPR review template."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AMLE Review"

    header_font = Font(bold=True)
    id_fill = PatternFill("solid", fgColor=_ID_FILL)
    current_fill = PatternFill("solid", fgColor=_CURRENT_FILL)
    proposed_fill = PatternFill("solid", fgColor=_PROPOSED_FILL)

    ws.append(REVIEW_HEADERS)
    for col_idx, _ in enumerate(REVIEW_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if col_idx <= 6:
            cell.fill = id_fill
        elif col_idx <= 10:
            cell.fill = current_fill
        else:
            cell.fill = proposed_fill

    for row in review_rows or []:
        current = row.get("current_values") or {}
        proposed = row.get("proposed_values") or {}
        ws.append([
            _cell_value(row.get("mrbts")),
            _cell_value(row.get("lnbts")),
            _cell_value(row.get("lncel")),
            _cell_value(row.get("amlepr")),
            row.get("highest_layer"),
            row.get("lowest_layer"),
            _cell_value(current.get("cacHeadroom")),
            _cell_value(current.get("deltaCac")),
            _cell_value(current.get("maxCacThreshold")),
            _cell_value(current.get("targetCarrierFreq")),
            _cell_value(proposed.get("cacHeadroom")),
            _cell_value(proposed.get("deltaCac")),
            _cell_value(proposed.get("maxCacThreshold")),
            _cell_value(proposed.get("targetCarrierFreq")),
        ])

    ws.freeze_panes = "A2"
    for col_idx in range(1, len(REVIEW_HEADERS) + 1):
        letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 16 if col_idx <= 6 else 14

    gaps = [g for g in (config_gaps or []) if "configuration gap" in g.lower()]
    if gaps:
        gap_ws = wb.create_sheet("Configuration gaps")
        gap_ws.append(["Sector / message"])
        gap_ws["A1"].font = header_font
        gap_ws["A1"].fill = PatternFill("solid", fgColor=_GAP_FILL)
        for message in gaps:
            gap_ws.append([message])
        gap_ws.column_dimensions["A"].width = 100

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_changes_xml(
    changes: list[dict[str, Any]],
    *,
    plan_name: str = "PrimeNet_Nokia_Load_Balancing",
) -> str:
    """
    Build Nokia RAML2 plan XML from a list of changes.

    Each change: {mo_class, target/dist_name/dn, parameter, new_value}
    """
    root = ET.Element("raml", {"version": "2.0", "xmlns": "raml21.xsd"})
    cm_data = ET.SubElement(
        root,
        "cmData",
        {
            "type": "plan",
            "name": plan_name,
            "version": os.environ.get("NOKIA_CM_REIMPORT_RAML_VERSION", "xL21A_2012_003"),
        },
    )
    header = ET.SubElement(cm_data, "header")
    ET.SubElement(
        header,
        "log",
        {
            "dateTime": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            "action": "created",
            "appInfo": "PrimeNet Nokia Load Balancing",
        },
    )

    grouped: dict[tuple[str, str], dict[str, str]] = {}
    for change in changes:
        mo_class = str(change.get("mo_class") or change.get("mo_class_id") or "").strip()
        target = str(change.get("target") or change.get("dist_name") or change.get("dn") or change.get("DN") or "").strip()
        parameter = str(change.get("parameter") or "").strip()
        new_value = str(change.get("new_value") if change.get("new_value") is not None else "").strip()
        if not mo_class or not target or not parameter:
            continue
        grouped.setdefault((mo_class, target), {})[parameter] = new_value

    mo_operation = os.environ.get("NOKIA_CM_REIMPORT_MO_OPERATION", "update")
    for (mo_class, target), params in sorted(grouped.items()):
        mo_elem = ET.SubElement(
            cm_data,
            "managedObject",
            {
                "class": mo_class,
                "distName": target,
                "operation": mo_operation,
            },
        )
        for name, value in sorted(params.items()):
            ET.SubElement(mo_elem, "p", {"name": name}).text = value

    tree = ET.ElementTree(root)
    ET.indent(tree, space="  ")
    body = ET.tostring(root, encoding="unicode")
    return "<?xml version='1.0' encoding='UTF-8'?>\n" + body


def build_backup_xml(
    changes: list[dict[str, Any]],
    *,
    plan_name: str = "PrimeNet_Nokia_Load_Balancing_backup",
) -> str:
    """Same RAML shape as ``build_changes_xml`` but with current (pre-change) values."""
    backup_changes = []
    for change in changes or []:
        old_value = change.get("old_value")
        if old_value is None or str(old_value).strip() == "":
            continue
        backup_changes.append({**change, "new_value": old_value})
    if not backup_changes:
        raise ValueError("No current values available for backup.")
    return build_changes_xml(backup_changes, plan_name=plan_name)
