"""Excel + MML export for Huawei CellMLB proposals."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from . import config

REVIEW_HEADERS = [
    "Sector",
    "Site",
    "Layer",
    "Role",
    "MO",
    "IdleMlbUeNumThd",
    "HoMlbUeNumThd",
    "Proposed IdleMlbUeNumThd",
    "Proposed HoMlbUeNumThd",
]


def _cell_value(value: Any) -> Any:
    if value is None or value == "":
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def build_review_excel(
    review_rows: list[dict[str, Any]],
    *,
    config_gaps: list[str] | None = None,
) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CellMLB Review"
    header_font = Font(bold=True)
    ws.append(REVIEW_HEADERS)
    for col_idx, _ in enumerate(REVIEW_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row in review_rows or []:
        current = row.get("current_values") or {}
        proposed = row.get("proposed_values") or {}
        ws.append([
            _cell_value(row.get("sector_id")),
            _cell_value(row.get("site_id")),
            _cell_value(row.get("layer")),
            _cell_value(row.get("role")),
            _cell_value(row.get("mo_class") or config.CELLMLB_MO),
            _cell_value(current.get("IdleMlbUeNumThd")),
            _cell_value(current.get("HoMlbUeNumThd")),
            _cell_value(proposed.get("IdleMlbUeNumThd")),
            _cell_value(proposed.get("HoMlbUeNumThd")),
        ])

    gaps = [g for g in (config_gaps or []) if str(g).strip()]
    if gaps:
        gap_ws = wb.create_sheet("Notes")
        gap_ws.append(["Message"])
        gap_ws["A1"].font = header_font
        gap_ws["A1"].fill = PatternFill("solid", fgColor="FCE4D6")
        for message in gaps:
            gap_ws.append([message])
        gap_ws.column_dimensions["A"].width = 100

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_mml(changes: list[dict[str, Any]]) -> str:
    """U2020-style MML grouped by sector/layer. LocalCellId is a review placeholder."""
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for change in changes or []:
        key = (str(change.get("sector_id") or ""), str(change.get("layer") or ""))
        grouped.setdefault(key, {"meta": change, "params": {}})
        grouped[key]["params"][str(change.get("parameter"))] = change.get("new_value")

    lines = [
        f"// PrimeNet Huawei Load Balancing CellMLB plan {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "// Review LocalCellId on U2020 before execution. OSS push is not enabled from PrimeNet.",
        "",
    ]
    for (sector_id, layer), bundle in sorted(grouped.items()):
        meta = bundle["meta"]
        params = bundle["params"]
        assignments = ", ".join(f"{name}={value}" for name, value in sorted(params.items()))
        site = meta.get("site_id") or sector_id.split("_")[0]
        lines.append(f"// Sector {sector_id} layer {layer} ({meta.get('role')})")
        lines.append(
            f'MOD CELLMLB: eNodeBId={site}, LocalCellId=REPLACE_ME, {assignments};'
        )
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"
