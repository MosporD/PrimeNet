"""
PCI / PSC directional conflict logic: co-band reuse, distance, azimuth vs. inter-site bearing.

Tune thresholds in CONFLICT_STRICTNESS_PROFILES below.
"""

from __future__ import annotations

import io
import math
import re
from datetime import datetime, timedelta
from itertools import combinations

from db.runtime import execute_query

from modules.reports.metadata_helpers import (
    _metadata_inventory_union_sql,
    _metadata_table_columns,
    _pick_col,
    _sql_ident,
)

CONFLICT_CACHE_TTL = timedelta(days=1)
_CONFLICT_CACHE: dict[str, dict] = {}

# ── Strictness (distance + azimuth vs. inter-site bearing) ─────────────────────
#
# Pair candidates are built once up to max(dist_max_km) across all profiles, then
# each profile filters by its own dist_max_km and recomputes High / Medium / Low.
#
# Rules (same for every profile; only thresholds change):
#   - d_a / d_b = |azimuth − geographic bearing toward the other site| (wrapped 0–180°).
#   - "Aligned" for a side means d ≤ az_near_deg.
#   - High: both sides aligned AND distance_km ≤ dist_high_km.
#   - Medium: both aligned OR exactly one side aligned (and not High).
#   - Low: neither side aligned.
CONFLICT_STRICTNESS_PROFILES: dict[str, dict[str, object]] = {
    'strict': {
        'label': 'Strict',
        'hint': 'Smallest search radius; tightest match between antenna azimuth and bearing to the neighbor.',
        'dist_max_km': 4.0,
        'dist_high_km': 2.5,
        'az_near_deg': 35.0,
    },
    'standard': {
        'label': 'Standard',
        'hint': 'Original default thresholds (6 km cap, 4 km for High when both aimed, ±50°).',
        'dist_max_km': 6.0,
        'dist_high_km': 4.0,
        'az_near_deg': 50.0,
    },
    'moderate': {
        'label': 'Moderate',
        'hint': 'Between Standard and Relaxed: more pairs qualify, looser bearing window.',
        'dist_max_km': 8.0,
        'dist_high_km': 5.0,
        'az_near_deg': 58.0,
    },
    'relaxed': {
        'label': 'Relaxed',
        'hint': 'Largest radius and loosest bearing match — exploratory / catch-all screening.',
        'dist_max_km': 10.0,
        'dist_high_km': 7.0,
        'az_near_deg': 65.0,
    },
}
DEFAULT_CONFLICT_STRICTNESS = 'standard'


def normalize_conflict_tech(technology: str = '4G') -> str:
    t = str(technology or '4G').strip().upper()
    return t if t in ('3G', '4G', '5G') else '4G'


def conflict_build_max_km() -> float:
    return max(float(p['dist_max_km']) for p in CONFLICT_STRICTNESS_PROFILES.values())


def normalize_strictness(slug: str | None) -> str:
    k = str(slug or '').strip().lower()
    if k in CONFLICT_STRICTNESS_PROFILES:
        return k
    return DEFAULT_CONFLICT_STRICTNESS


def _conflict_profile_thresholds(slug: str) -> dict[str, float]:
    p = CONFLICT_STRICTNESS_PROFILES[normalize_strictness(slug)]
    return {
        'dist_max_km': float(p['dist_max_km']),
        'dist_high_km': float(p['dist_high_km']),
        'az_near_deg': float(p['az_near_deg']),
    }


def _conflict_risk_for_metrics(
    dist_km: float,
    d_a: float | None,
    d_b: float | None,
    thresholds: dict[str, float],
) -> str:
    az = float(thresholds['az_near_deg'])
    dh = float(thresholds['dist_high_km'])
    both_aligned = d_a is not None and d_b is not None and d_a <= az and d_b <= az
    one_aligned = (d_a is not None and d_a <= az) or (d_b is not None and d_b <= az)
    if both_aligned and dist_km <= dh:
        return 'High'
    if both_aligned or one_aligned:
        return 'Medium'
    return 'Low'


def conflict_strictness_profiles_public() -> list[dict[str, object]]:
    out: list[dict[str, object]] = []
    for key, p in CONFLICT_STRICTNESS_PROFILES.items():
        out.append({
            'id': key,
            'label': str(p.get('label', key.title())),
            'hint': str(p.get('hint', '')),
            'dist_max_km': float(p['dist_max_km']),
            'dist_high_km': float(p['dist_high_km']),
            'az_near_deg': float(p['az_near_deg']),
        })
    return out


def apply_strictness_to_pairs(base_rows: list[dict], strictness: str | None) -> list[dict]:
    slug = normalize_strictness(strictness)
    thr = _conflict_profile_thresholds(slug)
    dmax = thr['dist_max_km']
    out: list[dict] = []
    for r in base_rows:
        dist = r.get('distance_km')
        if dist is None or float(dist) > dmax:
            continue
        da = r.get('_d_a')
        db = r.get('_d_b')
        rr = {k: v for k, v in r.items() if not str(k).startswith('_')}
        rr['risk'] = _conflict_risk_for_metrics(float(dist), da, db, thr)
        rr['strictness'] = slug
        out.append(rr)
    risk_rank = {'High': 0, 'Medium': 1, 'Low': 2}
    out.sort(key=lambda x: (risk_rank.get(x['risk'], 9), x['distance_km'], str(x.get('pci') or '')))
    return out


def _safe_float(v):
    try:
        if v is None:
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def _extract_coband_key(cell_name: str) -> str:
    s = str(cell_name or '').strip()
    if not s:
        return ''
    hits = re.findall(r'(\d+)', s)
    return hits[-1] if hits else ''


def _haversine_km(lat1, lon1, lat2, lon2) -> float | None:
    lat1 = _safe_float(lat1)
    lon1 = _safe_float(lon1)
    lat2 = _safe_float(lat2)
    lon2 = _safe_float(lon2)
    if None in (lat1, lon1, lat2, lon2):
        return None
    r = 6371.0
    p1 = math.radians(lat1)
    p2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlambda / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


def _bearing_deg(lat1, lon1, lat2, lon2) -> float | None:
    lat1 = _safe_float(lat1)
    lon1 = _safe_float(lon1)
    lat2 = _safe_float(lat2)
    lon2 = _safe_float(lon2)
    if None in (lat1, lon1, lat2, lon2):
        return None
    p1 = math.radians(lat1)
    p2 = math.radians(lat2)
    dlambda = math.radians(lon2 - lon1)
    y = math.sin(dlambda) * math.cos(p2)
    x = math.cos(p1) * math.sin(p2) - math.sin(p1) * math.cos(p2) * math.cos(dlambda)
    b = math.degrees(math.atan2(y, x))
    return (b + 360.0) % 360.0


def _az_diff_deg(a, b) -> float | None:
    a = _safe_float(a)
    b = _safe_float(b)
    if a is None or b is None:
        return None
    d = abs((a - b) % 360.0)
    return min(d, 360.0 - d)


def _meta():
    from modules.reports.routes import _meta as reports_meta

    return reports_meta()


def build_conflict_base_pairs(technology: str = '4G'):
    """Return pair geometry + co-PCI grouping; no risk tier (that depends on strictness)."""
    conn = _meta()
    inv_union = _metadata_inventory_union_sql(conn)
    site_col_names = _metadata_table_columns(conn, 'sites')
    site_low_to_real = {str(c).strip().lower(): c for c in site_col_names}
    site_area_col = _pick_col(['area', 'region', 'market'], site_low_to_real)
    site_cluster_col = _pick_col(['cluster', 'cluster_name'], site_low_to_real)
    site_area_expr = f"s.{_sql_ident(site_area_col)}" if site_area_col else 'NULL'
    site_cluster_expr = f"s.{_sql_ident(site_cluster_col)}" if site_cluster_col else 'NULL'

    tech_req = normalize_conflict_tech(technology)
    if tech_req == '3G':
        tech_filter = ('3G',)
    elif tech_req == '4G':
        tech_filter = ('4G-FDD', '4G-TDD')
    else:
        tech_filter = ('5G',)
    filter_sql = ', '.join(['?'] * len(tech_filter))

    rows = execute_query(
        conn,
        f'''
        SELECT v.cell_name, v.technology, v.vendor, v.pci, v.azimuth, v.frequency_band,
               v.area AS cell_area,
               s.site_id, s.site_name, s.latitude, s.longitude,
               {site_area_expr} AS site_area,
               {site_cluster_expr} AS site_cluster
        FROM ({inv_union}) v
        LEFT JOIN sites s ON CAST(s.site_id AS TEXT) = CAST(v.site_id AS TEXT)
        WHERE v.cell_name IS NOT NULL
          AND TRIM(CAST(v.cell_name AS TEXT)) <> ''
          AND v.pci IS NOT NULL
          AND TRIM(CAST(v.pci AS TEXT)) <> ''
          AND (
                v.status IS NULL
                OR TRIM(COALESCE(v.status, '')) = ''
                OR LOWER(TRIM(COALESCE(v.status, ''))) = 'active'
          )
          AND v.technology IN ({filter_sql})
        ORDER BY v.pci, s.site_name
    ''',
        tech_filter,
    ).fetchall()
    conn.close()

    groups: dict[tuple[str, str], list] = {}
    for r in rows:
        rd = dict(r)
        coband = _extract_coband_key(rd.get('cell_name'))
        if not coband:
            continue
        key = (str(rd.get('pci')).strip(), coband)
        groups.setdefault(key, []).append(rd)

    dist_build_max_km = conflict_build_max_km()
    pair_rows = []
    for (pci, coband), grp in groups.items():
        if len({g['site_id'] for g in grp}) < 2:
            continue
        for a, b in combinations(grp, 2):
            if str(a.get('site_id') or '') == str(b.get('site_id') or ''):
                continue
            dist_km = _haversine_km(a.get('latitude'), a.get('longitude'), b.get('latitude'), b.get('longitude'))
            if dist_km is None or dist_km > dist_build_max_km:
                continue
            brg_ab = _bearing_deg(a.get('latitude'), a.get('longitude'), b.get('latitude'), b.get('longitude'))
            brg_ba = _bearing_deg(b.get('latitude'), b.get('longitude'), a.get('latitude'), a.get('longitude'))
            d_a = _az_diff_deg(a.get('azimuth'), brg_ab)
            d_b = _az_diff_deg(b.get('azimuth'), brg_ba)

            pair_rows.append(
                {
                    'pci': pci,
                    'coband': coband,
                    'technology': a.get('technology') or b.get('technology'),
                    'distance_km': round(dist_km, 3),
                    'bearing_ab': None if brg_ab is None else round(brg_ab, 1),
                    'bearing_ba': None if brg_ba is None else round(brg_ba, 1),
                    'a_name': a.get('cell_name'),
                    'a_site': a.get('site_name') or a.get('site_id'),
                    'a_az': a.get('azimuth'),
                    'a_band': a.get('frequency_band'),
                    'a_area': a.get('cell_area') or a.get('site_area'),
                    'a_cluster': a.get('site_cluster'),
                    'a_lat': _safe_float(a.get('latitude')),
                    'a_lng': _safe_float(a.get('longitude')),
                    'b_name': b.get('cell_name'),
                    'b_site': b.get('site_name') or b.get('site_id'),
                    'b_az': b.get('azimuth'),
                    'b_band': b.get('frequency_band'),
                    'b_area': b.get('cell_area') or b.get('site_area'),
                    'b_cluster': b.get('site_cluster'),
                    'b_lat': _safe_float(b.get('latitude')),
                    'b_lng': _safe_float(b.get('longitude')),
                    'a_to_b_diff': None if d_a is None else round(d_a, 1),
                    'b_to_a_diff': None if d_b is None else round(d_b, 1),
                    '_d_a': d_a,
                    '_d_b': d_b,
                }
            )

    pair_rows.sort(key=lambda r: (r['distance_km'], str(r.get('pci') or '')))
    return tech_req, pair_rows


def get_cached_conflict_base(technology: str, force_refresh: bool = False):
    tech = normalize_conflict_tech(technology)
    now = datetime.utcnow()
    cached = _CONFLICT_CACHE.get(tech)
    if (not force_refresh) and cached and 'base_rows' in cached:
        gen = cached.get('generated_at')
        if isinstance(gen, datetime) and now - gen <= CONFLICT_CACHE_TTL:
            return tech, cached['base_rows'], gen, False
    tech_req, base_rows = build_conflict_base_pairs(tech)
    generated_at = datetime.utcnow()
    _CONFLICT_CACHE[tech_req] = {'base_rows': base_rows, 'generated_at': generated_at}
    return tech_req, base_rows, generated_at, True


def get_cached_conflict_pairs(technology: str, strictness: str | None = None, force_refresh: bool = False):
    tech, base_rows, gen, ref = get_cached_conflict_base(technology, force_refresh=force_refresh)
    rows = apply_strictness_to_pairs(base_rows, strictness)
    return tech, rows, gen, ref


def kmlline_style_id(risk: str) -> str:
    r = str(risk or '').lower()
    if r == 'high':
        return 'risk-high'
    if r == 'medium':
        return 'risk-medium'
    return 'risk-low'


def destination_point(lat: float, lng: float, bearing_deg: float, distance_km: float):
    r = 6371.0
    br = math.radians(bearing_deg)
    p1 = math.radians(lat)
    l1 = math.radians(lng)
    d = distance_km / r
    p2 = math.asin(math.sin(p1) * math.cos(d) + math.cos(p1) * math.sin(d) * math.cos(br))
    l2 = l1 + math.atan2(math.sin(br) * math.sin(d) * math.cos(p1), math.cos(d) - math.sin(p1) * math.sin(p2))
    return math.degrees(p2), math.degrees(l2)


def wedge_polygon_coords(lat, lng, azimuth, width_deg=40.0, distance_km=0.8, segments=8):
    lat = _safe_float(lat)
    lng = _safe_float(lng)
    az = _safe_float(azimuth)
    if None in (lat, lng, az):
        return []
    start = az - (width_deg / 2.0)
    step = width_deg / max(1, segments)
    pts = [(lat, lng)]
    for i in range(segments + 1):
        b = start + (i * step)
        pts.append(destination_point(lat, lng, b, distance_km))
    pts.append((lat, lng))
    return pts


def generate_pci_conflicts_workbook(technology: str = '4G', strictness: str | None = None):
    """Return (BytesIO xlsx, filename, row_count) for the PCI conflict Excel report."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as e:
        raise RuntimeError('openpyxl required') from e

    st = normalize_strictness(strictness)
    tech_req, base_rows, _, _ = get_cached_conflict_base(technology, force_refresh=False)
    pair_rows = apply_strictness_to_pairs(base_rows, st)

    wb = Workbook()
    ws = wb.active
    ws.title = 'PCI Conflicts'

    headers = [
        'Strictness',
        'Risk',
        'PCI',
        'CoBand',
        'Distance_km',
        'Bearing_A_to_B_deg',
        'Bearing_B_to_A_deg',
        'Cell_A',
        'Site_A',
        'Area_A',
        'Cluster_A',
        'Azimuth_A',
        'A_to_B_Azimuth_Diff_deg',
        'Band_A',
        'Cell_B',
        'Site_B',
        'Area_B',
        'Cluster_B',
        'Azimuth_B',
        'B_to_A_Azimuth_Diff_deg',
        'Band_B',
    ]
    hdr_fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font

    for r in pair_rows:
        ws.append(
            [
                r.get('strictness', st),
                r['risk'],
                r['pci'],
                r['coband'],
                r['distance_km'],
                r['bearing_ab'],
                r['bearing_ba'],
                r['a_name'],
                r['a_site'],
                r['a_area'],
                r['a_cluster'],
                r['a_az'],
                r['a_to_b_diff'],
                r['a_band'],
                r['b_name'],
                r['b_site'],
                r['b_area'],
                r['b_cluster'],
                r['b_az'],
                r['b_to_a_diff'],
                r['b_band'],
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f'PCI_Conflicts_{tech_req}_{st}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    return buf, fn, len(pair_rows)
