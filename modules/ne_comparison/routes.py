"""
NE Comparison Routes
Handles XML configuration comparison functionality
"""

from flask import Blueprint, request, jsonify, send_file, render_template, redirect, url_for
from werkzeug.utils import secure_filename
from collections import Counter, defaultdict
import os
import tempfile
import uuid
from functools import wraps
from typing import Any

from ncm_core import XMLComparator
from database_enhanced import get_user_by_session, log_activity
from core.cm_extractor.extraction import build_huawei_client, build_nokia_client
from core.cm_extractor.huawei_client import HuaweiCmError
from core.cm_extractor.huawei_semantics import build_mml_command, get_mo_object_catalog
from core.cm_extractor.mml_parser import repair_mml_rows
from core.cm_extractor.nokia_client import NokiaCmError
from core.cm_extractor.nokia_semantics import extract_full_mo_class, get_mo_class_catalog
from core.cm_extractor.site_catalog import list_huawei_db_sites, list_nokia_db_sites

ne_comparison_bp = Blueprint(
    'ne_comparison', __name__,
    template_folder='templates',
    static_folder='static',
    static_url_path='/ne_comparison/static',
)

TEMP_FILES = {}

NOKIA_DEFAULT_MO_CLASSES = {
    'MRBTS': ['MRBTS:MRBTS', 'NOKLTE:LNBTS', 'NOKLTE:LNCEL', 'NOKNR:NRBTS', 'NOKNR:NRCELL'],
    'RNC': ['NOKRNC:RNC', 'NOKRNC:WBTS', 'NOKRNC:WCEL'],
    'BSC': ['NOKBSC:BSC', 'NOKBSC:BCF', 'NOKBSC:BTS', 'NOKBSC:TRX'],
}

HUAWEI_DEFAULT_MO_CLASSES = ['ENODEBFUNCTION', 'CELL', 'CNOPERATOR']
MAX_COMPARE_DIFFS = 2500
MAX_AUDIT_VALUE_SAMPLES = 8
AUDIT_IDENTITY_COLUMNS = {
    'dn', 'distname', 'moid', '$instance', 'instance', 'ne',
    'cell name', 'object name',
}

def login_required(f):
    """Decorator to require login"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        session_token = request.cookies.get('session_token')
        if not session_token:
            return redirect(url_for('auth.login_page'))

        user = get_user_by_session(session_token)
        if not user:
            return redirect(url_for('auth.login_page'))

        request.current_user = user
        return f(*args, **kwargs)

    return decorated_function

def get_current_user():
    """Get current logged-in user"""
    session_token = request.cookies.get('session_token')
    if session_token:
        return get_user_by_session(session_token)
    return None
def format_user_data(user):
    """Format user data for templates"""
    if not user:
        return None
    if isinstance(user, dict):
        return {'username': user.get('username'), 'email': user.get('email'), 'role': user.get('role'), 'id': user.get('id')}
    return {'username': (user.get('username') if isinstance(user, dict) else user[1]), 'email': (user.get('email') if isinstance(user, dict) else user[2]), 'role': (user.get('role') if isinstance(user, dict) else user[6]), 'id': (user.get('id') if isinstance(user, dict) else user[0])}


def _user_id(user) -> int:
    return user.get('id') if isinstance(user, dict) else user[0]


def _json_body() -> dict:
    return request.get_json(silent=True) or {}


def _normalize_vendor(value: str) -> str:
    vendor = (value or 'nokia').strip().lower()
    if vendor not in ('nokia', 'huawei'):
        raise ValueError('Vendor must be nokia or huawei')
    return vendor


def _normalize_scope(vendor: str, value: str) -> str:
    scope = (value or ('MRBTS' if vendor == 'nokia' else 'ENODEB')).strip().upper()
    allowed = ('MRBTS', 'RNC', 'BSC') if vendor == 'nokia' else ('ENODEB',)
    if scope not in allowed:
        raise ValueError(f'Scope must be one of: {", ".join(allowed)}')
    return scope


def _as_list(value: Any) -> list[str]:
    if isinstance(value, str):
        return [v.strip() for v in value.split(',') if v.strip()]
    if isinstance(value, (list, tuple)):
        return [str(v).strip() for v in value if str(v).strip()]
    return []


def _nokia_default_classes(scope_level: str) -> list[str]:
    return list(NOKIA_DEFAULT_MO_CLASSES.get(scope_level, NOKIA_DEFAULT_MO_CLASSES['MRBTS']))


def _huawei_default_classes() -> list[str]:
    return list(HUAWEI_DEFAULT_MO_CLASSES)


def _split_nokia_mo_id(mo_id: str) -> tuple[str, str]:
    raw = (mo_id or '').strip()
    if ':' not in raw:
        raise ValueError(f'Invalid Nokia MO class id: {raw}')
    adaptation, abbreviation = raw.split(':', 1)
    if not adaptation or not abbreviation:
        raise ValueError(f'Invalid Nokia MO class id: {raw}')
    return adaptation, abbreviation


def _sheet_to_records(sheet: dict[str, Any], *, ignore_columns: set[str] | None = None) -> dict[str, dict[str, Any]]:
    headers = [str(h) for h in (sheet.get('headers') or [])]
    rows = sheet.get('rows') or []
    ignore = {c.lower() for c in (ignore_columns or set())}
    records: dict[str, dict[str, Any]] = {}
    for idx, row in enumerate(rows):
        record = {
            header: row[pos] if pos < len(row) else ''
            for pos, header in enumerate(headers)
            if header.lower() not in ignore
        }
        key = _record_key(record, fallback=f'row-{idx + 1}')
        records[key] = record
    return records


def _rows_to_records(rows: list[dict[str, Any]], *, ignore_columns: set[str] | None = None) -> dict[str, dict[str, Any]]:
    ignore = {c.lower() for c in (ignore_columns or set())}
    records: dict[str, dict[str, Any]] = {}
    for idx, row in enumerate(rows):
        record = {str(k): v for k, v in row.items() if str(k).lower() not in ignore}
        key = _record_key(record, fallback=f'row-{idx + 1}')
        records[key] = record
    return records


def _record_key(record: dict[str, Any], *, fallback: str) -> str:
    priority = (
        'DN', 'distName', 'moId', '$instance', 'instance',
        'Local Cell ID', 'Cell Name', 'Cell ID', 'eNodeB ID',
        'Nr Cell ID', 'BTS ID', 'TRX ID', 'Object Name',
    )
    for col in priority:
        value = record.get(col)
        if value not in (None, ''):
            return f'{col}={value}'
    for col, value in record.items():
        if value not in (None, ''):
            return f'{col}={value}'
    return fallback


def _diff_records(
    left: dict[str, dict[str, Any]],
    right: dict[str, dict[str, Any]],
    *,
    section: str,
    limit_remaining: int,
) -> tuple[list[dict[str, Any]], dict[str, int]]:
    stats = {'added': 0, 'removed': 0, 'modified': 0, 'same': 0}
    diffs: list[dict[str, Any]] = []
    all_keys = sorted(set(left) | set(right))
    for key in all_keys:
        lrow = left.get(key)
        rrow = right.get(key)
        if lrow is None:
            stats['added'] += 1
            if len(diffs) < limit_remaining:
                diffs.append({'type': 'added', 'section': section, 'path': key, 'new_value': rrow})
            continue
        if rrow is None:
            stats['removed'] += 1
            if len(diffs) < limit_remaining:
                diffs.append({'type': 'removed', 'section': section, 'path': key, 'old_value': lrow})
            continue
        columns = sorted(set(lrow) | set(rrow))
        changed = [
            {
                'parameter': col,
                'old_value': '' if lrow.get(col) is None else lrow.get(col),
                'new_value': '' if rrow.get(col) is None else rrow.get(col),
            }
            for col in columns
            if str(lrow.get(col, '')) != str(rrow.get(col, ''))
        ]
        if changed:
            stats['modified'] += 1
            if len(diffs) < limit_remaining:
                diffs.append({'type': 'modified', 'section': section, 'path': key, 'changes': changed})
        else:
            stats['same'] += 1
    return diffs, stats


def _merge_stats(target: dict[str, int], source: dict[str, int]) -> None:
    for key, value in source.items():
        target[key] = int(target.get(key, 0)) + int(value or 0)


def _compare_nokia_cm(
    *,
    scope_level: str,
    ne1: dict[str, str],
    ne2: dict[str, str],
    mo_classes: list[str],
    conf_id: int,
) -> dict[str, Any]:
    client = build_nokia_client()
    summary: list[dict[str, Any]] = []
    differences: list[dict[str, Any]] = []
    stats = {'added': 0, 'removed': 0, 'modified': 0, 'same': 0}
    warnings: list[str] = []

    used_classes = mo_classes or _nokia_default_classes(scope_level)
    for mo_id in used_classes:
        try:
            adaptation, abbreviation = _split_nokia_mo_id(mo_id)
            sheet1 = extract_full_mo_class(
                client,
                adaptation,
                abbreviation,
                conf_id=conf_id,
                site_id=str(ne1.get('site_id') or ''),
                scope_level=scope_level,
                site_name=str(ne1.get('site_name') or ''),
            )
            sheet2 = extract_full_mo_class(
                client,
                adaptation,
                abbreviation,
                conf_id=conf_id,
                site_id=str(ne2.get('site_id') or ''),
                scope_level=scope_level,
                site_name=str(ne2.get('site_name') or ''),
            )
            left = _sheet_to_records(sheet1)
            right = _sheet_to_records(sheet2)
            section_diffs, section_stats = _diff_records(
                left,
                right,
                section=mo_id,
                limit_remaining=max(0, MAX_COMPARE_DIFFS - len(differences)),
            )
            differences.extend(section_diffs)
            _merge_stats(stats, section_stats)
            summary.append({
                'section': mo_id,
                'left_count': len(left),
                'right_count': len(right),
                **section_stats,
            })
        except Exception as exc:
            warnings.append(f'{mo_id}: {exc}')

    return {
        'vendor': 'nokia',
        'scope_level': scope_level,
        'mo_classes': used_classes,
        'stats': stats,
        'summary': summary,
        'differences': differences,
        'warnings': warnings,
        'truncated': len(differences) >= MAX_COMPARE_DIFFS,
    }


def _compare_huawei_cm(*, ne1_name: str, ne2_name: str, mo_classes: list[str]) -> dict[str, Any]:
    client = build_huawei_client()
    summary: list[dict[str, Any]] = []
    differences: list[dict[str, Any]] = []
    stats = {'added': 0, 'removed': 0, 'modified': 0, 'same': 0}
    warnings: list[str] = []

    used_classes = mo_classes or _huawei_default_classes()
    for mo_id in used_classes:
        try:
            command = build_mml_command(mo_id)
            rows1 = repair_mml_rows(client.run_mml_chunked(command, [ne1_name]))
            errors1 = client.consume_mml_errors()
            rows2 = repair_mml_rows(client.run_mml_chunked(command, [ne2_name]))
            errors2 = client.consume_mml_errors()
            for err in errors1 + errors2:
                warnings.append(f'{mo_id}: {err}')
            left = _rows_to_records(rows1, ignore_columns={'NE'})
            right = _rows_to_records(rows2, ignore_columns={'NE'})
            section_diffs, section_stats = _diff_records(
                left,
                right,
                section=mo_id,
                limit_remaining=max(0, MAX_COMPARE_DIFFS - len(differences)),
            )
            differences.extend(section_diffs)
            _merge_stats(stats, section_stats)
            summary.append({
                'section': mo_id,
                'left_count': len(left),
                'right_count': len(right),
                **section_stats,
            })
        except Exception as exc:
            warnings.append(f'{mo_id}: {exc}')

    return {
        'vendor': 'huawei',
        'scope_level': 'ENODEB',
        'mo_classes': used_classes,
        'stats': stats,
        'summary': summary,
        'differences': differences,
        'warnings': warnings,
        'truncated': len(differences) >= MAX_COMPARE_DIFFS,
    }


def _ne_display_name(ne: dict[str, Any]) -> str:
    return str(
        ne.get('label')
        or ne.get('site_name')
        or ne.get('u2020_ne_name')
        or ne.get('ne_name')
        or ne.get('site_id')
        or 'NE'
    )


def _normalize_audit_value(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, (dict, list, tuple)):
        return str(value)
    return str(value).strip()


def _audit_status(score: float) -> str:
    if score <= 0:
        return 'consistent'
    if score < 0.05:
        return 'low'
    if score < 0.2:
        return 'medium'
    return 'high'


def _add_audit_records(
    buckets: dict[tuple[str, str], list[dict[str, str]]],
    *,
    section: str,
    ne_name: str,
    records: dict[str, dict[str, Any]],
) -> None:
    for object_key, record in records.items():
        for parameter, raw_value in record.items():
            if str(parameter).strip().lower() in AUDIT_IDENTITY_COLUMNS:
                continue
            value = _normalize_audit_value(raw_value)
            buckets[(section, str(parameter))].append({
                'value': value,
                'ne': ne_name,
                'object': object_key,
            })


def _build_audit_summary(
    buckets: dict[tuple[str, str], list[dict[str, str]]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for (section, parameter), samples in buckets.items():
        if not samples:
            continue
        value_counts = Counter(sample['value'] for sample in samples)
        most_common_value, most_common_count = value_counts.most_common(1)[0]
        total = len(samples)
        inconsistent = total - most_common_count
        score = inconsistent / total if total else 0.0
        ne_count = len({sample['ne'] for sample in samples})
        object_count = len({sample['object'] for sample in samples})

        value_samples: list[dict[str, Any]] = []
        for value, count in value_counts.most_common(MAX_AUDIT_VALUE_SAMPLES):
            sample_nes = []
            seen_nes = set()
            for sample in samples:
                if sample['value'] == value and sample['ne'] not in seen_nes:
                    seen_nes.add(sample['ne'])
                    sample_nes.append(sample['ne'])
                if len(sample_nes) >= 5:
                    break
            value_samples.append({
                'value': value,
                'count': count,
                'percent': round((count / total) * 100, 2) if total else 0,
                'sample_nes': sample_nes,
            })

        rows.append({
            'section': section,
            'parameter': parameter,
            'total_samples': total,
            'ne_count': ne_count,
            'object_count': object_count,
            'distinct_values': len(value_counts),
            'most_common_value': most_common_value,
            'most_common_count': most_common_count,
            'inconsistent_count': inconsistent,
            'inconsistency_pct': round(score * 100, 2),
            'status': _audit_status(score),
            'values': value_samples,
        })

    return sorted(
        rows,
        key=lambda row: (
            -float(row['inconsistency_pct']),
            -int(row['distinct_values']),
            str(row['section']),
            str(row['parameter']),
        ),
    )


def _audit_nokia_cm(
    *,
    scope_level: str,
    nes: list[dict[str, Any]],
    mo_classes: list[str],
    conf_id: int,
) -> dict[str, Any]:
    client = build_nokia_client()
    used_classes = mo_classes or _nokia_default_classes(scope_level)
    buckets: dict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    warnings: list[str] = []
    section_counts: dict[str, dict[str, int]] = {
        mo_id: {'ne_count': 0, 'object_count': 0}
        for mo_id in used_classes
    }

    for ne in nes:
        ne_name = _ne_display_name(ne)
        for mo_id in used_classes:
            try:
                adaptation, abbreviation = _split_nokia_mo_id(mo_id)
                sheet = extract_full_mo_class(
                    client,
                    adaptation,
                    abbreviation,
                    conf_id=conf_id,
                    site_id=str(ne.get('site_id') or ''),
                    scope_level=scope_level,
                    site_name=str(ne.get('site_name') or ''),
                )
                records = _sheet_to_records(sheet)
                _add_audit_records(buckets, section=mo_id, ne_name=ne_name, records=records)
                section_counts[mo_id]['ne_count'] += 1
                section_counts[mo_id]['object_count'] += len(records)
            except Exception as exc:
                warnings.append(f'{ne_name} / {mo_id}: {exc}')

    parameter_summary = _build_audit_summary(buckets)
    return {
        'vendor': 'nokia',
        'scope_level': scope_level,
        'mo_classes': used_classes,
        'ne_count': len(nes),
        'section_summary': [
            {'section': section, **counts}
            for section, counts in section_counts.items()
        ],
        'parameter_summary': parameter_summary,
        'warnings': warnings,
    }


def _audit_huawei_cm(*, nes: list[dict[str, Any]], mo_classes: list[str]) -> dict[str, Any]:
    client = build_huawei_client()
    used_classes = mo_classes or _huawei_default_classes()
    buckets: dict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    warnings: list[str] = []
    section_counts: dict[str, dict[str, int]] = {
        mo_id: {'ne_count': 0, 'object_count': 0}
        for mo_id in used_classes
    }

    for mo_id in used_classes:
        try:
            command = build_mml_command(mo_id)
        except Exception as exc:
            warnings.append(f'{mo_id}: {exc}')
            continue
        for ne in nes:
            ne_name = str(ne.get('u2020_ne_name') or ne.get('ne_name') or ne.get('site_name') or '').strip()
            if not ne_name:
                warnings.append(f'{_ne_display_name(ne)} / {mo_id}: missing Huawei NE name')
                continue
            try:
                rows = repair_mml_rows(client.run_mml_chunked(command, [ne_name]))
                for err in client.consume_mml_errors():
                    warnings.append(f'{ne_name} / {mo_id}: {err}')
                records = _rows_to_records(rows, ignore_columns={'NE'})
                _add_audit_records(buckets, section=mo_id, ne_name=ne_name, records=records)
                section_counts[mo_id]['ne_count'] += 1
                section_counts[mo_id]['object_count'] += len(records)
            except Exception as exc:
                warnings.append(f'{ne_name} / {mo_id}: {exc}')

    parameter_summary = _build_audit_summary(buckets)
    return {
        'vendor': 'huawei',
        'scope_level': 'ENODEB',
        'mo_classes': used_classes,
        'ne_count': len(nes),
        'section_summary': [
            {'section': section, **counts}
            for section, counts in section_counts.items()
        ],
        'parameter_summary': parameter_summary,
        'warnings': warnings,
    }


@ne_comparison_bp.route('/ne-comparison')
@login_required
def ne_comparison_page():
    """Render NE Comparison page"""
    user = get_current_user()
    return render_template('ne_comparison.html', user=format_user_data(user))


@ne_comparison_bp.route('/api/ne-comparison/cm/nes', methods=['GET'])
def cm_ne_options():
    """Return selectable NEs from the same CM inventory sources used by CM Extractor."""
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    try:
        vendor = _normalize_vendor(request.args.get('vendor', 'nokia'))
        scope_level = _normalize_scope(vendor, request.args.get('scope_level', ''))
        query = (request.args.get('q') or '').strip()
        limit = int(request.args.get('limit') or 500)
        if vendor == 'nokia':
            items = list_nokia_db_sites(query, scope_level=scope_level, limit=limit)
        else:
            items = list_huawei_db_sites(query, scope_level=scope_level, limit=limit)
        return jsonify({'success': True, 'vendor': vendor, 'scope_level': scope_level, 'items': items})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@ne_comparison_bp.route('/api/ne-comparison/cm/mo-classes', methods=['GET'])
def cm_mo_class_options():
    """Return MO classes that can be compared for the selected vendor/scope."""
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    try:
        vendor = _normalize_vendor(request.args.get('vendor', 'nokia'))
        scope_level = _normalize_scope(vendor, request.args.get('scope_level', ''))
        if vendor == 'nokia':
            defaults = set(_nokia_default_classes(scope_level))
            items: list[dict[str, Any]] = []
            try:
                client = build_nokia_client()
                items = get_mo_class_catalog(client, scope_level=scope_level)
            except Exception:
                items = [
                    {
                        'id': mo_id,
                        'label': mo_id.split(':', 1)[-1],
                        'group': mo_id.split(':', 1)[0],
                    }
                    for mo_id in _nokia_default_classes(scope_level)
                ]
            for item in items:
                item['recommended'] = item.get('id') in defaults
            return jsonify({
                'success': True,
                'vendor': vendor,
                'scope_level': scope_level,
                'items': items,
                'default_mo_classes': list(defaults),
            })

        defaults = set(_huawei_default_classes())
        items = get_mo_object_catalog()
        for item in items:
            item['recommended'] = item.get('id') in defaults
        return jsonify({
            'success': True,
            'vendor': vendor,
            'scope_level': scope_level,
            'items': items,
            'default_mo_classes': list(defaults),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@ne_comparison_bp.route('/api/ne-comparison/cm/compare', methods=['POST'])
def compare_cm_nes():
    """Compare two same-level NEs using CM API data instead of uploaded XML files."""
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    try:
        data = _json_body()
        vendor = _normalize_vendor(data.get('vendor', 'nokia'))
        scope_level = _normalize_scope(vendor, data.get('scope_level', ''))
        mo_classes = _as_list(data.get('mo_classes'))
        conf_id = int(data.get('conf_id') or 1)
        ne1 = data.get('ne1') or {}
        ne2 = data.get('ne2') or {}
        if not isinstance(ne1, dict) or not isinstance(ne2, dict):
            return jsonify({'error': 'Both NEs are required'}), 400
        if vendor == 'nokia':
            if not ne1.get('site_id') or not ne2.get('site_id'):
                return jsonify({'error': 'Select two Nokia NEs from the same level'}), 400
            result = _compare_nokia_cm(
                scope_level=scope_level,
                ne1=ne1,
                ne2=ne2,
                mo_classes=mo_classes,
                conf_id=conf_id,
            )
        else:
            ne1_name = str(ne1.get('u2020_ne_name') or ne1.get('ne_name') or '').strip()
            ne2_name = str(ne2.get('u2020_ne_name') or ne2.get('ne_name') or '').strip()
            if not ne1_name or not ne2_name:
                return jsonify({'error': 'Select two resolved Huawei NEs from the same level'}), 400
            result = _compare_huawei_cm(ne1_name=ne1_name, ne2_name=ne2_name, mo_classes=mo_classes)

        result['success'] = True
        result['left_ne'] = ne1
        result['right_ne'] = ne2
        log_activity(
            _user_id(user),
            'ne_cm_comparison',
            f'Compared {vendor} {scope_level}: {ne1.get("site_id") or ne1.get("ne_name")} vs {ne2.get("site_id") or ne2.get("ne_name")}',
        )
        return jsonify(result)
    except (NokiaCmError, HuaweiCmError, ValueError) as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@ne_comparison_bp.route('/api/ne-comparison/cm/audit', methods=['POST'])
def audit_cm_network():
    """Run a CM API network audit and summarize per-parameter inconsistency."""
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    try:
        data = _json_body()
        vendor = _normalize_vendor(data.get('vendor', 'nokia'))
        scope_level = _normalize_scope(vendor, data.get('scope_level', ''))
        mo_classes = _as_list(data.get('mo_classes'))
        conf_id = int(data.get('conf_id') or 1)
        nes = data.get('nes') or []
        if not isinstance(nes, list) or not nes:
            return jsonify({'error': 'Select at least one NE for the NW audit'}), 400
        nes = [ne for ne in nes if isinstance(ne, dict)]
        if not nes:
            return jsonify({'error': 'Select at least one valid NE for the NW audit'}), 400

        if vendor == 'nokia':
            result = _audit_nokia_cm(
                scope_level=scope_level,
                nes=nes,
                mo_classes=mo_classes,
                conf_id=conf_id,
            )
        else:
            result = _audit_huawei_cm(nes=nes, mo_classes=mo_classes)

        result['success'] = True
        result['stats'] = {
            'parameters': len(result.get('parameter_summary') or []),
            'high': sum(1 for row in result.get('parameter_summary') or [] if row.get('status') == 'high'),
            'medium': sum(1 for row in result.get('parameter_summary') or [] if row.get('status') == 'medium'),
            'low': sum(1 for row in result.get('parameter_summary') or [] if row.get('status') == 'low'),
            'consistent': sum(1 for row in result.get('parameter_summary') or [] if row.get('status') == 'consistent'),
        }
        log_activity(
            _user_id(user),
            'ne_cm_audit',
            f'Audited {vendor} {scope_level}: {len(nes)} NE(s), {len(result.get("mo_classes") or [])} MO class(es)',
        )
        return jsonify(result)
    except (NokiaCmError, HuaweiCmError, ValueError) as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@ne_comparison_bp.route('/api/ne-comparison/compare', methods=['POST'])
def compare_files():
    """Compare two XML files - returns Excel file directly like old version"""
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    if 'file1' not in request.files or 'file2' not in request.files:
        return jsonify({'error': 'Both files required'}), 400

    file1 = request.files['file1']
    file2 = request.files['file2']

    if file1.filename == '' or file2.filename == '':
        return jsonify({'error': 'Both files must be selected'}), 400

    if not (file1.filename.endswith('.xml') and file2.filename.endswith('.xml')):
        return jsonify({'error': 'Both files must be XML'}), 400

    try:
        file_id = str(uuid.uuid4())

        filename1 = secure_filename(file1.filename)
        filename2 = secure_filename(file2.filename)

        temp_path1 = os.path.join(tempfile.gettempdir(), f"{file_id}_1_{filename1}")
        temp_path2 = os.path.join(tempfile.gettempdir(), f"{file_id}_2_{filename2}")

        file1.save(temp_path1)
        file2.save(temp_path2)

        # Create output filename with timestamp like old version
        from datetime import datetime
        output_filename = f"XML_Comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join(tempfile.gettempdir(), f"{file_id}_{output_filename}")

        # Compare
        comparator = XMLComparator(temp_path1, temp_path2, output_path)
        success, diff_count = comparator.compare()

        if not success:
            return jsonify({'error': 'Comparison failed'}), 500

        # Log activity
        log_activity((user.get('id') if isinstance(user, dict) else user[0]), 'ne_comparison', f'Compared {filename1} and {filename2}')

        # Send file directly like old version
        return send_file(
            output_path,
            as_attachment=True,
            download_name=output_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@ne_comparison_bp.route('/api/ne-comparison/download-report', methods=['POST'])
def download_report():
    """Download comparison report"""
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        data = request.get_json()

        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Comparison Report"

        ws['A1'] = 'Type'
        ws['B1'] = 'Parameter/MO'
        ws['C1'] = 'Old Value'
        ws['D1'] = 'New Value'
        ws['E1'] = 'Path'

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

        row = 2
        for diff in data.get('differences', []):
            ws[f'A{row}'] = diff.get('type', '')
            ws[f'B{row}'] = diff.get('parameter', diff.get('mo_class', ''))
            ws[f'C{row}'] = str(diff.get('old_value', ''))
            ws[f'D{row}'] = str(diff.get('new_value', ''))
            ws[f'E{row}'] = diff.get('path', '')

            if diff.get('type') == 'added':
                fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            elif diff.get('type') == 'removed':
                fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            elif diff.get('type') == 'modified':
                fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            else:
                fill = None

            if fill:
                for cell in ws[row]:
                    cell.fill = fill

            row += 1

        temp_path = os.path.join(tempfile.gettempdir(), f"report_{uuid.uuid4()}.xlsx")
        wb.save(temp_path)

        log_activity((user.get('id') if isinstance(user, dict) else user[0]), 'report_download', 'Downloaded comparison report')

        return send_file(
            temp_path,
            as_attachment=True,
            download_name='comparison_report.xlsx'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500
