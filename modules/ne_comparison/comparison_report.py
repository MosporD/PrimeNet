"""Build MAE-style NE comparison Excel reports for Nokia and Huawei."""

from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill(start_color='6699CC', end_color='6699CC', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')
TITLE_FONT = Font(bold=True)
SECTION_FONT = Font(bold=True, size=11)

VENDOR_LABELS = {
    'nokia': 'Nokia NetAct',
    'huawei': 'Huawei U2020',
}

SCOPE_COMPARE_TYPES = {
    'MRBTS': 'MRBTS Configuration Data',
    'RNC': 'RNC Configuration Data',
    'BSC': 'BSC Configuration Data',
    'ENODEB': 'eNodeB Configuration Data',
}

NOKIA_IDENTITY_COLUMNS = {
    'dn', 'distname', 'moid', 'instance', '$instance', 'ne', 'object name',
}


def _safe_str(value: Any) -> str:
    if value is None:
        return ''
    return str(value).strip()


def _ne_label(ne: dict[str, Any] | None) -> str:
    if not isinstance(ne, dict):
        return ''
    return _safe_str(
        ne.get('label')
        or ne.get('site_name')
        or ne.get('u2020_ne_name')
        or ne.get('ne_name')
        or ne.get('site_id')
    )


def _vendor_label(vendor: str) -> str:
    return VENDOR_LABELS.get(_safe_str(vendor).lower(), 'PrimeNet')


def _compare_type(scope_level: str, vendor: str) -> str:
    scope = _safe_str(scope_level).upper()
    if scope in SCOPE_COMPARE_TYPES:
        return SCOPE_COMPARE_TYPES[scope]
    vendor_name = _vendor_label(vendor)
    return f'{scope.title()} {vendor_name} Configuration Data'


def _moc_name(section: str, vendor: str) -> str:
    section = _safe_str(section)
    if not section:
        return ''
    if _safe_str(vendor).lower() == 'nokia' and ':' in section:
        return section.split(':', 1)[-1]
    return section


def _reference_data(payload: dict[str, Any]) -> str:
    left_ne = payload.get('left_ne') or {}
    ref_ne_name = _ne_label(left_ne)
    vendor = _safe_str(payload.get('vendor')).lower()
    scope_level = _safe_str(payload.get('scope_level')).upper() or 'MRBTS'

    if vendor == 'nokia':
        site_id = _safe_str(left_ne.get('site_id')) or ref_ne_name
        parts = [f'{scope_level}={site_id}']
        conf_id = payload.get('conf_id')
        if conf_id not in (None, ''):
            parts.append(f'ConfId={conf_id}')
        return ','.join(parts)

    return ref_ne_name or 'Reference NE'


def _format_object_info(record: Any, *, vendor: str = '') -> str:
    if not isinstance(record, dict):
        return _safe_str(record)
    parts: list[str] = []
    for key, raw in record.items():
        if raw is None or raw == '':
            continue
        if _safe_str(vendor).lower() == 'nokia' and str(key).strip().lower() in NOKIA_IDENTITY_COLUMNS:
            continue
        value = _safe_str(raw)
        if re.search(r'[,"\s=&]', value):
            parts.append(f'{key}="{value}"')
        else:
            parts.append(f'{key}={value}')
    return ','.join(parts)


def _reference_object_identity(ref_ne_name: str, path: str, *, vendor: str = '') -> str:
    path = _safe_str(path)
    if not path:
        if _safe_str(vendor).lower() == 'nokia':
            return f'NE="{ref_ne_name}"'
        return f'NE="{ref_ne_name}"'
    if path.upper().startswith('NE='):
        return path
    if _safe_str(vendor).lower() == 'nokia' and path.upper().startswith('DN='):
        return path
    return f'NE="{ref_ne_name}",{path}'


def _selection_summary(selections: list[dict[str, Any]], vendor: str) -> str:
    if not selections:
        return 'NE Configuration Data'
    labels: list[str] = []
    for sel in selections:
        if _safe_str(vendor).lower() == 'nokia':
            mo_id = _safe_str(sel.get('mo_class_id') or sel.get('id'))
            labels.append(_moc_name(mo_id, 'nokia'))
        else:
            labels.append(_safe_str(sel.get('mo_id') or sel.get('id')))
    joined = ', '.join(label for label in labels if label)
    return joined or 'NE Configuration Data'


def _selection_version(selections: list[dict[str, Any]], section: str, vendor: str) -> str:
    section_key = _moc_name(section, vendor) if vendor == 'nokia' else _safe_str(section).upper()
    for sel in selections or []:
        if _safe_str(vendor).lower() == 'nokia':
            mo_id = _safe_str(sel.get('mo_class_id') or sel.get('id'))
            if section and mo_id and _moc_name(mo_id, 'nokia') != section_key and mo_id != section:
                continue
            version = _safe_str(sel.get('version'))
            if version:
                return version
        else:
            mo_id = _safe_str(sel.get('mo_id') or sel.get('id')).upper()
            if section and mo_id and mo_id != section_key:
                continue
    return ''


def _style_table_header(ws: Worksheet, row: int, headers: list[str]) -> None:
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def _write_section_title(ws: Worksheet, row: int, title: str, *, back_link: bool = True) -> None:
    ws.cell(row=row, column=1, value=title).font = SECTION_FONT
    if back_link:
        ws.cell(row=row, column=6, value='back to cover')


def _split_differences(
    differences: list[dict[str, Any]],
    *,
    vendor: str = '',
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    object_rows: list[dict[str, Any]] = []
    parameter_rows: list[dict[str, Any]] = []
    moc_rows: list[dict[str, Any]] = []
    is_nokia = _safe_str(vendor).lower() == 'nokia'

    for diff in differences or []:
        diff_type = _safe_str(diff.get('type')).lower()
        section = _safe_str(diff.get('section'))
        path = _safe_str(diff.get('path'))

        if diff_type in ('added', 'removed'):
            object_rows.append(diff)
            continue

        if diff_type == 'modified':
            changes = diff.get('changes') or []
            if not changes:
                continue
            for change in changes:
                parameter = _safe_str(change.get('parameter'))
                if is_nokia and parameter.lower() in NOKIA_IDENTITY_COLUMNS:
                    continue
                parameter_rows.append({
                    'section': section,
                    'path': path,
                    'parameter': parameter,
                    'old_value': change.get('old_value', ''),
                    'new_value': change.get('new_value', ''),
                })
            continue

        moc_rows.append(diff)

    return moc_rows, object_rows, parameter_rows


def build_comparison_workbook(payload: dict[str, Any]) -> tuple[io.BytesIO, str]:
    """Return (bytes buffer, suggested filename) for a MAE-style comparison report."""
    vendor = _safe_str(payload.get('vendor')).lower() or 'nokia'
    scope_level = _safe_str(payload.get('scope_level')) or ('MRBTS' if vendor == 'nokia' else 'ENODEB')
    left_ne = payload.get('left_ne') or {}
    right_ne = payload.get('right_ne') or {}
    ref_ne_name = _ne_label(left_ne)
    target_ne_name = _ne_label(right_ne)
    selections = payload.get('selections') or []
    differences = payload.get('differences') or []
    operator = _safe_str(payload.get('operator')) or 'PrimeNet User'
    warnings = payload.get('warnings') or []

    moc_diffs, object_diffs, parameter_diffs = _split_differences(differences, vendor=vendor)
    object_count = len(object_diffs)
    parameter_count = len(parameter_diffs)
    moc_count = len(moc_diffs)

    end_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    compare_type = _compare_type(scope_level, vendor)
    reference_data = _reference_data(payload)
    task_name = target_ne_name or ref_ne_name or 'NE Comparison'
    selection_text = _selection_summary(selections, vendor)
    vendor_name = _vendor_label(vendor)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    cover = wb.create_sheet('Cover', 0)
    cover.column_dimensions['A'].width = 30
    for idx, title in enumerate([
        'Summary', 'Statistics', 'NE Information', 'MOC Difference',
        'Object Difference', 'Parameter Value Difference',
    ], start=1):
        cover.cell(row=idx, column=1, value=title)

    summary = wb.create_sheet('Summary', 1)
    summary.column_dimensions['A'].width = 30
    _write_section_title(summary, 1, 'Basic Info')
    _style_table_header(summary, 2, ['Item', 'Property'])
    summary_rows = [
        ('Version', f'PrimeNet NE Comparison — {vendor_name}'),
        ('Unit', 'Metric System'),
        ('Vendor', vendor_name),
        ('Compare Type', compare_type),
        ('Reference Data', reference_data),
        ('Target Data Count', '1'),
        ('Task Name', task_name),
        ('End Time', end_time),
        ('Operator', operator),
        ('Workspace', 'Current Area'),
        ('MO Selection', selection_text),
    ]

    for offset, (item, prop) in enumerate(summary_rows, start=3):
        summary.cell(row=offset, column=1, value=item)
        summary.cell(row=offset, column=2, value=prop)

    legend_row = len(summary_rows) + 4
    summary.cell(row=legend_row, column=1, value='Legend').font = TITLE_FONT
    _style_table_header(summary, legend_row + 1, ['Legend', 'Description'])
    summary.cell(row=legend_row + 2, column=1, value='ADDED')
    summary.cell(
        row=legend_row + 2, column=2,
        value="Added object entity or MOC or MOC's parameter to the reference data.",
    )
    summary.cell(row=legend_row + 3, column=1, value='MISSING')
    summary.cell(
        row=legend_row + 3, column=2,
        value="Missing object entity or MOC or MOC's parameter from the reference data.",
    )
    error_row = legend_row + 5
    summary.cell(row=error_row, column=1, value=f'Error Report({len(warnings)})')
    if warnings:
        summary.cell(row=error_row + 1, column=1, value='Detail Information').font = TITLE_FONT
        for idx, warning in enumerate(warnings[:20], start=error_row + 2):
            summary.cell(row=idx, column=1, value=_safe_str(warning))

    statistics = wb.create_sheet('Statistics', 2)
    statistics.column_dimensions['A'].width = 30
    _write_section_title(statistics, 1, 'Statistics')
    _style_table_header(statistics, 2, ['Statistic Item', 'Statistic Result'])
    stat_rows = [
        ('MOC Difference Number', str(moc_count)),
        ('MOC Key Parameter Difference Number', '0'),
        ('MOC Normal Parameter Difference Number', '0'),
        ('Object Difference Number', str(object_count)),
        ('Parameter Difference Number', str(parameter_count)),
    ]
    for offset, (item, result) in enumerate(stat_rows, start=3):
        statistics.cell(row=offset, column=1, value=item)
        statistics.cell(row=offset, column=2, value=result)

    ne_info = wb.create_sheet('NE Information', 3)
    ne_info.column_dimensions['A'].width = 30
    _write_section_title(ne_info, 1, 'NE(2)')
    _style_table_header(ne_info, 2, ['NE Compare Type', 'NE Name', 'Component Version', 'Version'])
    ne_info.cell(row=3, column=1, value='Base NE')
    ne_info.cell(row=3, column=2, value=ref_ne_name)
    ne_info.cell(row=3, column=3, value=_safe_str(left_ne.get('site_id')))
    ne_info.cell(row=3, column=4, value=scope_level)
    ne_info.cell(row=4, column=1, value='Target NE')
    ne_info.cell(row=4, column=2, value=target_ne_name)
    ne_info.cell(row=4, column=3, value=_safe_str(right_ne.get('site_id')))
    ne_info.cell(row=4, column=4, value=scope_level)

    moc_sheet = wb.create_sheet('MOC Difference', 4)
    moc_sheet.column_dimensions['A'].width = 30
    _write_section_title(moc_sheet, 1, f'MOC Difference Detail({moc_count})')
    _style_table_header(moc_sheet, 2, ['NE Name', 'Version', 'MOC Name', 'ADDED/MISSING'])
    moc_row = 3
    for diff in moc_diffs:
        status = 'ADDED' if _safe_str(diff.get('type')).lower() == 'added' else 'MISSING'
        section = _safe_str(diff.get('section'))
        moc_sheet.cell(row=moc_row, column=1, value=target_ne_name)
        moc_sheet.cell(row=moc_row, column=2, value=_selection_version(selections, section, vendor))
        moc_sheet.cell(row=moc_row, column=3, value=_moc_name(section, vendor))
        moc_sheet.cell(row=moc_row, column=4, value=status)
        moc_row += 1

    key_header_row = moc_row + 1
    moc_sheet.cell(row=key_header_row, column=1, value='MOC Key Parameter Difference Detail(0)').font = TITLE_FONT
    _style_table_header(
        moc_sheet,
        key_header_row + 1,
        ['NE Name', 'Version', 'MOC Name', 'Key Parameter of Target', 'Key Parameter of Reference'],
    )

    normal_header_row = key_header_row + 3
    moc_sheet.cell(row=normal_header_row, column=1, value='MOC Normal Parameter Difference Detail(0)').font = TITLE_FONT
    _style_table_header(
        moc_sheet,
        normal_header_row + 1,
        ['NE Name', 'Version', 'MOC Name', 'Normal Parameter', 'ADDED/MISSING'],
    )

    object_sheet = wb.create_sheet('Object Difference', 5)
    object_sheet.column_dimensions['A'].width = 30
    _write_section_title(object_sheet, 1, f'Object Difference({object_count})')
    _style_table_header(object_sheet, 2, [
        'NE Name', 'MOC Name', 'Object Identity', 'ADDED/MISSING',
        'Object Info', 'Support Fixing', 'Fix by Reference Value',
    ])
    obj_row = 3
    for diff in object_diffs:
        diff_type = _safe_str(diff.get('type')).lower()
        status = 'ADDED' if diff_type == 'added' else 'MISSING'
        record = diff.get('new_value') if diff_type == 'added' else diff.get('old_value')
        section = _safe_str(diff.get('section'))
        object_sheet.cell(row=obj_row, column=1, value=target_ne_name)
        object_sheet.cell(row=obj_row, column=2, value=_moc_name(section, vendor))
        object_sheet.cell(row=obj_row, column=3, value=_safe_str(diff.get('path')))
        object_sheet.cell(row=obj_row, column=4, value=status)
        object_sheet.cell(row=obj_row, column=5, value=_format_object_info(record, vendor=vendor))
        object_sheet.cell(row=obj_row, column=6, value='Yes')
        object_sheet.cell(row=obj_row, column=7, value='')
        obj_row += 1

    param_sheet = wb.create_sheet('Parameter Value Difference', 6)
    param_sheet.column_dimensions['A'].width = 30
    _write_section_title(param_sheet, 1, f'Parameter Value Difference({parameter_count})')
    _style_table_header(param_sheet, 2, [
        'NE Name', 'Object Identity', 'MOC Name', 'Normal Parameter',
        'Target Value', 'Reference Value', 'Reference Object Identity',
        'Support Fixing', 'Fix by Reference Value',
    ])
    param_row = 3
    for diff in parameter_diffs:
        section = _safe_str(diff.get('section'))
        path = _safe_str(diff.get('path'))
        param_sheet.cell(row=param_row, column=1, value=target_ne_name)
        param_sheet.cell(row=param_row, column=2, value=path)
        param_sheet.cell(row=param_row, column=3, value=_moc_name(section, vendor))
        param_sheet.cell(row=param_row, column=4, value=_safe_str(diff.get('parameter')))
        param_sheet.cell(row=param_row, column=5, value=_safe_str(diff.get('new_value')))
        param_sheet.cell(row=param_row, column=6, value=_safe_str(diff.get('old_value')))
        param_sheet.cell(
            row=param_row, column=7,
            value=_reference_object_identity(ref_ne_name, path, vendor=vendor),
        )
        param_sheet.cell(row=param_row, column=8, value='Yes')
        param_sheet.cell(row=param_row, column=9, value='')
        param_row += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_ref = re.sub(r'[^\w\-]+', '_', ref_ne_name)[:40].strip('_') or 'ref'
    safe_target = re.sub(r'[^\w\-]+', '_', target_ne_name)[:40].strip('_') or 'target'
    vendor_tag = 'Nokia' if vendor == 'nokia' else 'Huawei'
    filename = (
        f'{vendor_tag}_NE_Comparison_{safe_ref}_vs_{safe_target}_'
        f'{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )
    return buffer, filename


# Backward-compatible alias used during module rename.
build_huawei_comparison_workbook = build_comparison_workbook
