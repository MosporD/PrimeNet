"""Excel export for CM Parameter Audit live scan results."""

from __future__ import annotations

import io
import re
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def _safe_filename_part(value: str, *, fallback: str = 'param') -> str:
    cleaned = re.sub(r'[^\w\-]+', '_', (value or '').strip())
    return cleaned.strip('_') or fallback


def build_audit_workbook(payload: dict[str, Any]) -> tuple[io.BytesIO, str]:
    """Build a multi-sheet workbook from a live audit payload."""
    summary = payload.get('summary') or {}
    ne_scope = payload.get('ne_scope') or {}
    distribution = (
        summary.get('value_distribution_all')
        or summary.get('value_distribution')
        or []
    )
    rows = payload.get('rows') or []
    warnings = payload.get('warnings') or []

    vendor = str(payload.get('vendor') or '')
    mo_class = str(payload.get('mo_class') or '')
    parameter = str(payload.get('parameter') or '')
    scope_level = str(payload.get('scope_level') or '')
    query_column = str(payload.get('query_column') or parameter)
    query_mode = str(payload.get('query_mode') or '')
    area = str(ne_scope.get('area') or payload.get('area') or 'all')

    wb = Workbook()
    hdr_fill = PatternFill(start_color='1F6FEB', end_color='1F6FEB', fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True)

    def style_header(ws, headers: list[str]) -> None:
        for col, label in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal='center')

    # --- Summary ---
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    ws_summary.column_dimensions['A'].width = 28
    ws_summary.column_dimensions['B'].width = 48
    generated = datetime.now(timezone.utc).replace(microsecond=0).isoformat()

    meta_rows = [
        ('Generated (UTC)', generated),
        ('Vendor', vendor),
        ('Scope', scope_level),
        ('MO Class', mo_class),
        ('Parameter', parameter),
        ('Query Column', query_column),
        ('Query Mode', query_mode),
        ('Area', area),
        ('NEs Queried', ne_scope.get('queried', '')),
        ('NEs Available', ne_scope.get('available', '')),
        ('Scope Truncated', 'Yes' if ne_scope.get('truncated') else 'No'),
        ('Objects', summary.get('object_count', 0)),
        ('Distinct NEs', summary.get('ne_count', 0)),
        ('Distinct Values', summary.get('distinct_values', 0)),
        ('Dominant Value', summary.get('most_common_value', '')),
        ('Dominant Count', summary.get('most_common_count', 0)),
        ('Inconsistent Count', summary.get('inconsistent_count', 0)),
        ('Inconsistency %', summary.get('inconsistency_pct', 0)),
        ('Consistency Status', summary.get('status', '')),
    ]
    ws_summary['A1'] = 'Field'
    ws_summary['B1'] = 'Value'
    style_header(ws_summary, ['Field', 'Value'])
    for idx, (field, value) in enumerate(meta_rows, start=2):
        ws_summary.cell(row=idx, column=1, value=field)
        ws_summary.cell(row=idx, column=2, value=value)

    if payload.get('note'):
        note_row = len(meta_rows) + 3
        ws_summary.cell(row=note_row, column=1, value='Note')
        ws_summary.cell(row=note_row, column=2, value=str(payload.get('note')))

    # --- Value distribution ---
    ws_dist = wb.create_sheet('Value Distribution')
    dist_headers = ['Value', 'Count', 'Percent']
    style_header(ws_dist, dist_headers)
    ws_dist.column_dimensions['A'].width = 36
    ws_dist.column_dimensions['B'].width = 12
    ws_dist.column_dimensions['C'].width = 12
    for item in distribution:
        ws_dist.append([
            item.get('value', ''),
            item.get('count', 0),
            item.get('percent', 0),
        ])
    if not distribution:
        ws_dist.append(['(no data)', 0, 0])

    # --- Network status ---
    ws_rows = wb.create_sheet('Network Status')
    row_headers = ['NE', 'Site ID', 'Area', 'Cell / Object', 'DN', 'Value', 'Status']
    style_header(ws_rows, row_headers)
    ws_rows.column_dimensions['A'].width = 34
    ws_rows.column_dimensions['B'].width = 14
    ws_rows.column_dimensions['C'].width = 16
    ws_rows.column_dimensions['D'].width = 28
    ws_rows.column_dimensions['E'].width = 42
    ws_rows.column_dimensions['F'].width = 18
    ws_rows.column_dimensions['G'].width = 12
    for row in rows:
        object_label = row.get('cell_name') or row.get('object') or row.get('dn') or ''
        status = 'Dominant' if row.get('matches_dominant') else 'Variant'
        ws_rows.append([
            row.get('ne', ''),
            row.get('site_id', ''),
            row.get('area', ''),
            object_label,
            row.get('dn', ''),
            row.get('value', ''),
            status,
        ])
    if not rows:
        ws_rows.append(['(no data)', '', '', '', '', '', ''])

    # --- Warnings ---
    if warnings or payload.get('note'):
        ws_warn = wb.create_sheet('Warnings')
        style_header(ws_warn, ['Message'])
        ws_warn.column_dimensions['A'].width = 100
        if payload.get('note'):
            ws_warn.append([str(payload.get('note'))])
        for message in warnings:
            ws_warn.append([str(message)])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    stamp = datetime.now().strftime('%Y%m%d_%H%M')
    filename = (
        f'CM_Parameter_Audit_{_safe_filename_part(vendor)}_'
        f'{_safe_filename_part(mo_class)}_{_safe_filename_part(parameter)}_{stamp}.xlsx'
    )
    return buf, filename
