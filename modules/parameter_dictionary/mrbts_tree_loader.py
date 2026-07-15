"""Nokia Single RAN MRBTS graphical MO hierarchy (levels 1–8)."""

from __future__ import annotations

import json
import os
from typing import Any

import openpyxl

_DIR = os.path.dirname(__file__)
NOKIA_PARAMS_DIR = os.path.join(_DIR, "Nokia Parameters")
MRBTS_EXCEL_PATH = os.path.join(NOKIA_PARAMS_DIR, "Parameter Graphical Tree MRBTS.xlsx")
MRBTS_CACHE_PATH = os.path.join(_DIR, "data", "nokia_mrbts_tree.json")

_cache: dict[str, Any] | None = None


def _excel_mtime() -> float:
    try:
        return os.path.getmtime(MRBTS_EXCEL_PATH)
    except OSError:
        return 0.0


def _cache_mtime() -> float:
    try:
        return os.path.getmtime(MRBTS_CACHE_PATH)
    except OSError:
        return 0.0


def _build_tree_from_rows(rows: list[tuple]) -> dict[str, Any]:
    roots: list[dict[str, Any]] = []
    parent_stack: list[dict[str, Any] | None] = [None] * 8

    for row in rows:
        level_idx = None
        name = None
        for i in range(8):
            value = row[i] if i < len(row) else None
            if value is not None and str(value).strip():
                level_idx = i
                name = str(value).strip()
                break
        if level_idx is None or not name:
            continue

        meaning = str(row[8] or "").strip() if len(row) > 8 else ""
        for j in range(level_idx + 1, 8):
            parent_stack[j] = None

        path_names: list[str] = []
        for i in range(level_idx):
            parent = parent_stack[i]
            if parent is None:
                raise RuntimeError(f"Missing parent at level {i + 1} for {name}")
            path_names.append(parent["name"])
        path_names.append(name)

        node: dict[str, Any] = {
            "name": name,
            "meaning": meaning,
            "level": level_idx + 1,
            "path": path_names,
            "id": "/".join(path_names),
            "children": [],
        }
        parent_stack[level_idx] = node

        if level_idx == 0:
            roots.append(node)
        else:
            parent = parent_stack[level_idx - 1]
            assert parent is not None
            parent["children"].append(node)

    if len(roots) != 1:
        raise RuntimeError(f"Expected one MRBTS root, got {len(roots)}")
    return roots[0]


def _count_nodes(node: dict[str, Any]) -> int:
    return 1 + sum(_count_nodes(child) for child in node["children"])


def _max_depth(node: dict[str, Any]) -> int:
    if not node["children"]:
        return int(node["level"])
    return max(_max_depth(child) for child in node["children"])


def _flatten(node: dict[str, Any], out: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    if out is None:
        out = []
    out.append(
        {
            "id": node["id"],
            "name": node["name"],
            "meaning": node["meaning"],
            "level": node["level"],
            "path": node["path"],
            "child_count": len(node["children"]),
        }
    )
    for child in node["children"]:
        _flatten(child, out)
    return out


def rebuild_mrbts_tree_cache(excel_path: str | None = None) -> dict[str, Any]:
    """Parse the MRBTS Excel hierarchy and write the JSON cache."""
    path = excel_path or MRBTS_EXCEL_PATH
    if not os.path.isfile(path):
        raise FileNotFoundError(f"MRBTS hierarchy Excel not found: {path}")

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))[1:]
    tree = _build_tree_from_rows(rows)
    flat = _flatten(tree)
    node_count = _count_nodes(tree)

    payload: dict[str, Any] = {
        "meta": {
            "source": os.path.basename(path),
            "root": tree["name"],
            "title": "Single RAN MRBTS MO Hierarchy",
            "description": (
                "Graphical tree of Nokia Single RAN MRBTS managed objects "
                "(levels 1-8) with meanings."
            ),
            "node_count": node_count,
            "max_level": _max_depth(tree),
            "level2_branches": [
                {
                    "name": child["name"],
                    "meaning": child["meaning"],
                    "child_count": len(child["children"]),
                }
                for child in tree["children"]
            ],
        },
        "tree": tree,
        "flat": flat,
    }

    os.makedirs(os.path.dirname(MRBTS_CACHE_PATH), exist_ok=True)
    with open(MRBTS_CACHE_PATH, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, separators=(",", ":"))
    return payload


def _load_cache_file() -> dict[str, Any]:
    with open(MRBTS_CACHE_PATH, encoding="utf-8") as handle:
        return json.load(handle)


def load_mrbts_tree(*, force_reload: bool = False) -> dict[str, Any]:
    """Load MRBTS tree cache, rebuilding from Excel when newer."""
    global _cache
    if _cache is not None and not force_reload:
        return _cache

    excel_mtime = _excel_mtime()
    cache_mtime = _cache_mtime()
    if excel_mtime and excel_mtime > cache_mtime:
        _cache = rebuild_mrbts_tree_cache()
        return _cache

    if cache_mtime:
        _cache = _load_cache_file()
        return _cache

    if excel_mtime:
        _cache = rebuild_mrbts_tree_cache()
        return _cache

    raise FileNotFoundError(
        "MRBTS hierarchy data not found. Place "
        "'Parameter Graphical Tree MRBTS.xlsx' under Nokia Parameters/ "
        "or ensure data/nokia_mrbts_tree.json exists."
    )


def get_mrbts_tree_payload() -> dict[str, Any]:
    """API-facing payload (tree + meta; omit flat to keep response smaller)."""
    data = load_mrbts_tree()
    return {
        "meta": data.get("meta") or {},
        "tree": data.get("tree") or {},
        "flat": data.get("flat") or [],
    }
