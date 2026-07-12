"""
Scheduled Report Generation Routes
Generates Excel reports on demand or on a schedule.
"""

from flask import Blueprint, request, jsonify, render_template, redirect, url_for, send_file
from functools import wraps
import sqlite3
import os
import io
import json
import urllib.request
import urllib.parse
from datetime import datetime

from database_enhanced import get_user_by_session, log_activity
from db.runtime import connect_app, connect_metadata, execute_query
from core.elevation import coord_key as _shared_coord_key, elevation_for_points as _shared_elevation_for_points
from .metadata_helpers import _metadata_table_columns, _pick_col, _sql_ident
from modules.sync.metadata_active_sql import PER_TABLE_ACTIVE_WHERE
from sync_config import PROJECT_ROOT

reports_bp = Blueprint(
    'reports', __name__,
    template_folder='templates',
    static_folder='static',
    static_url_path='/reports/static',
)

REPORTS_DIR = os.path.join(PROJECT_ROOT, 'generated_reports')
os.makedirs(REPORTS_DIR, exist_ok=True)


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.cookies.get('session_token')
        if not token:
            return redirect(url_for('auth.login_page'))
        user = get_user_by_session(token)
        if not user:
            return redirect(url_for('auth.login_page'))
        request.current_user = user
        return f(*args, **kwargs)
    return decorated


def get_current_user():
    token = request.cookies.get('session_token')
    return get_user_by_session(token) if token else None


def format_user(user):
    if not user:
        return None
    return {'id': user.get('id'), 'username': user.get('username'), 'role': user.get('role')}


def _ncm():
    conn = connect_app()
    if isinstance(conn, sqlite3.Connection):
        conn.row_factory = sqlite3.Row
        conn.execute('''
            CREATE TABLE IF NOT EXISTS elevation_cache (
                coord_key TEXT PRIMARY KEY,
                lat REAL NOT NULL,
                lng REAL NOT NULL,
                elevation_m REAL,
                updated_at TEXT NOT NULL
            )
        ''')
        conn.commit()
    return conn


def _meta():
    conn = connect_metadata()
    if isinstance(conn, sqlite3.Connection):
        try:
            conn.execute('PRAGMA journal_mode=WAL')
        except Exception:
            pass
        conn.row_factory = sqlite3.Row
    return conn


def _safe_float(v):
    try:
        if v is None:
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def _coord_key(lat: float, lng: float) -> str:
    return _shared_coord_key(lat, lng)


def _fetch_elevation_remote(points: list[tuple[float, float]]) -> dict[str, float | None]:
    """
    Query elevation service for a list of (lat, lng).
    Uses OpenTopoData (SRTM90m). Returns map by coord_key.
    """
    if not points:
        return {}
    out: dict[str, float | None] = {}
    batch_size = 80
    for i in range(0, len(points), batch_size):
        batch = points[i:i + batch_size]
        loc = '|'.join(f'{lat:.6f},{lng:.6f}' for lat, lng in batch)
        url = 'https://api.opentopodata.org/v1/srtm90m?locations=' + urllib.parse.quote(loc, safe='|,')
        try:
            req = urllib.request.Request(url, headers={'User-Agent': 'PrimeNet/1.0'})
            with urllib.request.urlopen(req, timeout=12) as resp:
                payload = json.loads(resp.read().decode('utf-8'))
            results = payload.get('results') if isinstance(payload, dict) else None
            if not isinstance(results, list):
                results = []
            for idx, item in enumerate(results):
                if idx >= len(batch):
                    break
                lat, lng = batch[idx]
                elev = None
                if isinstance(item, dict):
                    v = item.get('elevation')
                    try:
                        elev = float(v) if v is not None else None
                    except (TypeError, ValueError):
                        elev = None
                out[_coord_key(lat, lng)] = elev
        except Exception:
            for lat, lng in batch:
                out[_coord_key(lat, lng)] = None
    return out


def _elevation_for_points(points: list[tuple[float, float]]) -> dict[str, float | None]:
    """
    Resolve elevations with DB cache first, then remote fetch for misses.
    """
    return _shared_elevation_for_points(points)


# ── Page ──────────────────────────────────────────────────────────────────────

@reports_bp.route('/reports')
@login_required
def reports_page():
    user = get_current_user()
    return render_template('reports.html', user=format_user(user))


# ── Report generators ─────────────────────────────────────────────────────────

def _generate_site_inventory(technology: str = 'all'):
    """Excel: one row per cell with dynamic full metadata columns."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError('openpyxl is required for report generation')

    conn = _meta()
    tech_key = (technology or 'all').strip().upper()
    tech_filters = {
        'ALL': None,
        '2G': ['2G'],
        '3G': ['3G'],
        '4G': ['4G-FDD', '4G-TDD'],
        '5G': ['5G'],
    }
    selected_techs = tech_filters.get(tech_key)
    if selected_techs is None and tech_key != 'ALL':
        selected_techs = None

    specs = [
        ('cells_2g', '2G', ['site_id', 'bcf id', 'bts id']),
        ('cells_3g', '3G', ['site_id', 'nodeb_id']),
        ('cells_4g_fdd', '4G-FDD', ['site_id', 'enb_id_actual', 'enodeb id']),
        ('cells_4g_tdd', '4G-TDD', ['site_id', 'enb_id_actual', 'enodeb id']),
        ('cells_5g', '5G', ['site_id', 'gnb_id_actual', 'gnb id']),
    ]

    site_cols = _metadata_table_columns(conn, 'sites')
    site_rows = execute_query(conn, 'SELECT * FROM "sites"', ()).fetchall() if site_cols else []
    site_id_col = _pick_col(['site_id'], {c.strip().lower(): c for c in site_cols}) if site_cols else None
    site_by_id: dict[str, dict] = {}
    if site_id_col:
        for sr in site_rows:
            sid = sr[site_id_col]
            if sid is None:
                continue
            site_by_id[str(sid)] = dict(sr)

    tech_records: dict[str, list[dict]] = {}
    tech_columns: dict[str, list[str]] = {}
    tech_seen_columns: dict[str, set[str]] = {}

    def _format_header(col_name: str) -> str:
        header = str(col_name).replace('_', ' ').title().replace(' Id', ' ID')
        for duplicated_prefix in ('Site Site ', 'Cell Cell '):
            if header.startswith(duplicated_prefix):
                header = header.replace(duplicated_prefix, duplicated_prefix.split()[0] + ' ', 1)
        return header

    def _register_col(tech_name: str, col_name: str):
        if tech_name not in tech_columns:
            tech_columns[tech_name] = []
            tech_seen_columns[tech_name] = set()
        if col_name not in tech_seen_columns[tech_name]:
            tech_seen_columns[tech_name].add(col_name)
            tech_columns[tech_name].append(col_name)

    for table, tech, site_aliases in specs:
        if selected_techs and tech not in selected_techs:
            continue

        table_cols = _metadata_table_columns(conn, table)
        if not table_cols:
            continue

        low_to_real = {str(c).strip().lower(): c for c in table_cols}
        site_col = _pick_col(site_aliases, low_to_real)
        cell_col = _pick_col(
            ['cell_name', 'cell name', 'wcel name', 'lncel name', 'nrcel name', 'bts name'],
            low_to_real,
        )
        status_col = _pick_col(['status', 'activity_status'], low_to_real)

        table_rows = execute_query(conn, f'SELECT * FROM {_sql_ident(table)}', ()).fetchall()
        for tr in table_rows:
            if cell_col:
                cell_val = tr[cell_col]
                if cell_val is None or str(cell_val).strip() == '':
                    continue
            else:
                continue

            if status_col:
                status_val = tr[status_col]
                status_txt = '' if status_val is None else str(status_val).strip().lower()
                if status_txt not in ('', 'active'):
                    continue

            row_dict = dict(tr)
            record: dict = {'technology': tech}

            sid = str(row_dict.get(site_col)) if site_col and row_dict.get(site_col) is not None else None
            if sid is not None:
                record['site_id'] = sid

            if cell_col:
                record['cell_name'] = row_dict.get(cell_col)

            # Attach all site columns (if found) to enrich output.
            if sid is not None and sid in site_by_id:
                for sc, sv in site_by_id[sid].items():
                    if site_id_col and str(sc).strip().lower() == str(site_id_col).strip().lower():
                        continue
                    if sc not in record:
                        record[sc] = sv
                    else:
                        record[f'site_{sc}'] = sv

            # Attach every metadata column from the selected tech table.
            for tc, tv in row_dict.items():
                if tc not in record:
                    record[tc] = tv
                else:
                    record[f'cell_{tc}'] = tv

            if tech not in tech_records:
                tech_records[tech] = []
                for base_col in ['site_id', 'site_name', 'technology', 'cell_name']:
                    _register_col(tech, base_col)

            for key in record.keys():
                _register_col(tech, key)
            tech_records[tech].append(record)

    conn.close()

    tech_label = tech_key if tech_key in tech_filters else 'ALL'
    pull_date = datetime.now().strftime('%Y-%m-%d')

    wb = Workbook()
    hdr_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True)
    ordered_techs = [t for _, t, _ in specs if t in tech_records]

    if not ordered_techs:
        ws = wb.active
        ws.title = 'Site Inventory'
        ws.cell(row=1, column=1, value='No inventory data found')
    else:
        # Use one sheet per technology when multiple technologies are included.
        for idx, tech in enumerate(ordered_techs):
            ws = wb.active if idx == 0 else wb.create_sheet()
            ws.title = tech[:31]

            cols_for_tech = tech_columns.get(tech, [])
            lead_cols = [c for c in ['site_id', 'site_name', 'technology', 'cell_name'] if c in cols_for_tech]
            other_cols = [c for c in cols_for_tech if c not in lead_cols]
            headers = lead_cols + other_cols

            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=_format_header(h))
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal='center')

            for rec in tech_records.get(tech, []):
                ws.append([rec.get(h) for h in headers])

            # Auto-width per sheet
            for col in ws.columns:
                max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    total_rows = sum(len(v) for v in tech_records.values())
    return buf, f"Cells_{tech_label}_{pull_date}.xlsx", total_rows


def _generate_pci_conflicts(technology: str = '4G', strictness: str | None = None):
    """Excel: co-band PCI conflict candidates (delegates to conflict_map.logic)."""
    from modules.conflict_map.logic import generate_pci_conflicts_workbook

    return generate_pci_conflicts_workbook(technology, strictness)


def _generate_config_versions_report():
    """Excel: all config versions uploaded."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise RuntimeError('openpyxl required')

    conn = _ncm()
    rows = execute_query(conn, '''
        SELECT cv.ne_name, cv.version_num, cv.file_name, cv.comment,
               cv.created_at, u.username as uploaded_by,
               LENGTH(CAST(cv.xml_content AS TEXT)) as file_size_bytes
        FROM config_versions cv
        LEFT JOIN users u ON cv.uploaded_by = u.id
        ORDER BY cv.ne_name, cv.version_num
    ''', ()).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Config Versions'

    headers = ['NE Name', 'Version', 'File Name', 'Comment', 'Uploaded At', 'Uploaded By', 'Size (bytes)']
    hdr_fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font

    for r in rows:
        ws.append([r['ne_name'], r['version_num'], r['file_name'],
                   r['comment'], r['created_at'], r['uploaded_by'], r['file_size_bytes']])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, f"Config_Versions_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx", len(rows)


def _generate_sector_health_excel(*, active_only: bool = True):
    """Excel: one row per site-sector with Yes/No columns for every tech/band."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError('openpyxl is required for report generation')

    from .sector_coverage_data import load_sector_coverage_rows

    sector_list, sorted_tb = load_sector_coverage_rows(active_only=active_only)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sector Health' if active_only else 'All Cells'

    hdr_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    hdr_font = Font(color='FFFFFF', bold=True, size=10)
    yes_fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
    yes_font = Font(color='FFFFFF', bold=True, size=10)
    no_fill  = PatternFill(start_color='F2F3F4', end_color='F2F3F4', fill_type='solid')
    no_font  = Font(color='BDC3C7', size=10)
    thin_border = Border(
        left=Side(style='thin', color='D5D8DC'),
        right=Side(style='thin', color='D5D8DC'),
        top=Side(style='thin', color='D5D8DC'),
        bottom=Side(style='thin', color='D5D8DC'),
    )

    fixed_headers = ['Site ID', 'Site Name', 'Vendor', 'Area', 'Sector']
    all_headers = fixed_headers + sorted_tb

    for col_idx, h in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[1].height = 40

    for row_idx, sec in enumerate(sector_list, 2):
        ws.cell(row=row_idx, column=1, value=sec['site_id']).border = thin_border
        ws.cell(row=row_idx, column=2, value=sec['site_name']).border = thin_border
        ws.cell(row=row_idx, column=3, value=' / '.join(sorted(sec['vendors']))).border = thin_border
        ws.cell(row=row_idx, column=4, value=sec['area']).border = thin_border
        ws.cell(row=row_idx, column=5, value=sec['sector']).border = thin_border

        for tb_idx, tb in enumerate(sorted_tb):
            cell = ws.cell(row=row_idx, column=6 + tb_idx)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            if tb in sec['tech_bands']:
                cell.value = 'Yes'
                cell.fill = yes_fill
                cell.font = yes_font
            else:
                cell.value = '—'
                cell.fill = no_fill
                cell.font = no_font

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 3, 10), 30)

    ws.freeze_panes = 'F2'

    summary = wb.create_sheet('Summary')
    summary_hdr_fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
    summary_hdr_font = Font(color='FFFFFF', bold=True)
    sector_col_label = 'Active Sectors' if active_only else 'Sectors'
    for ci, h in enumerate(['Tech / Band', sector_col_label, 'Total Sites'], 1):
        cell = summary.cell(row=1, column=ci, value=h)
        cell.fill = summary_hdr_fill
        cell.font = summary_hdr_font
        cell.alignment = Alignment(horizontal='center')

    for ri, tb in enumerate(sorted_tb, 2):
        sector_count = sum(1 for s in sector_list if tb in s['tech_bands'])
        site_ids = set(s['site_id'] for s in sector_list if tb in s['tech_bands'])
        summary.cell(row=ri, column=1, value=tb)
        summary.cell(row=ri, column=2, value=sector_count)
        summary.cell(row=ri, column=3, value=len(site_ids))

    for col in summary.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        summary.column_dimensions[col[0].column_letter].width = max(max_len + 4, 14)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    pull_date = datetime.now().strftime('%Y-%m-%d')
    if active_only:
        filename = f"Sector_Health_{pull_date}.xlsx"
    else:
        filename = f"Sector_Health_All_Cells_{pull_date}.xlsx"
    return buf, filename, len(sector_list)


def _generate_sector_health():
    return _generate_sector_health_excel(active_only=True)


def _generate_sector_health_all():
    return _generate_sector_health_excel(active_only=False)


REPORT_TYPES = {
    'site_inventory':     ('Site Inventory',       _generate_site_inventory),
    'pci_conflicts':      ('Conflict Report',      _generate_pci_conflicts),
    'config_versions':    ('Configuration Log',   _generate_config_versions_report),
    'sector_health':      ('Sector Health',        _generate_sector_health),
    'sector_health_all':  ('Sector Health (All Cells)', _generate_sector_health_all),
}

# Legacy report type id from before rename
_REPORT_TYPE_ALIASES = {'sector_coverage': 'sector_health'}


# ── API: generate on-demand ───────────────────────────────────────────────────

@reports_bp.route('/api/reports/generate', methods=['POST'])
@login_required
def generate_report():
    user = get_current_user()
    data = request.get_json()
    report_type = data.get('report_type', '')
    report_type = _REPORT_TYPE_ALIASES.get(report_type, report_type)

    if report_type not in REPORT_TYPES:
        return jsonify({'error': f'Unknown report type: {report_type}'}), 400

    label, generator = REPORT_TYPES[report_type]
    try:
        if report_type == 'site_inventory':
            technology = str(data.get('technology', 'all') or 'all')
            buf, filename, row_count = generator(technology)
        elif report_type == 'pci_conflicts':
            technology = str(data.get('technology', '4G') or '4G')
            pci_strict = str(data.get('strictness', '') or '').strip() or None
            buf, filename, row_count = generator(technology, pci_strict)
        else:
            buf, filename, row_count = generator()
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    # Save to disk and record in archive
    file_path = os.path.join(REPORTS_DIR, filename)
    with open(file_path, 'wb') as fh:
        fh.write(buf.getvalue())

    conn = _ncm()
    params = (filename, report_type, file_path, os.path.getsize(file_path), user['id'])
    if isinstance(conn, sqlite3.Connection):
        execute_query(
            conn,
            '''
            INSERT INTO report_archive (report_name, report_type, file_path, file_size, generated_by)
            VALUES (?, ?, ?, ?, ?)
            ''',
            params,
        )
        archive_id = execute_query(conn, 'SELECT last_insert_rowid()', ()).fetchone()[0]
    else:
        archive_id = execute_query(
            conn,
            '''
            INSERT INTO report_archive (report_name, report_type, file_path, file_size, generated_by)
            VALUES (?, ?, ?, ?, ?) RETURNING id
            ''',
            params,
        ).fetchone()['id']
    conn.commit()
    conn.close()

    log_activity(user['id'], 'report_generated', f'Generated {label} report ({row_count} rows)')
    return jsonify({'success': True, 'archive_id': archive_id, 'filename': filename, 'rows': row_count})


# ── API: download report ──────────────────────────────────────────────────────

@reports_bp.route('/api/reports/download/<int:report_id>')
@login_required
def download_report(report_id):
    conn = _ncm()
    row = execute_query(conn, 'SELECT * FROM report_archive WHERE id = ?', (report_id,)).fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'Report not found'}), 404
    row = dict(row)
    if not os.path.exists(row['file_path']):
        return jsonify({'error': 'Report file no longer exists on disk'}), 404
    return send_file(row['file_path'], as_attachment=True, download_name=row['report_name'],
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── API: report archive ───────────────────────────────────────────────────────

@reports_bp.route('/api/reports/archive')
@login_required
def report_archive():
    conn = _ncm()
    rows = execute_query(conn, '''
        SELECT ra.id, ra.report_name, ra.report_type, ra.file_size, ra.generated_at,
               u.username as generated_by_name
        FROM report_archive ra
        LEFT JOIN users u ON ra.generated_by = u.id
        ORDER BY ra.generated_at DESC
        LIMIT 100
    ''', ()).fetchall()
    conn.close()
    return jsonify({'success': True, 'reports': [dict(r) for r in rows]})


# ── API: delete report from archive ──────────────────────────────────────────

@reports_bp.route('/api/reports/archive/<int:report_id>', methods=['DELETE'])
@login_required
def delete_report(report_id):
    user = get_current_user()
    if not user or str(user.get('role', '')).strip().lower() != 'admin':
        return jsonify({'error': 'Admin privileges required to delete archived reports'}), 403

    conn = _ncm()
    row = execute_query(conn, 'SELECT * FROM report_archive WHERE id = ?', (report_id,)).fetchone()
    if row:
        try:
            if os.path.exists(row['file_path']):
                os.remove(row['file_path'])
        except OSError:
            pass
        execute_query(conn, 'DELETE FROM report_archive WHERE id = ?', (report_id,))
        conn.commit()
    conn.close()
    return jsonify({'success': True})


# ── API: list report types ────────────────────────────────────────────────────

@reports_bp.route('/api/reports/types')
@login_required
def report_types():
    return jsonify({
        'success': True,
        'types': [{'id': k, 'label': v[0]} for k, v in REPORT_TYPES.items()]
    })
