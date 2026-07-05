"""
Network Map Routes
Handles network visualization and KPI display
"""

from flask import Blueprint, request, jsonify, render_template, redirect, url_for, make_response
import gzip
import json
from functools import wraps
from collections import defaultdict
import math
import sqlite3
import os
import sys
import re

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from sync_config import (
    HUAWEI_NEIGHBOR_RAW_DB,
    NOKIA_PM_DB,
    HUAWEI_PM_DB,
    NEIGHBOR_KPI_DB,
    pm_table_name,
)
from db.runtime import connect_metadata, execute_query
from .neighbor_raw_linking import build_raw_neighbor_lines, neighbor_ho_failures
from .repeater_loader import load_all_repeaters, repeaters_for_map
from database_enhanced import get_user_by_session, log_activity
from modules.sync.metadata_active_sql import (
    _STATUS_2G,
    _STATUS_3G_FDD,
    _STATUS_4G_FDD,
    _STATUS_4G_TDD,
    _STATUS_5G,
)

_CLUSTER_AREA = {
    3: 'East Amman', 13: 'East Amman', 17: 'East Amman', 21: 'East Amman',
    23: 'East Amman', 27: 'East Amman', 48: 'East Amman', 49: 'East Amman',
    50: 'East Amman', 51: 'East Amman', 52: 'East Amman', 54: 'East Amman',
    10: 'East Jordan', 11: 'East Jordan', 19: 'East Jordan', 28: 'East Jordan',
    31: 'East Jordan', 42: 'East Jordan', 43: 'East Jordan', 47: 'East Jordan',
    1: 'South Amman', 6: 'South Amman', 9: 'South Amman', 18: 'South Amman',
    30: 'South Amman', 36: 'South Amman', 38: 'South Amman', 39: 'South Amman',
    53: 'South Amman', 57: 'South Amman', 59: 'South Amman',
    7: 'South Jordan', 8: 'South Jordan', 12: 'South Jordan', 15: 'South Jordan',
    33: 'South Jordan', 41: 'South Jordan', 58: 'South Jordan',
    2: 'West Amman', 5: 'West Amman', 16: 'West Amman', 20: 'West Amman',
    22: 'West Amman', 25: 'West Amman', 26: 'West Amman', 32: 'West Amman',
    35: 'West Amman', 40: 'West Amman', 55: 'West Amman', 56: 'West Amman',
    4: 'North Jordan', 14: 'North Jordan', 24: 'North Jordan', 29: 'North Jordan',
    34: 'North Jordan', 37: 'North Jordan', 44: 'North Jordan', 45: 'North Jordan',
    46: 'North Jordan', 65: 'North Jordan',
}


def _derive_cluster_area(site_id: object) -> tuple[int | None, str]:
    try:
        cluster = int(site_id) // 100
    except (TypeError, ValueError):
        return None, 'Unknown'
    return cluster, _CLUSTER_AREA.get(cluster, 'Unknown')


def _haversine_km(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    r = 6371.0
    d_lat = math.radians(lat2 - lat1)
    d_lng = math.radians(lng2 - lng1)
    a = (
        math.sin(d_lat / 2) ** 2
        + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(d_lng / 2) ** 2
    )
    return 2 * r * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def _json_gzip_response(payload: dict, *, min_compress_bytes: int = 16_384):
    """JSON response with optional gzip when the client accepts it."""
    body = json.dumps(payload, ensure_ascii=False, separators=(',', ':')).encode('utf-8')
    accept = (request.headers.get('Accept-Encoding') or '').lower()
    if 'gzip' in accept and len(body) >= min_compress_bytes:
        resp = make_response(gzip.compress(body, compresslevel=6))
        resp.headers['Content-Type'] = 'application/json; charset=utf-8'
        resp.headers['Content-Encoding'] = 'gzip'
        resp.headers['Vary'] = 'Accept-Encoding'
        return resp
    resp = make_response(body)
    resp.headers['Content-Type'] = 'application/json; charset=utf-8'
    return resp


def _assign_area_from_nearest_known(sites: list[dict]) -> None:
    known = []
    unknown_idx = []
    for i, s in enumerate(sites):
        lat = s.get('latitude')
        lng = s.get('longitude')
        if lat is None or lng is None:
            continue
        try:
            lat_f = float(lat)
            lng_f = float(lng)
        except (TypeError, ValueError):
            continue
        if s.get('area') and s.get('area') != 'Unknown':
            known.append((lat_f, lng_f, s.get('area')))
        elif s.get('area') == 'Unknown':
            unknown_idx.append((i, lat_f, lng_f))

    if not known:
        return

    for idx, lat_f, lng_f in unknown_idx:
        nearest_area = None
        nearest_dist = None
        for k_lat, k_lng, k_area in known:
            d = _haversine_km(lat_f, lng_f, k_lat, k_lng)
            if nearest_dist is None or d < nearest_dist:
                nearest_dist = d
                nearest_area = k_area
        if nearest_area:
            sites[idx]['area'] = nearest_area


def _per_tech_union_sql(tech: str | None = None) -> tuple[str, list]:
    """
    Build a UNION ALL subquery over per-technology tables in metadata.db.
    Returns (sql, params) where sql is a SELECT ... UNION ALL ... string producing:
      cell_name, site_id, site_name, technology, vendor,
      latitude, longitude, azimuth, mechanical_tilt, electrical_tilt,
      frequency_band, pci, activity_status, status
    ``activity_status`` / ``status``: Active | Inactive from sync/metadata_active_sql.py
    (2G: Huawei active_state, Nokia admin_state; 3G/5G: Huawei activated / Nokia unlocked;
     4G-FDD: Nokia Unlocked or CELL_ACTIVE on active_state, Huawei activated or CELL_ACTIVE;
     4G-TDD: Huawei CELL_ACTIVE, Nokia admin_state Unlocked).
    """
    parts = {
        '2G': f"""
            SELECT
                cell_name,
                site_id      AS site_id,
                site_name    AS site_name,
                '2G'         AS technology,
                vendor,
                CAST(lat  AS REAL) AS latitude,
                CAST(long AS REAL) AS longitude,
                CAST(azimuth AS REAL) AS azimuth,
                CAST(mtilt   AS REAL) AS mechanical_tilt,
                CAST(etilt   AS REAL) AS electrical_tilt,
                frequency_band AS frequency_band,
                CAST(COALESCE(NULLIF(TRIM(bcch), ''), NULLIF(TRIM(bcc), '')) AS INTEGER) AS pci,
                {_STATUS_2G} AS activity_status,
                {_STATUS_2G} AS status
            FROM cells_2g
        """,
        '3G': f"""
            SELECT
                cell_name,
                nodeb_id     AS site_id,
                nodeb_name   AS site_name,
                '3G'         AS technology,
                vendor,
                CAST(lat  AS REAL) AS latitude,
                CAST(long AS REAL) AS longitude,
                CAST(azimuth AS REAL) AS azimuth,
                CAST(mtilt   AS REAL) AS mechanical_tilt,
                CAST(etilt   AS REAL) AS electrical_tilt,
                dl_uarfcn     AS frequency_band,
                CAST(psc AS INTEGER) AS pci,
                {_STATUS_3G_FDD} AS activity_status,
                {_STATUS_3G_FDD} AS status
            FROM cells_3g
        """,
        '4G-FDD': f"""
            SELECT
                cell_name,
                enb_id_actual AS site_id,
                enb_name      AS site_name,
                '4G-FDD'      AS technology,
                vendor,
                CAST(lat  AS REAL) AS latitude,
                CAST(long AS REAL) AS longitude,
                CAST(azimuth AS REAL) AS azimuth,
                CAST(mtilt   AS REAL) AS mechanical_tilt,
                CAST(etilt   AS REAL) AS electrical_tilt,
                band          AS frequency_band,
                CAST(pci AS INTEGER) AS pci,
                {_STATUS_4G_FDD} AS activity_status,
                {_STATUS_4G_FDD} AS status
            FROM cells_4g_fdd
        """,
        '4G-TDD': f"""
            SELECT
                cell_name,
                enb_id_actual AS site_id,
                enb_name      AS site_name,
                '4G-TDD'      AS technology,
                vendor,
                CAST(lat  AS REAL) AS latitude,
                CAST(long AS REAL) AS longitude,
                CAST(azimuth AS REAL) AS azimuth,
                CAST(mtilt   AS REAL) AS mechanical_tilt,
                CAST(etilt   AS REAL) AS electrical_tilt,
                band          AS frequency_band,
                CAST(pci AS INTEGER) AS pci,
                {_STATUS_4G_TDD} AS activity_status,
                {_STATUS_4G_TDD} AS status
            FROM cells_4g_tdd
        """,
        '5G': f"""
            SELECT
                cell_name,
                gnb_id_actual AS site_id,
                gnb_name      AS site_name,
                '5G'          AS technology,
                vendor,
                CAST(lat  AS REAL) AS latitude,
                CAST(long AS REAL) AS longitude,
                CAST(azimuth AS REAL) AS azimuth,
                CAST(mtilt   AS REAL) AS mechanical_tilt,
                CAST(etilt   AS REAL) AS electrical_tilt,
                bw            AS frequency_band,
                CAST(pci AS INTEGER) AS pci,
                {_STATUS_5G} AS activity_status,
                {_STATUS_5G} AS status
            FROM cells_5g
        """,
    }

    if tech and tech != 'all':
        # Tech can come from UI as '2G','3G','4G-FDD','4G-TDD','5G'
        if tech not in parts:
            return "SELECT NULL AS cell_name, NULL AS site_id, NULL AS site_name, NULL AS technology, NULL AS vendor, NULL AS latitude, NULL AS longitude, NULL AS azimuth, NULL AS mechanical_tilt, NULL AS electrical_tilt, NULL AS frequency_band, NULL AS pci, NULL AS activity_status, NULL AS status WHERE 1=0", []
        return parts[tech], []

    # All technologies
    return " UNION ALL ".join(parts[t] for t in ['2G', '3G', '4G-FDD', '4G-TDD', '5G']), []


def _normalize_ui_map_tech_token(tech: str) -> str:
    """Map chip / query-param variants to canonical tokens (case, unicode dashes, spacing)."""
    s = (tech or "").strip()
    for dash in ("\u2011", "\u2013", "\u2212"):
        s = s.replace(dash, "-")
    s = re.sub(r"\s+", " ", s)
    low = s.lower()
    chips = {
        "2g-2g": "2G-2G",
        "3g-3g": "3G-3G",
        "4g-4g intra-enb": "4G-4G",
        "4g-4g inter-enb": "4G-4G",
        # legacy UI tokens
        "4g-4g intra": "4G-4G",
        "4g-4g inter": "4G-4G",
        "4g-4g": "4G-4G",
        "2g": "2G",
        "3g": "3G",
        "4g-fdd": "4G-FDD",
        "4g-tdd": "4G-TDD",
        "5g": "5G",
    }
    return chips.get(low, s)


def _union_sql_for_map_filter(tech: str | None = None) -> tuple[str, list]:
    """
    UNION for Network Map list/search/export when the UI uses relation-style chips:
    All → 2G + 3G + LTE FDD/TDD (no 5G on this map); 2G-2G / 3G-3G / 4G-4G use LTE FDD only (no TDD in neighbor HO scope).
    """
    raw = (tech or "").strip()
    if not raw or raw.lower() == "all":
        s2, _ = _per_tech_union_sql("2G")
        s3, _ = _per_tech_union_sql("3G")
        sf, _ = _per_tech_union_sql("4G-FDD")
        st, _ = _per_tech_union_sql("4G-TDD")
        return " UNION ALL ".join([s2, s3, sf, st]), []
    t = _normalize_ui_map_tech_token(raw)
    if t == "2G-2G":
        return _per_tech_union_sql("2G")
    if t == "3G-3G":
        return _per_tech_union_sql("3G")
    if t in ("4G-4G", "4G-4G Intra-eNB", "4G-4G Inter-eNB", "4G-4G Intra", "4G-4G Inter"):
        sf, _ = _per_tech_union_sql("4G-FDD")
        return sf, []
    if t in ("2G", "3G", "4G-FDD", "4G-TDD", "5G"):
        return _per_tech_union_sql(t)
    return (
        "SELECT NULL AS cell_name, NULL AS site_id, NULL AS site_name, NULL AS technology, NULL AS vendor, "
        "NULL AS latitude, NULL AS longitude, NULL AS azimuth, NULL AS mechanical_tilt, NULL AS electrical_tilt, "
        "NULL AS frequency_band, NULL AS pci, NULL AS activity_status, NULL AS status WHERE 1=0",
        [],
    )


def _map_request_tech_filter_to_bcch_band_clause(tech: str) -> bool:
    """Tech-specific dropdown always scopes band / UARFCN (``frequency_band``), not PCI/BCCH/PSC."""
    return False


def _cell_kpi_sql_technologies(req_tech: str) -> list[str] | None:
    """Map map-page technology chip to v.technology IN (...) for KPI lookup."""
    t = _normalize_ui_map_tech_token((req_tech or "").strip())
    if not t:
        return None
    if t == "2G-2G":
        return ["2G"]
    if t == "3G-3G":
        return ["3G"]
    if t in ("4G-4G", "4G-4G Intra-eNB", "4G-4G Inter-eNB", "4G-4G Intra", "4G-4G Inter"):
        return ["4G-FDD"]
    return [t]


# KML export: match map.js sector geometry (wedges) and tech colours
_KML_TECH_COLORS = {
    '2G': '#7f8c8d',
    '3G': '#27ae60',
    '4G-FDD': '#1a5276',
    '4G-TDD': '#148f77',
    '5G': '#9b59b6',
}
_KML_SECTOR_RADIUS_M = 240.0
_KML_SECTOR_BEAMWIDTH = 65.0


def _kml_color_aabbggrr(hex6: str, alpha: int = 255) -> str:
    """KML colour: aabbggrr (alpha, blue, green, red)."""
    h = hex6.lstrip('#')
    if len(h) != 6:
        return 'FF000000'
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return f'{max(0, min(255, alpha)):02X}{b:02X}{g:02X}{r:02X}'


def _kml_cell_active(row) -> bool:
    raw = (row['activity_status'] if row['activity_status'] is not None else row['status']) or ''
    return str(raw).strip().lower() == 'active'


def _kml_wedge_ring_coords(site_lat: float, site_lng: float, azimuth: float,
                           half_beam: float = _KML_SECTOR_BEAMWIDTH / 2.0,
                           radius_m: float = _KML_SECTOR_RADIUS_M,
                           step_deg: float = 3.0) -> list[tuple[float, float]]:
    """Closed (lon, lat) ring, same construction as network_map/static/map.js."""
    r_lat = radius_m / 111320.0
    c = math.cos(math.radians(site_lat))
    denom = 111320.0 * (c if abs(c) > 1e-9 else 1e-9)
    r_lng = radius_m / denom
    coords: list[tuple[float, float]] = [(site_lng, site_lat)]
    a = azimuth - half_beam
    end = azimuth + half_beam + 0.001
    while a <= end:
        rad = math.radians(a)
        lat = site_lat + r_lat * math.cos(rad)
        lng = site_lng + r_lng * math.sin(rad)
        coords.append((lng, lat))
        a += step_deg
    coords.append((site_lng, site_lat))
    return coords


def _kml_circle_ring_coords(lat0: float, lng0: float, radius_m: float, n: int = 36,
                            reverse: bool = False) -> list[tuple[float, float]]:
    """One closed ring (lon, lat); reverse=True for inner donut hole winding."""
    r_lat = radius_m / 111320.0
    c = math.cos(math.radians(lat0))
    r_lng = radius_m / (111320.0 * (c if abs(c) > 1e-9 else 1e-9))
    idx = range(n)
    if reverse:
        idx = range(n - 1, -1, -1)
    pts: list[tuple[float, float]] = []
    for i in idx:
        ang = 2 * math.pi * (i / n)
        lat = lat0 + r_lat * math.cos(ang)
        lng = lng0 + r_lng * math.sin(ang)
        pts.append((lng, lat))
    pts.append(pts[0])
    return pts


def _kml_ring_to_coordinates(coords: list[tuple[float, float]]) -> str:
    return ' '.join(f'{lng},{lat},0' for lng, lat in coords)


def _kml_cdata_fragment(t: str) -> str:
    """CDATA cannot contain the literal sequence ]]> ."""
    return (t or '').replace(']]>', ']] >')


def _kml_group_cells_into_wedges(cells: list) -> dict[tuple[str, float], list]:
    """Technology + 5° azimuth bucket → cells (same keying as map.js)."""
    by_key: dict[tuple[str, float], list] = defaultdict(list)
    for c in cells:
        if not _kml_cell_active(c):
            continue
        az = c['azimuth']
        if az is None:
            continue
        try:
            a = float(az)
        except (TypeError, ValueError):
            continue
        if not math.isfinite(a):
            continue
        bucket = round(a / 5.0) * 5.0
        tech = c['technology'] or 'Unknown'
        by_key[(tech, bucket)].append(c)
    return by_key


def _metadata_table_for_tech(tech: str) -> str | None:
    mapping = {
        '2G': 'cells_2g',
        '3G': 'cells_3g',
        '4G-FDD': 'cells_4g_fdd',
        '4G-TDD': 'cells_4g_tdd',
        '5G': 'cells_5g',
    }
    return mapping.get((tech or '').strip())


def _norm_cell_key(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def _sqlite_row(conn) -> None:
    if isinstance(conn, sqlite3.Connection):
        conn.row_factory = sqlite3.Row


def _safe_int(value: object, default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _safe_float(value: object, default: float) -> float:
    try:
        if value is None or str(value).strip() == "":
            return default
        return float(str(value).replace(",", ""))
    except (TypeError, ValueError):
        return default


def _load_neighbor_coords(vendor: str, technology: str) -> dict[str, dict]:
    conn = connect_metadata()
    _sqlite_row(conn)
    try:
        union_sql, params = _union_sql_for_map_filter(technology)
        rows = execute_query(conn,
            f"""
            SELECT
                v.cell_name,
                v.site_id,
                v.latitude,
                v.longitude,
                s.region
            FROM ({union_sql}) v
            LEFT JOIN sites s ON s.site_id = v.site_id
            WHERE LOWER(TRIM(COALESCE(v.vendor, ''))) = LOWER(TRIM(?))
              AND v.latitude IS NOT NULL
              AND v.longitude IS NOT NULL
            """,
            params + [vendor],
        ).fetchall()
        out: dict[str, dict] = {}
        for r in rows:
            key = _norm_cell_key(r["cell_name"])
            if key and key not in out:
                out[key] = {
                    "cell_name": r["cell_name"],
                    "site_id": r["site_id"],
                    "lat": float(r["latitude"]),
                    "lng": float(r["longitude"]),
                    "region": r["region"],
                    "cluster": (int(r["site_id"]) // 100) if str(r["site_id"]).isdigit() else None,
                }
        return out
    finally:
        conn.close()


def _neighbor_table_exists(conn, name: str) -> bool:
    row = execute_query(
        conn,
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
        (name,),
    ).fetchone()
    return bool(row)


def _neighbor_table_non_empty(conn: sqlite3.Connection, name: str) -> bool:
    """True if table exists and has at least one row (empty slim shells must not block wide fallback)."""
    if not _neighbor_table_exists(conn, name):
        return False
    row = conn.execute(f'SELECT 1 AS x FROM "{name}" LIMIT 1').fetchone()
    return row is not None


def _neighbor_raw_base_prefix_for_slug(vendor_slug: str) -> str:
    return "huawei_neighbor" if (vendor_slug or "").strip().lower() == "huawei" else "nokia_neighbor"


def _neighbor_raw_table_name_for_technology(technology: str, prefix: str) -> str | None:
    tok = _normalize_ui_map_tech_token((technology or "").strip())
    t = tok.upper()
    if tok == "2G-2G" or t.startswith("2G"):
        return f"{prefix}_2g"
    if tok == "3G-3G" or t.startswith("3G"):
        return f"{prefix}_3g"
    if "4G" in t or "LTE" in t:
        return f"{prefix}_4g"
    return None


def _resolve_raw_neighbor_tables_for_vendor(
    conn: sqlite3.Connection, technology: str, vendor_slug: str
) -> list[str]:
    """Resolve one logical RAT to one or more raw tables.

    4G is intentionally a single UI/API technology. Prefer a populated combined
    ``<prefix>_4g`` table, but also support old split intra/inter tables as
    implementation details if a loader produced them.
    """
    prefix = _neighbor_raw_base_prefix_for_slug(vendor_slug)
    preferred = _neighbor_raw_table_name_for_technology(technology, prefix)
    if not preferred:
        return []
    t = (technology or "").strip().upper()
    if "4G" not in t and "LTE" not in t:
        return [preferred] if _neighbor_table_non_empty(conn, preferred) else []

    tables: list[str] = []
    combined = f"{prefix}_4g"
    if _neighbor_table_non_empty(conn, combined):
        tables.append(combined)
    for split in (f"{prefix}_4g_intra", f"{prefix}_4g_inter"):
        if _neighbor_table_non_empty(conn, split):
            tables.append(split)
    return tables


def _resolve_raw_neighbor_table_for_vendor(
    conn: sqlite3.Connection, technology: str, vendor_slug: str
) -> str | None:
    tables = _resolve_raw_neighbor_tables_for_vendor(conn, technology, vendor_slug)
    return tables[0] if tables else None


def _resolve_huawei_neighbor_export_table(_conn: sqlite3.Connection, technology: str) -> str | None:
    """Huawei wide DB table names (no Nokia-style intra/inter split)."""
    tok = _normalize_ui_map_tech_token((technology or "").strip())
    t = tok.upper()
    if tok == "2G-2G" or t.startswith("2G"):
        return "huawei_neighbor_export_2g"
    if tok == "3G-3G" or t.startswith("3G"):
        return "huawei_neighbor_export_3g"
    if "4G" in t or "LTE" in t:
        return "huawei_neighbor_export_4g"
    return None


def _any_raw_neighbor_table_exists(nokia_conn: sqlite3.Connection, technology: str) -> bool:
    """Nokia tables in ``neighbor_kpis.db`` or Huawei export tables in ``huawei_neighbor_raw.db``."""
    if _resolve_raw_neighbor_tables_for_vendor(nokia_conn, technology, "nokia"):
        return True
    if not os.path.isfile(HUAWEI_NEIGHBOR_RAW_DB):
        return False
    hconn = sqlite3.connect(HUAWEI_NEIGHBOR_RAW_DB, timeout=20)
    try:
        ht = _resolve_huawei_neighbor_export_table(hconn, technology)
        return bool(ht and _neighbor_table_non_empty(hconn, ht))
    finally:
        hconn.close()


def _neighbor_hourly_tech_aliases(technology: str) -> list[str]:
    """Map UI technology tokens to likely legacy ``neighbor_hourly.technology`` values."""
    tok = _normalize_ui_map_tech_token((technology or "").strip())
    t = tok.upper()
    if tok == "2G-2G" or t.startswith("2G"):
        return ["2G", "2G-2G"]
    if tok == "3G-3G" or t.startswith("3G"):
        return ["3G", "3G-3G"]
    if tok in ("4G-4G", "4G-4G Intra-eNB", "4G-4G Inter-eNB", "4G-4G Intra", "4G-4G Inter") or "4G" in t or "LTE" in t:
        return ["4G", "4G-FDD", "LTE", "4G-4G", "4G-4G Intra-eNB", "4G-4G Inter-eNB"]
    return [tok] if tok else []


def _ensure_neighbor_schema(conn: sqlite3.Connection) -> None:
    """No legacy neighbor_hourly DDL; data lives in ``nokia_neighbor_*`` / ``huawei_neighbor_*`` raw tables."""
    return

network_map_bp = Blueprint(
    'network_map', __name__,
    template_folder='templates',
    static_folder='static',
    static_url_path='/network_map/static',
)


def login_required(f):
    """Decorator to require login"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        session_token = request.cookies.get('session_token')
        if not session_token:
            return redirect(url_for('auth.login_page'))

        user = get_user_by_session(session_token)
        if not user:
            return redirect(url_for('auth.login_page'))

        request.current_user = user
        return f(*args, **kwargs)

    return decorated_function

def get_current_user():
    """Get current logged-in user"""
    session_token = request.cookies.get('session_token')
    if session_token:
        return get_user_by_session(session_token)
    return None

def format_user_data(user):
    """Format user data for templates"""
    if not user:
        return None
    if isinstance(user, dict):
        return {'username': user.get('username'), 'email': user.get('email'), 'role': user.get('role'), 'id': user.get('id')}
    return {'username': (user.get('username') if isinstance(user, dict) else user[1]), 'email': (user.get('email') if isinstance(user, dict) else user[2]), 'role': (user.get('role') if isinstance(user, dict) else user[6]), 'id': (user.get('id') if isinstance(user, dict) else user[0])}


@network_map_bp.route('/network-map')
@login_required
def network_map_page():
    """Render Network Map page"""
    user = get_current_user()
    return render_template('network_map.html', user=format_user_data(user))


@network_map_bp.route('/neighbor-analysis')
@login_required
def neighbor_analysis_page():
    """Render dedicated Neighbor Analysis page."""
    user = get_current_user()
    return render_template('neighbor_analysis.html', user=format_user_data(user))

@network_map_bp.route('/api/map/sites', methods=['GET'])
def get_all_sites():
    """Get all network sites, optionally filtered by technology.

    IMPORTANT: For Network Map we use the per-technology tables (cells_2g/3g/4g_fdd/4g_tdd/5g)
    so counts match the CSV snapshots (no de-duplication via the unified cells table).
    """
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    tech = request.args.get('tech', '').strip()
    tech_value = request.args.get('tech_value', '').strip()

    try:
        conn = connect_metadata()
        _sqlite_row(conn)

        union_sql, params = _union_sql_for_map_filter(tech if tech else None)
        scope_clause = ''
        sql_params = list(params)
        if tech_value:
            if _map_request_tech_filter_to_bcch_band_clause(tech):
                scope_clause = ' AND CAST(v.pci AS TEXT) = ?'
            else:
                scope_clause = ' AND CAST(v.frequency_band AS TEXT) = ?'
            sql_params.append(tech_value)

        cur = execute_query(conn, f'''
            WITH v AS (
                {union_sql}
            ),
            filtered AS (
                SELECT *
                FROM v
                WHERE 1=1
                  {scope_clause}
            ),
            offline AS (
                SELECT
                    site_id,
                    SUM(
                        CASE WHEN LOWER(TRIM(COALESCE(activity_status, status, ''))) = 'inactive'
                             THEN 1 ELSE 0 END
                    ) AS offline_cell_count
                FROM filtered
                GROUP BY site_id
            ),
            wedge_agg AS (
                SELECT
                    site_id,
                    technology,
                    ROUND(CAST(azimuth AS REAL) / 5.0) * 5 AS az_bin,
                    COUNT(*) AS cell_count,
                    SUM(
                        CASE WHEN LOWER(TRIM(COALESCE(activity_status, status, ''))) = 'inactive'
                             THEN 1 ELSE 0 END
                    ) AS offline_count
                FROM filtered
                WHERE azimuth IS NOT NULL
                  AND TRIM(CAST(azimuth AS TEXT)) <> ''
                GROUP BY site_id, technology,
                         ROUND(CAST(azimuth AS REAL) / 5.0) * 5
            ),
            full_sector_off AS (
                SELECT site_id, COUNT(*) AS full_sector_offline_count
                FROM wedge_agg
                WHERE cell_count > 0 AND offline_count = cell_count
                GROUP BY site_id
            )
            SELECT DISTINCT
                s.site_id, s.site_name, s.latitude, s.longitude,
                s.region, s.site_type, s.vendor, s.status,
                COALESCE(o.offline_cell_count, 0) AS offline_cell_count,
                COALESCE(fs.full_sector_offline_count, 0) AS full_sector_offline_count
            FROM filtered f
            JOIN sites s ON s.site_id = f.site_id
            LEFT JOIN offline o ON o.site_id = s.site_id
            LEFT JOIN full_sector_off fs ON fs.site_id = s.site_id
            WHERE s.latitude IS NOT NULL AND s.longitude IS NOT NULL
            ORDER BY s.site_name
        ''', sql_params)

        sites = [dict(row) for row in cur.fetchall()]
        for s in sites:
            cluster, area = _derive_cluster_area(s.get('site_id'))
            s['cluster'] = cluster
            s['area'] = area
        _assign_area_from_nearest_known(sites)
        conn.close()

        log_activity((user.get('id') if isinstance(user, dict) else user[0]), 'map_view', 'Viewed network map sites')
        return jsonify({'success': True, 'sites': sites})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@network_map_bp.route('/api/map/repeaters', methods=['GET'])
def get_map_repeaters():
    """Repeater devices from manual Excel/CSV (network-map/repeater)."""
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        repeaters, source = load_all_repeaters()
        if not repeaters and source and not os.path.isfile(source):
            return jsonify({'success': True, 'repeaters': [], 'total': 0, 'message': source})

        map_rows = repeaters_for_map(repeaters)

        log_activity(
            (user.get('id') if isinstance(user, dict) else user[0]),
            'map_view',
            f'Viewed {len(map_rows)} repeaters on network map',
        )
        return _json_gzip_response({
            'success': True,
            'repeaters': map_rows,
            'total': len(map_rows),
            'source_file': source if repeaters else None,
            'message': source if not repeaters else None,
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@network_map_bp.route('/api/map/repeater/<path:repeater_id>', methods=['GET'])
def get_map_repeater_detail(repeater_id):
    """Detail for one repeater record."""
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        repeaters, _source = load_all_repeaters()
        rid = str(repeater_id or '').strip()
        match = next((r for r in repeaters if str(r.get('repeater_id')) == rid), None)
        if not match:
            return jsonify({'error': 'Repeater not found'}), 404
        return jsonify({'success': True, 'repeater': match})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@network_map_bp.route('/api/map/tech-filter-options', methods=['GET'])
def get_tech_filter_options():
    """Get dynamic dropdown values for tech-specific filters (UARFCN/Band/BCCH)."""
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    tech = request.args.get('tech', '').strip()
    if tech in ("2G-2G",):
        tech = "2G"
    elif tech in ("3G-3G",):
        tech = "3G"
    if tech not in ('2G', '3G', '4G-FDD', '4G-TDD'):
        return jsonify({'success': True, 'label': '', 'values': []})

    try:
        conn = connect_metadata()
        _sqlite_row(conn)
        union_sql, params = _per_tech_union_sql(tech)

        if tech == '2G':
            rows = execute_query(conn, f'''
                SELECT DISTINCT CAST(v.frequency_band AS TEXT) AS val
                FROM ({union_sql}) v
                WHERE v.frequency_band IS NOT NULL AND TRIM(CAST(v.frequency_band AS TEXT)) <> ''
                ORDER BY v.frequency_band
            ''', params).fetchall()
            label = 'Band'
        elif tech == '3G':
            rows = execute_query(conn, f'''
                SELECT DISTINCT CAST(v.frequency_band AS TEXT) AS val
                FROM ({union_sql}) v
                WHERE v.frequency_band IS NOT NULL AND TRIM(CAST(v.frequency_band AS TEXT)) <> ''
                ORDER BY CAST(v.frequency_band AS INTEGER), v.frequency_band
            ''', params).fetchall()
            label = 'UARFCN'
        else:
            rows = execute_query(conn, f'''
                SELECT DISTINCT CAST(v.frequency_band AS TEXT) AS val
                FROM ({union_sql}) v
                WHERE v.frequency_band IS NOT NULL AND TRIM(CAST(v.frequency_band AS TEXT)) <> ''
                ORDER BY v.frequency_band
            ''', params).fetchall()
            label = 'Band'

        conn.close()
        values = [r['val'] for r in rows if r['val'] is not None]
        return jsonify({'success': True, 'label': label, 'values': values})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@network_map_bp.route('/api/map/site/<site_id>', methods=['GET'])
def get_site_details(site_id):
    """Get detailed information about a specific site"""
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        conn = connect_metadata()
        _sqlite_row(conn)

        site_cur = execute_query(conn, '''
            SELECT site_id, site_name, latitude, longitude, region, site_type, vendor, status
            FROM sites
            WHERE site_id = ?
        ''', (site_id,))

        site = site_cur.fetchone()
        if not site:
            conn.close()
            return jsonify({'error': 'Site not found'}), 404

        site_data = dict(site)
        cluster, area = _derive_cluster_area(site_data.get('site_id'))
        site_data['cluster'] = cluster
        site_data['area'] = area

        # Cells are the sectors — pull from per-technology tables (no de-dupe).
        union_sql, params = _per_tech_union_sql(None)
        cells_cur = execute_query(conn, f'''
            SELECT
                cell_name,
                technology,
                vendor,
                frequency_band,
                azimuth,
                mechanical_tilt,
                electrical_tilt,
                pci,
                activity_status,
                status
            FROM ({union_sql}) v
            WHERE v.site_id = ?
            ORDER BY technology, cell_name
        ''', params + [site_id])
        cells = []
        for row in cells_cur.fetchall():
            d = dict(row)
            # Map JS expects an identifier; per-tech tables don't have integer ids.
            d['cell_id'] = d.get('cell_name')  # backward-compatible field name
            cells.append(d)
        site_data['cells'] = cells

        conn.close()

        log_activity((user.get('id') if isinstance(user, dict) else user[0]), 'site_view', f'Viewed site {site_id}')
        return jsonify({'success': True, 'site': site_data})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@network_map_bp.route('/api/map/cells/wedge-data', methods=['GET'])
@login_required
def cells_wedge_data():
    """Return lat/lng/azimuth and display fields for drawing sector wedges for named cells (neighbor scope)."""
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    technology = _normalize_ui_map_tech_token((request.args.get('technology') or '').strip())
    raw = request.args.getlist('cell')
    if not technology or technology.lower() == 'all' or not raw:
        return jsonify({'success': True, 'cells': []})

    names_norm: list[str] = []
    seen: set[str] = set()
    for c in raw:
        k = _norm_cell_key(c)
        if not k or k in seen:
            continue
        seen.add(k)
        names_norm.append(k)
        if len(names_norm) >= 280:
            break

    try:
        union_sql, u_params = _union_sql_for_map_filter(technology)
        ph = ",".join(["?"] * len(names_norm))
        sql = f"""
            SELECT
                v.cell_name,
                v.site_id,
                v.site_name,
                v.latitude,
                v.longitude,
                v.azimuth,
                v.technology,
                v.vendor,
                v.frequency_band,
                v.pci,
                v.activity_status,
                v.status,
                v.mechanical_tilt,
                v.electrical_tilt
            FROM ({union_sql}) v
            WHERE LOWER(TRIM(v.cell_name)) IN ({ph})
        """
        conn = connect_metadata()
        _sqlite_row(conn)
        rows = [
            dict(r)
            for r in execute_query(conn, sql, u_params + names_norm).fetchall()
            if r["latitude"] is not None and r["longitude"] is not None
        ]
        conn.close()
        return jsonify({'success': True, 'cells': rows})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@network_map_bp.route('/api/map/cell/<int:cell_id>/kpis', methods=['GET'])
def get_cell_kpis(cell_id):
    """Get the latest KPI snapshot for a cell from the appropriate PM database."""
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        meta_conn = connect_metadata()
        _sqlite_row(meta_conn)
        cell = execute_query(meta_conn, '''
            SELECT c.cell_id, c.cell_name, c.technology, c.vendor,
                   c.azimuth, c.mechanical_tilt, c.electrical_tilt, c.pci,
                   st.site_id, st.site_name, st.region
            FROM cells c
            LEFT JOIN sites st ON c.site_id = st.site_id
            WHERE c.cell_id = ?
        ''', (cell_id,)).fetchone()
        meta_conn.close()

        if not cell:
            return jsonify({'success': False, 'error': 'Cell not found'}), 404

        cell_data = dict(cell)
        vendor    = cell_data.get('vendor', '')
        cell_name = cell_data['cell_name']
        cell_tech = cell_data.get('technology', '4G')
        pm_db     = NOKIA_PM_DB if vendor == 'Nokia' else HUAWEI_PM_DB
        if vendor == 'Huawei':
            from modules.sync.pm_processor import huawei_pm_table_for_cell

            table = huawei_pm_table_for_cell(cell_name, cell_tech, pm_db)
        else:
            table = pm_table_name(cell_tech)

        try:
            if not table:
                cell_data['kpis'] = None
            else:
                pm_conn = sqlite3.connect(pm_db)
                pm_conn.row_factory = sqlite3.Row
                try:
                    kpi = execute_query(pm_conn, f'''
                        SELECT *
                        FROM "{table}"
                        WHERE cell_name = ?
                        ORDER BY timestamp DESC
                        LIMIT 1
                    ''', (cell_name,)).fetchone()
                    cell_data['kpis'] = dict(kpi) if kpi else None
                finally:
                    pm_conn.close()
        except Exception:
            cell_data['kpis'] = None

        log_activity((user.get('id') if isinstance(user, dict) else user[0]),
                     'cell_kpi_view', f'Viewed KPIs for cell {cell_name}')
        return jsonify({'success': True, 'cell': cell_data})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@network_map_bp.route('/api/map/cell/kpis', methods=['GET'])
def get_cell_kpis_by_name():
    """
    KPI endpoint keyed by cell_name (used by the Network Map per-tech tables).
    Query param: ?cell_name=<name>
    """
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    cell_name = (request.args.get('cell_name') or '').strip()
    req_tech = _normalize_ui_map_tech_token((request.args.get('technology') or '').strip())
    req_vendor = (request.args.get('vendor') or '').strip()
    if not cell_name:
        return jsonify({'error': 'cell_name is required'}), 400

    try:
        conn = connect_metadata()
        _sqlite_row(conn)
        union_sql, params = (
            _per_tech_union_sql(None)
            if not req_tech
            else _union_sql_for_map_filter(req_tech)
        )
        where_filters = ['v.cell_name = ?']
        where_params = [cell_name]
        if req_vendor:
            where_filters.append("LOWER(TRIM(COALESCE(v.vendor,''))) = LOWER(TRIM(?))")
            where_params.append(req_vendor)
        if req_tech:
            tech_vals = _cell_kpi_sql_technologies(req_tech)
            if tech_vals and len(tech_vals) == 1:
                where_filters.append("v.technology = ?")
                where_params.append(tech_vals[0])
            elif tech_vals and len(tech_vals) > 1:
                ph = ",".join("?" * len(tech_vals))
                where_filters.append(f"v.technology IN ({ph})")
                where_params.extend(tech_vals)
        row = execute_query(conn, f'''
            SELECT
                v.cell_name, v.technology, v.vendor,
                v.frequency_band, v.azimuth, v.mechanical_tilt, v.electrical_tilt,
                v.pci, v.activity_status, v.status,
                s.site_id, s.site_name, s.region
            FROM ({union_sql}) v
            LEFT JOIN sites s ON s.site_id = v.site_id
            WHERE {' AND '.join(where_filters)}
            LIMIT 1
        ''', params + where_params).fetchone()
        conn.close()

        if not row:
            return jsonify({'success': False, 'error': 'Cell not found'}), 404

        cell_data = dict(row)

        vendor = cell_data.get('vendor', '')
        tech   = cell_data.get('technology', '4G-FDD')
        pm_db  = NOKIA_PM_DB if vendor == 'Nokia' else HUAWEI_PM_DB
        if vendor == 'Huawei':
            from modules.sync.pm_processor import huawei_pm_table_for_cell

            table = huawei_pm_table_for_cell(cell_name, tech, pm_db)
        else:
            table = pm_table_name(tech)

        # Add full metadata row from the technology-specific staging table so the
        # wedge-cell popup can show all available fields for that exact cell.
        cell_data['metadata'] = None
        meta_table = _metadata_table_for_tech(tech)
        if meta_table:
            try:
                meta_conn = connect_metadata()
                _sqlite_row(meta_conn)
                md = execute_query(meta_conn,
                    f'SELECT * FROM "{meta_table}" WHERE cell_name = ? LIMIT 1',
                    (cell_name,)
                ).fetchone()
                meta_conn.close()
                cell_data['metadata'] = dict(md) if md else None
            except Exception:
                cell_data['metadata'] = None

        try:
            if not table:
                cell_data['kpis'] = None
            else:
                pm_conn = sqlite3.connect(pm_db)
                pm_conn.row_factory = sqlite3.Row
                try:
                    kpi = execute_query(pm_conn, f'''
                        SELECT *
                        FROM "{table}"
                        WHERE cell_name = ?
                        ORDER BY timestamp DESC
                        LIMIT 1
                    ''', (cell_name,)).fetchone()
                    cell_data['kpis'] = dict(kpi) if kpi else None
                finally:
                    pm_conn.close()
        except Exception:
            cell_data['kpis'] = None

        log_activity((user.get('id') if isinstance(user, dict) else user[0]),
                     'cell_kpi_view', f'Viewed KPIs for cell {cell_name}')
        return jsonify({'success': True, 'cell': cell_data})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@network_map_bp.route('/api/map/stats', methods=['GET'])
def get_network_stats():
    """Get overall network statistics for Network Map (per-tech tables)."""
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        conn = connect_metadata()
        _sqlite_row(conn)

        def _count_active_distinct_sites(union_sql: str, params: list) -> int:
            val = execute_query(conn, f'''
                SELECT COUNT(DISTINCT v.site_id)
                FROM ({union_sql}) v
                WHERE LOWER(TRIM(COALESCE(v.activity_status, v.status, ''))) = 'active'
                  AND v.site_id IS NOT NULL
                  AND TRIM(CAST(v.site_id AS TEXT)) <> ''
            ''', params).fetchone()[0]
            return int(val or 0)

        tech_counts: dict[str, int] = {}
        for tech_key in ("2G", "3G", "4G-FDD", "4G-TDD", "5G"):
            ut, pt = _per_tech_union_sql(tech_key)
            tech_counts[tech_key] = _count_active_distinct_sites(ut, pt)

        union_cells, pc = _union_sql_for_map_filter(None)
        total_cells = int(
            execute_query(conn,
                f'''
                SELECT COUNT(*)
                FROM ({union_cells}) v
                WHERE LOWER(TRIM(COALESCE(v.activity_status, v.status, ''))) = 'active'
                ''',
                pc,
            ).fetchone()[0]
            or 0
        )

        # Sites shown on map are those with coordinates and at least one active cell in current snapshot.
        union_sql, params = _union_sql_for_map_filter(None)
        total_sites = execute_query(conn, f'''
            SELECT COUNT(DISTINCT s.site_id)
            FROM ({union_sql}) v
            JOIN sites s ON s.site_id = v.site_id
            WHERE s.latitude IS NOT NULL AND s.longitude IS NOT NULL
              AND LOWER(TRIM(COALESCE(v.activity_status, v.status, ''))) = 'active'
        ''', params).fetchone()[0]

        conn.close()

        repeater_total = 0
        try:
            all_repeaters, _ = load_all_repeaters()
            repeater_total = len(all_repeaters)
        except Exception:
            repeater_total = 0

        return jsonify({'success': True, 'stats': {
            'total_sites':   total_sites,
            'total_sectors': total_cells,
            'total_cells':   total_cells,
            'tech_counts':   tech_counts,
            'total_repeaters': repeater_total,
        }})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@network_map_bp.route('/api/map/search/cell-code', methods=['GET'])
def search_by_cell_code():
    """Search cells by PCI (4G/5G), PSC (3G), or BCCH (2G) using per-tech tables."""
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    code = request.args.get('code', '').strip()
    tech = request.args.get('tech', '').strip()
    tech_value = request.args.get('tech_value', '').strip()

    if not code or not code.lstrip('-').isdigit():
        return jsonify({'success': True, 'matches': []})

    try:
        conn = connect_metadata()
        _sqlite_row(conn)

        union_sql, u_params = _union_sql_for_map_filter(tech if tech else None)
        where_scope = ''
        q_params = u_params + [int(code)]
        if tech_value:
            if _map_request_tech_filter_to_bcch_band_clause(tech):
                where_scope = ' AND CAST(v.pci AS TEXT) = ?'
            else:
                where_scope = ' AND CAST(v.frequency_band AS TEXT) = ?'
            q_params.append(tech_value)
        search_cur = execute_query(conn, f'''
            SELECT
                v.cell_name,
                v.technology,
                v.vendor,
                v.frequency_band,
                v.azimuth,
                v.mechanical_tilt,
                v.electrical_tilt,
                v.pci,
                v.activity_status,
                v.status,
                s.site_id,
                s.site_name,
                COALESCE(s.latitude, v.latitude) AS latitude,
                COALESCE(s.longitude, v.longitude) AS longitude
            FROM ({union_sql}) v
            JOIN sites s ON s.site_id = v.site_id
            WHERE v.pci = ?
              {where_scope}
              AND COALESCE(s.latitude, v.latitude) IS NOT NULL
              AND COALESCE(s.longitude, v.longitude) IS NOT NULL
            ORDER BY s.site_name, v.technology, v.cell_name
        ''', q_params)

        matches = []
        for row in search_cur.fetchall():
            d = dict(row)
            d['cell_id'] = d.get('cell_name')  # map.js expects an id-like field
            matches.append(d)
        conn.close()

        return jsonify({'success': True, 'matches': matches})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@network_map_bp.route('/api/map/export/cell-code', methods=['GET'])
@login_required
def export_cell_code():
    """Export cells matching a PCI / SC / BCCH code as an Excel file."""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import send_file

    code = request.args.get('code', '').strip()
    tech = request.args.get('tech', '').strip()
    tech_value = request.args.get('tech_value', '').strip()

    if not code or not code.lstrip('-').isdigit():
        return jsonify({'error': 'Invalid code'}), 400

    try:
        conn = connect_metadata()
        _sqlite_row(conn)

        union_sql, u_params = _union_sql_for_map_filter(tech if tech else None)
        where_scope = ''
        q_params = u_params + [int(code)]
        if tech_value:
            if _map_request_tech_filter_to_bcch_band_clause(tech):
                where_scope = ' AND CAST(v.pci AS TEXT) = ?'
            else:
                where_scope = ' AND CAST(v.frequency_band AS TEXT) = ?'
            q_params.append(tech_value)
        rows = execute_query(conn, f'''
            SELECT
                v.cell_name,
                s.site_name,
                s.site_id,
                v.technology,
                v.vendor,
                v.frequency_band,
                v.pci,
                v.azimuth,
                v.mechanical_tilt,
                v.electrical_tilt,
                v.activity_status,
                COALESCE(s.latitude, v.latitude) AS latitude,
                COALESCE(s.longitude, v.longitude) AS longitude
            FROM ({union_sql}) v
            JOIN sites s ON s.site_id = v.site_id
            WHERE v.pci = ?
              {where_scope}
              AND COALESCE(s.latitude, v.latitude) IS NOT NULL
              AND COALESCE(s.longitude, v.longitude) IS NOT NULL
            ORDER BY s.site_name, v.technology, v.cell_name
        ''', q_params).fetchall()
        conn.close()

        code_label = (
            'PSC'
            if tech in ('3G', '3G-3G')
            else ('BCCH' if tech in ('2G', '2G-2G') else 'PCI')
        )

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = f'{code_label}_{code}'

        headers = ['Cell Name', 'Site Name', 'Site ID', 'Technology', 'Vendor',
                   'Band', code_label, 'Azimuth (°)', 'M.Tilt (°)', 'E.Tilt (°)',
                   'Activity status', 'Latitude', 'Longitude']
        ws.append(headers)

        hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        for cell in ws[1]:
            cell.font      = Font(bold=True, color='FFFFFF')
            cell.fill      = hdr_fill
            cell.alignment = Alignment(horizontal='center')

        for row in rows:
            ws.append([
                row['cell_name'], row['site_name'], row['site_id'],
                row['technology'], row['vendor'], row['frequency_band'],
                row['pci'], row['azimuth'], row['mechanical_tilt'],
                row['electrical_tilt'], row['activity_status'],
                row['latitude'], row['longitude']
            ])

        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        uid = (request.current_user.get('id')
               if isinstance(request.current_user, dict)
               else request.current_user[0])
        log_activity(uid, 'export', f'Exported {len(rows)} cells with {code_label}={code}')

        return send_file(
            buf,
            download_name=f'{code_label}_{code}_cells.xlsx',
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@network_map_bp.route('/api/map/export/sites', methods=['GET'])
@login_required
def export_sites_excel():
    """Export all sites + cells for the current filter as an Excel file (two sheets)."""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import send_file
    from datetime import datetime as dt

    tech       = request.args.get('tech',       '').strip()
    tech_value = request.args.get('tech_value', '').strip()
    vendor     = request.args.get('vendor',     '').strip()
    search     = request.args.get('search',     '').strip()

    try:
        conn = connect_metadata()
        _sqlite_row(conn)

        # ── Sites ─────────────────────────────────────────────────────────────
        s_cond   = ['s.latitude IS NOT NULL', 's.longitude IS NOT NULL']
        s_params = []
        if vendor:
            s_cond.append('s.vendor = ?');  s_params.append(vendor)
        if search:
            s_cond.append('(s.site_name LIKE ? OR CAST(s.site_id AS TEXT) LIKE ?)')
            s_params += [f'%{search}%', f'%{search}%']
        if tech_value:
            if _map_request_tech_filter_to_bcch_band_clause(tech):
                s_cond.append('CAST(v.pci AS TEXT) = ?')
            else:
                s_cond.append('CAST(v.frequency_band AS TEXT) = ?')
            s_params.append(tech_value)

        union_sql, u_params = _union_sql_for_map_filter(tech if tech else None)
        # Sites are those matching the filter and having at least one cell row.
        sites = execute_query(conn, f'''
            SELECT DISTINCT
                s.site_id, s.site_name, s.latitude, s.longitude,
                s.region, s.site_type, s.vendor, s.status
            FROM ({union_sql}) v
            JOIN sites s ON s.site_id = v.site_id
            WHERE {" AND ".join(s_cond)}
            ORDER BY s.site_name
        ''', u_params + s_params).fetchall()

        # ── Cells ─────────────────────────────────────────────────────────────
        c_cond, c_params = [], []
        if vendor:
            c_cond.append('s.vendor = ?');      c_params.append(vendor)
        if search:
            c_cond.append('(s.site_name LIKE ? OR CAST(s.site_id AS TEXT) LIKE ?)')
            c_params += [f'%{search}%', f'%{search}%']
        if tech_value:
            if _map_request_tech_filter_to_bcch_band_clause(tech):
                c_cond.append('CAST(v.pci AS TEXT) = ?')
            else:
                c_cond.append('CAST(v.frequency_band AS TEXT) = ?')
            c_params.append(tech_value)
        c_where = ('WHERE ' + ' AND '.join(c_cond)) if c_cond else ''
        union_sql, u_params = _union_sql_for_map_filter(tech if tech else None)
        cells = execute_query(conn, f'''
            SELECT
                v.cell_name,
                v.site_id,
                s.site_name,
                v.technology,
                v.frequency_band,
                v.azimuth,
                v.mechanical_tilt,
                v.electrical_tilt,
                v.pci,
                v.activity_status,
                v.status,
                v.vendor,
                s.latitude,
                s.longitude
            FROM ({union_sql}) v
            JOIN sites s ON s.site_id = v.site_id
            {c_where}
            ORDER BY s.site_name, v.technology, v.cell_name
        ''', u_params + c_params).fetchall()
        conn.close()

        # ── Build workbook ────────────────────────────────────────────────────
        HDR_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        HDR_FONT = Font(bold=True, color='FFFFFF')
        CTR      = Alignment(horizontal='center')

        def _style_header(ws):
            for cell in ws[1]:
                cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.alignment = CTR

        def _autofit(ws):
            for col in ws.columns:
                w = max((len(str(c.value or '')) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(w + 4, 45)

        wb   = openpyxl.Workbook()
        ws_s = wb.active
        ws_s.title = 'Sites'
        ws_s.append(['Site ID', 'Site Name', 'Latitude', 'Longitude',
                     'Region', 'Site Type', 'Vendor', 'Status'])
        _style_header(ws_s)
        for r in sites:
            ws_s.append([r['site_id'], r['site_name'], r['latitude'], r['longitude'],
                         r['region'], r['site_type'], r['vendor'], r['status']])
        _autofit(ws_s)

        ws_c = wb.create_sheet('Cells')
        ws_c.append(['Cell Name', 'Site ID', 'Site Name', 'Technology', 'Frequency Band',
                     'Azimuth (°)', 'Mech. Tilt (°)', 'Elec. Tilt (°)',
                     'PCI / SC / BCCH', 'Activity status', 'Vendor', 'Latitude', 'Longitude'])
        _style_header(ws_c)
        for r in cells:
            ws_c.append([r['cell_name'], r['site_id'], r['site_name'], r['technology'],
                         r['frequency_band'], r['azimuth'], r['mechanical_tilt'],
                         r['electrical_tilt'], r['pci'],
                         r['activity_status'] or r['status'], r['vendor'],
                         r['latitude'], r['longitude']])
        _autofit(ws_c)

        buf = BytesIO(); wb.save(buf); buf.seek(0)
        tech_label = (tech or 'All').replace('-', '_')
        fname = f'network_export_{tech_label}_{dt.now().strftime("%Y%m%d_%H%M")}.xlsx'
        uid = (request.current_user.get('id')
               if isinstance(request.current_user, dict) else request.current_user[0])
        log_activity(uid, 'export',
                     f'Excel export: tech={tech or "All"}, {len(sites)} sites, {len(cells)} cells')
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@network_map_bp.route('/api/map/export/kml', methods=['GET'])
@login_required
def export_sites_kml():
    """Export sites as a KML file for Google Earth."""
    from flask import Response
    from datetime import datetime as dt
    import xml.sax.saxutils as sx

    tech       = request.args.get('tech',       '').strip()
    tech_value = request.args.get('tech_value', '').strip()
    vendor     = request.args.get('vendor',     '').strip()
    search     = request.args.get('search',     '').strip()

    try:
        conn = connect_metadata()
        _sqlite_row(conn)

        s_cond   = ['s.latitude IS NOT NULL', 's.longitude IS NOT NULL']
        s_params = []
        if vendor:
            s_cond.append('s.vendor = ?');  s_params.append(vendor)
        if search:
            s_cond.append('(s.site_name LIKE ? OR CAST(s.site_id AS TEXT) LIKE ?)')
            s_params += [f'%{search}%', f'%{search}%']
        if tech_value:
            if _map_request_tech_filter_to_bcch_band_clause(tech):
                s_cond.append('CAST(v.pci AS TEXT) = ?')
            else:
                s_cond.append('CAST(v.frequency_band AS TEXT) = ?')
            s_params.append(tech_value)
        union_sql, u_params = _union_sql_for_map_filter(tech if tech else None)
        sites = execute_query(conn, f'''
            SELECT DISTINCT
                s.site_id, s.site_name, s.latitude, s.longitude,
                s.region, s.vendor, s.status
            FROM ({union_sql}) v
            JOIN sites s ON s.site_id = v.site_id
            WHERE {" AND ".join(s_cond)}
            ORDER BY s.site_name
        ''', u_params + s_params).fetchall()

        # Fetch cells for each site to populate the description balloon
        cells_by_site = {}
        if sites:
            ids = [r['site_id'] for r in sites]
            ph  = ','.join('?' * len(ids))
            union_sql, u_params = _union_sql_for_map_filter(tech if tech else None)
            for r in execute_query(conn, f'''
                SELECT
                    v.site_id,
                    v.cell_name,
                    v.technology,
                    v.azimuth,
                    v.frequency_band,
                    v.pci,
                    v.activity_status,
                    v.status
                FROM ({union_sql}) v
                WHERE v.site_id IN ({ph})
                  {("AND CAST(v.pci AS TEXT) = ?" if tech_value and _map_request_tech_filter_to_bcch_band_clause(tech) else "")}
                  {("AND CAST(v.frequency_band AS TEXT) = ?" if tech_value and not _map_request_tech_filter_to_bcch_band_clause(tech) else "")}
                ORDER BY v.technology, v.cell_name
            ''', u_params + ids + ([tech_value] if tech_value else [])).fetchall():
                cells_by_site.setdefault(r['site_id'], []).append(r)
        conn.close()

        # ── Build KML (site = donut polygon; active cells = sector wedges) ───
        tech_label = tech or 'All Technologies'
        lines = [
            '<?xml version="1.0" encoding="UTF-8"?>',
            '<kml xmlns="http://www.opengis.net/kml/2.2">',
            '<Document>',
            f'<name>{sx.escape(tech_label)} Network Export</name>',
            f'<description>Exported {dt.now().strftime("%Y-%m-%d %H:%M")} — sites as rings, active sectors as wedges.</description>',
            '<Style id="site-donut">',
            '  <LineStyle><color>FF333333</color><width>2</width></LineStyle>',
            f'  <PolyStyle><color>{_kml_color_aabbggrr("#ffffff", 140)}</color><outline>1</outline></PolyStyle>',
            '</Style>',
        ]
        for tname, hx in _KML_TECH_COLORS.items():
            sid = f'wedge-{tname.replace(" ", "_")}'
            fill_a = int(0.35 * 255)
            lines += [
                f'<Style id="{sid}">',
                f'  <LineStyle><color>{_kml_color_aabbggrr(hx, 255)}</color><width>2</width></LineStyle>',
                f'  <PolyStyle><color>{_kml_color_aabbggrr(hx, fill_a)}</color><outline>1</outline></PolyStyle>',
                '</Style>',
            ]
        lines += [
            '<Style id="wedge-default">',
            f'  <LineStyle><color>{_kml_color_aabbggrr("#34495e", 255)}</color><width>2</width></LineStyle>',
            f'  <PolyStyle><color>{_kml_color_aabbggrr("#34495e", int(0.35 * 255))}</color><outline>1</outline></PolyStyle>',
            '</Style>',
        ]

        lines += ['<Folder>', f'<name>{sx.escape("Sites")}</name>']
        wedge_count = 0
        for site in sites:
            site_cells = cells_by_site.get(site['site_id'], [])
            rows_html = ''.join(
                f'<tr><td>{sx.escape(c["cell_name"])}</td><td>{sx.escape(str(c["technology"] or ""))}</td>'
                f'<td>{c["azimuth"] if c["azimuth"] is not None else "—"}°</td>'
                f'<td>{sx.escape(str(c["frequency_band"] or "—"))}</td>'
                f'<td>{sx.escape(str((c["activity_status"] or c["status"]) or "—"))}</td></tr>'
                for c in site_cells
            )
            table = (
                '<table border="1" cellpadding="3">'
                '<tr><th>Cell</th><th>Tech</th><th>Azimuth</th><th>Band</th><th>Activity</th></tr>'
                f'{rows_html}</table>'
            ) if rows_html else ''
            desc = (
                f'<b>Site ID:</b> {sx.escape(str(site["site_id"]))}<br/>'
                f'<b>Vendor:</b> {sx.escape(site["vendor"] or "")}<br/>'
                f'<b>Region:</b> {sx.escape(site["region"] or "")}<br/>'
                f'<b>Status:</b> {sx.escape(site["status"] or "")}<br/><br/>{table}'
            )
            try:
                slat = float(site['latitude'])
                slng = float(site['longitude'])
            except (TypeError, ValueError):
                continue
            outer = _kml_circle_ring_coords(slat, slng, 24.0, 36, reverse=False)
            inner = _kml_circle_ring_coords(slat, slng, 12.0, 36, reverse=True)
            lines += [
                '<Placemark>',
                f'<name>{sx.escape(site["site_name"])}</name>',
                f'<description><![CDATA[{desc}]]></description>',
                '<styleUrl>#site-donut</styleUrl>',
                '<Polygon>',
                '  <extrude>0</extrude><altitudeMode>clampToGround</altitudeMode>',
                '  <outerBoundaryIs><LinearRing>',
                f'    <coordinates>{_kml_ring_to_coordinates(outer)}</coordinates>',
                '  </LinearRing></outerBoundaryIs>',
                '  <innerBoundaryIs><LinearRing>',
                f'    <coordinates>{_kml_ring_to_coordinates(inner)}</coordinates>',
                '  </LinearRing></innerBoundaryIs>',
                '</Polygon>',
                '</Placemark>',
            ]
        lines += ['</Folder>', '<Folder>', f'<name>{sx.escape("Sectors (wedges)")}</name>']

        for site in sites:
            site_cells = cells_by_site.get(site['site_id'], [])
            if not site_cells:
                continue
            try:
                slat = float(site['latitude'])
                slng = float(site['longitude'])
            except (TypeError, ValueError):
                continue
            wedges = _kml_group_cells_into_wedges(site_cells)
            for (wtech, bucket), group in sorted(wedges.items(), key=lambda x: (x[0][0], x[0][1])):
                ring = _kml_wedge_ring_coords(slat, slng, float(bucket))
                if len(ring) < 4:
                    continue
                sid = f'wedge-{wtech.replace(" ", "_")}'
                if wtech not in _KML_TECH_COLORS:
                    sid = 'wedge-default'

                names = ', '.join(_kml_cdata_fragment(str(c['cell_name'])) for c in group[:6])
                if len(group) > 6:
                    names += f' (+{len(group) - 6} more)'
                wdesc = (
                    f'<b>Site:</b> {_kml_cdata_fragment(site["site_name"])} (ID {_kml_cdata_fragment(str(site["site_id"]))})<br/>'
                    f'<b>Technology:</b> {_kml_cdata_fragment(wtech)}<br/>'
                    f'<b>Azimuth (bin centre):</b> {bucket}°<br/>'
                    f'<b>Cells:</b> {names}'
                )
                lines += [
                    '<Placemark>',
                    f'<name>{sx.escape(site["site_name"])} — {wtech} {int(bucket)}°</name>',
                    f'<description><![CDATA[{wdesc}]]></description>',
                    f'<styleUrl>#{sid}</styleUrl>',
                    '<Polygon>',
                    '  <extrude>0</extrude><altitudeMode>clampToGround</altitudeMode>',
                    '  <outerBoundaryIs><LinearRing>',
                    f'    <coordinates>{_kml_ring_to_coordinates(ring)}</coordinates>',
                    '  </LinearRing></outerBoundaryIs>',
                    '</Polygon>',
                    '</Placemark>',
                ]
                wedge_count += 1

        lines += ['</Folder>', '</Document>', '</kml>']

        tech_fn = (tech or 'All').replace('-', '_')
        fname   = f'network_export_{tech_fn}_{dt.now().strftime("%Y%m%d_%H%M")}.kml'
        uid = (request.current_user.get('id')
               if isinstance(request.current_user, dict) else request.current_user[0])
        log_activity(uid, 'export',
                     f'KML export: tech={tech or "All"}, {len(sites)} sites, {wedge_count} sector wedges')
        return Response('\n'.join(lines),
                        mimetype='application/vnd.google-earth.kml+xml',
                        headers={'Content-Disposition': f'attachment; filename="{fname}"'})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@network_map_bp.route('/api/network-map/neighbors/lines', methods=['GET'])
def get_neighbor_lines():
    """Neighbor lines; ``vendor`` scopes the **source** for raw export linking (targets may be any vendor).

    Query ``cell_name`` (normalized) matches the **source** cell only: outgoing handovers from that cell.

    ``failures_only=1`` (or ``true``): return only links with estimated failures meeting
    ``min_failures`` (default **1.0**): failures = ``attempts × (1 − SR/100)`` (or ``attempts − successes``);
    requires SR or success counts. ``min_attempts`` is ignored for filtering in this mode (only attempts ≥ 1).
    """
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    vendor = (request.args.get('vendor') or '').strip()
    technology = _normalize_ui_map_tech_token((request.args.get('technology') or '').strip())
    site_id = (request.args.get('site_id') or '').strip()
    cell_name = (request.args.get('cell_name') or '').strip()
    min_attempts = max(0, _safe_int(request.args.get('min_attempts'), 10))
    max_lines = min(2000, max(10, _safe_int(request.args.get('max_lines'), 300)))
    failures_only = (request.args.get('failures_only') or '').strip().lower() in (
        '1', 'true', 'yes', 'on',
    )
    min_failures = max(0.0, _safe_float(request.args.get('min_failures'), 1.0))

    if not technology:
        return jsonify({'error': 'technology is required'}), 400

    try:
        nconn = sqlite3.connect(NEIGHBOR_KPI_DB, timeout=30)
        nconn.row_factory = sqlite3.Row
        _ensure_neighbor_schema(nconn)

        if not _neighbor_table_exists(nconn, "neighbor_hourly"):
            v_req = (vendor or "").strip()
            v_low = v_req.lower()
            # (SQLite path, raw table, metadata source-vendor label for linking)
            jobs: list[tuple[str, str, str]] = []
            if not v_req or v_low == "all":
                for nrt in _resolve_raw_neighbor_tables_for_vendor(nconn, technology, "nokia"):
                    jobs.append((NEIGHBOR_KPI_DB, nrt, "Nokia"))
                if os.path.isfile(HUAWEI_NEIGHBOR_RAW_DB):
                    htmp = sqlite3.connect(HUAWEI_NEIGHBOR_RAW_DB, timeout=30)
                    htmp.row_factory = sqlite3.Row
                    try:
                        hrt = _resolve_huawei_neighbor_export_table(htmp, technology)
                        if hrt and _neighbor_table_non_empty(htmp, hrt):
                            jobs.append((HUAWEI_NEIGHBOR_RAW_DB, hrt, "Huawei"))
                    finally:
                        htmp.close()
            else:
                if v_low == "huawei":
                    if os.path.isfile(HUAWEI_NEIGHBOR_RAW_DB):
                        htmp = sqlite3.connect(HUAWEI_NEIGHBOR_RAW_DB, timeout=30)
                        htmp.row_factory = sqlite3.Row
                        try:
                            hrt = _resolve_huawei_neighbor_export_table(htmp, technology)
                            if hrt and _neighbor_table_non_empty(htmp, hrt):
                                jobs.append((HUAWEI_NEIGHBOR_RAW_DB, hrt, v_req))
                        finally:
                            htmp.close()
                else:
                    for nrt in _resolve_raw_neighbor_tables_for_vendor(nconn, technology, "nokia"):
                        jobs.append((NEIGHBOR_KPI_DB, nrt, v_req))

            if jobs:
                cell_norm = _norm_cell_key(cell_name) if cell_name else ""
                all_lines: list[dict] = []
                skipped = 0
                total = 0
                periods: list[str] = []
                messages: list[str] = []
                for db_path, raw_tbl, v_label in jobs:
                    reuse = db_path == NEIGHBOR_KPI_DB
                    wconn = nconn if reuse else sqlite3.connect(db_path, timeout=30)
                    if not reuse:
                        wconn.row_factory = sqlite3.Row
                    try:
                        lines, sk, tot, period, raw_msg = build_raw_neighbor_lines(
                            neighbor_conn=wconn,
                            raw_table=raw_tbl,
                            technology=technology,
                            vendor=v_label,
                            cell_norm=cell_norm,
                            site_id_filter=site_id,
                            min_attempts=float(min_attempts),
                            max_lines=max_lines,
                            failures_only=failures_only,
                            min_failures=float(min_failures),
                        )
                        all_lines.extend(lines)
                        skipped += sk
                        total += tot
                        if period:
                            periods.append(str(period))
                        if raw_msg:
                            messages.append(raw_msg)
                    finally:
                        if not reuse:
                            wconn.close()

                nconn.close()
                all_lines.sort(key=lambda x: float(x.get("ho_attempts") or 0), reverse=True)
                all_lines = all_lines[:max_lines]
                period_out = max(periods) if periods else None
                payload: dict = {
                    "success": True,
                    "period_start": period_out,
                    "lines": all_lines,
                    "skipped_missing_coords": skipped,
                    "total_candidates": total,
                    "raw_neighbor_tables": True,
                }
                if messages:
                    payload["message"] = "; ".join(dict.fromkeys(messages))
                return jsonify(payload)

            nconn.close()
            return jsonify({
                "success": True,
                "period_start": None,
                "lines": [],
                "skipped_missing_coords": 0,
                "total_candidates": 0,
            })

        cell_norm = _norm_cell_key(cell_name) if cell_name else ''
        tech_vals = _neighbor_hourly_tech_aliases(technology)
        if not tech_vals:
            nconn.close()
            return jsonify({'error': 'technology is required'}), 400
        tech_ph = ",".join("?" for _ in tech_vals)
        where = [f"technology IN ({tech_ph})"]
        params: list[object] = [*tech_vals]
        if failures_only:
            where.append("COALESCE(ho_attempts, 0) >= 1")
        else:
            where.append("COALESCE(ho_attempts, 0) >= ?")
            params.append(min_attempts)
        if vendor and vendor.lower() != 'all':
            where.insert(0, "vendor = ?")
            params.insert(0, vendor)
        if cell_norm:
            where.append("source_cell_norm = ?")
            params.append(cell_norm)

        max_period = execute_query(nconn,
            f'SELECT MAX(period_start) AS p FROM neighbor_hourly WHERE {" AND ".join(where)}',
            params,
        ).fetchone()
        period = max_period["p"] if max_period else None
        if not period:
            nconn.close()
            return jsonify({
                'success': True,
                'period_start': None,
                'lines': [],
                'skipped_missing_coords': 0,
                'total_candidates': 0,
            })

        rows = execute_query(nconn,
            f"""
            SELECT *
            FROM neighbor_hourly
            WHERE {" AND ".join(where)} AND period_start = ?
            ORDER BY COALESCE(ho_attempts, 0) DESC
            LIMIT ?
            """,
            params + [period, max_lines * 4],
        ).fetchall()
        nconn.close()

        coords = _load_neighbor_coords(vendor, technology)
        lines = []
        skipped_missing = 0
        for r in rows:
            src = coords.get(r["source_cell_norm"] or _norm_cell_key(r["source_cell"]))
            dst = coords.get(r["target_cell_norm"] or _norm_cell_key(r["target_cell"]))
            if not src or not dst:
                skipped_missing += 1
                continue
            if site_id:
                if str(src.get("site_id")) != site_id and str(dst.get("site_id")) != site_id:
                    continue
            attempts_val = float(r["ho_attempts"] or 0)
            succ_raw = r["ho_successes"]
            try:
                succ_f = float(succ_raw) if succ_raw is not None else None
            except (TypeError, ValueError):
                succ_f = None
            rate = r["ho_success_rate"]
            if rate is None:
                try:
                    if attempts_val > 0 and succ_f is not None:
                        rate = (succ_f / attempts_val) * 100.0
                except (TypeError, ValueError, ZeroDivisionError):
                    rate = None
            rate_f = float(rate) if rate is not None else None
            failures = neighbor_ho_failures(attempts_val, rate_f, succ_f)
            failures_int = int(math.trunc(failures)) if failures is not None else None
            if failures_only:
                thr = min_failures if min_failures > 1e-12 else 1e-9
                if failures_int is None or failures_int < thr - 1e-12:
                    continue
            fr_pct = (failures_int / attempts_val * 100.0) if failures_int is not None and attempts_val > 0 else None
            src_site = src.get("site_id")
            dst_site = dst.get("site_id")
            is_intra = bool(str(src_site or "").strip() and str(src_site or "").strip() == str(dst_site or "").strip())
            lines.append({
                "period_start": r["period_start"],
                "vendor": r["vendor"],
                "technology": r["technology"],
                "source_cell": r["source_cell"],
                "target_cell": r["target_cell"],
                "source_site_id": src_site,
                "target_site_id": dst_site,
                "is_intra_relation": is_intra,
                "relation_scope": "intra" if is_intra else "inter",
                "source_lat": src["lat"],
                "source_lng": src["lng"],
                "target_lat": dst["lat"],
                "target_lng": dst["lng"],
                "ho_attempts": r["ho_attempts"],
                "ho_successes": r["ho_successes"],
                "ho_success_rate": rate,
                "ho_failures": failures_int,
                "ho_failure_rate_percent": fr_pct,
            })
            if len(lines) >= max_lines:
                break

        return jsonify({
            "success": True,
            "period_start": period,
            "lines": lines,
            "skipped_missing_coords": skipped_missing,
            "total_candidates": len(rows),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@network_map_bp.route('/api/network-map/neighbors/cell-summary', methods=['GET'])
def get_neighbor_cell_summary():
    """Top incoming/outgoing neighbors for a selected cell in latest hour."""
    user = get_current_user()
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401

    vendor = (request.args.get('vendor') or '').strip()
    technology = _normalize_ui_map_tech_token((request.args.get('technology') or '').strip())
    cell_name = (request.args.get('cell_name') or '').strip()
    top_n = min(50, max(1, _safe_int(request.args.get('top_n'), 10)))
    min_attempts = max(0, _safe_int(request.args.get('min_attempts'), 0))
    if not technology or not cell_name:
        return jsonify({'error': 'technology and cell_name are required'}), 400
    cell_norm = _norm_cell_key(cell_name)
    try:
        conn = sqlite3.connect(NEIGHBOR_KPI_DB, timeout=30)
        conn.row_factory = sqlite3.Row
        _ensure_neighbor_schema(conn)
        if not _neighbor_table_exists(conn, "neighbor_hourly"):
            if _any_raw_neighbor_table_exists(conn, technology):
                conn.close()
                return jsonify({
                    "success": True,
                    "period_start": None,
                    "cell_name": cell_name,
                    "outgoing": [],
                    "incoming": [],
                    "raw_neighbor_tables": True,
                })
            conn.close()
            return jsonify({
                "success": True,
                "period_start": None,
                "cell_name": cell_name,
                "outgoing": [],
                "incoming": [],
            })
        vendor_clause = ""
        vendor_params: list[object] = []
        if vendor and vendor.lower() != 'all':
            vendor_clause = "AND vendor = ?"
            vendor_params = [vendor]
        tech_vals = _neighbor_hourly_tech_aliases(technology)
        if not tech_vals:
            conn.close()
            return jsonify({'error': 'technology and cell_name are required'}), 400
        tech_ph = ",".join("?" for _ in tech_vals)

        period_row = execute_query(conn,
            f"""
            SELECT MAX(period_start) AS p
            FROM neighbor_hourly
            WHERE technology IN ({tech_ph})
              {vendor_clause}
              AND (source_cell_norm = ? OR target_cell_norm = ?)
              AND COALESCE(ho_attempts, 0) >= ?
            """,
            [*tech_vals] + vendor_params + [cell_norm, cell_norm, min_attempts],
        ).fetchone()
        period = period_row["p"] if period_row else None
        if not period:
            conn.close()
            return jsonify({"success": True, "period_start": None, "outgoing": [], "incoming": []})

        outgoing = execute_query(conn,
            f"""
            SELECT target_cell AS neighbor_cell,
                   SUM(COALESCE(ho_attempts,0)) AS ho_attempts,
                   SUM(COALESCE(ho_successes,0)) AS ho_successes,
                   CASE WHEN SUM(COALESCE(ho_attempts,0)) > 0
                        THEN (SUM(COALESCE(ho_successes,0))*100.0) / SUM(COALESCE(ho_attempts,0))
                        ELSE NULL END AS ho_success_rate
            FROM neighbor_hourly
            WHERE technology IN ({tech_ph}) AND period_start = ?
              {vendor_clause}
              AND source_cell_norm = ? AND COALESCE(ho_attempts,0) >= ?
            GROUP BY target_cell_norm, target_cell
            ORDER BY ho_attempts DESC
            LIMIT ?
            """,
            [*tech_vals, period] + vendor_params + [cell_norm, min_attempts, top_n],
        ).fetchall()
        incoming = execute_query(conn,
            f"""
            SELECT source_cell AS neighbor_cell,
                   SUM(COALESCE(ho_attempts,0)) AS ho_attempts,
                   SUM(COALESCE(ho_successes,0)) AS ho_successes,
                   CASE WHEN SUM(COALESCE(ho_attempts,0)) > 0
                        THEN (SUM(COALESCE(ho_successes,0))*100.0) / SUM(COALESCE(ho_attempts,0))
                        ELSE NULL END AS ho_success_rate
            FROM neighbor_hourly
            WHERE technology IN ({tech_ph}) AND period_start = ?
              {vendor_clause}
              AND target_cell_norm = ? AND COALESCE(ho_attempts,0) >= ?
            GROUP BY source_cell_norm, source_cell
            ORDER BY ho_attempts DESC
            LIMIT ?
            """,
            [*tech_vals, period] + vendor_params + [cell_norm, min_attempts, top_n],
        ).fetchall()
        conn.close()

        return jsonify({
            "success": True,
            "period_start": period,
            "cell_name": cell_name,
            "outgoing": [dict(r) for r in outgoing],
            "incoming": [dict(r) for r in incoming],
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@network_map_bp.route('/api/map/refresh', methods=['POST'])
@login_required
def refresh_metadata():
    """Trigger a metadata sync in the background and return immediately."""
    import threading
    try:
        from modules.sync.scheduler import trigger_metadata_now
        t = threading.Thread(target=trigger_metadata_now, daemon=True)
        t.start()
        uid = (request.current_user.get('id')
               if isinstance(request.current_user, dict)
               else request.current_user[0])
        log_activity(uid, 'metadata_refresh', 'Triggered metadata refresh from network map')
        return jsonify({'success': True, 'message': 'Metadata sync started'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
